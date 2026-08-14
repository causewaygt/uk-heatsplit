#!/usr/bin/env python3
"""
Irish Heat Split - daily build pipeline. PIPELINE_VERSION 1.0.0.

Changelog:
  1.0.0 - launch. Geothermal register + WGC2026 ROI anchors + per-capita
          what-if derivation with European reference points; per-
          jurisdiction hero; climate-modelled ASHP SPF; cool-side panel;
          events, flags and soft-feed registers.
  0.x   - see below.
  0.5.0 - ONS GB heating-oil feed (kj5u - same-tax control for the NI-GB
          market-structure gap); EVENTS register rendered as chart
          annotations; FEED_FLAGS register for value-level caveats distinct
          from fetch status; tariff anchors re-based on the July 2026
          sourced pass (Power NI/UR review, ROI standard rates);
          derive_heat_gap() - cost of useful heat by route per jurisdiction
          with break-even SPF vs the incumbent oil boiler.
  0.4.0 - oil bulletin fetches with- AND without-taxes files; ex-tax series.
  0.3.0 - gni_live implemented against the probed gasconsumption JSON API.
  0.2.0 - ANCHORS + derive_hero() weekly four-stat and geothermal what-if.
  0.1.x - scaffold, feed fixes from first-run logs.

House rules: self-resolve IDs/links at runtime; dump available names on
failure; every feed try/except with previous values retained and status
"stale"; fetch health vs data recency tracked separately; unit
autodetection; clip future-dated rows; en dashes in user-facing strings.
"""

import datetime as dt
import html as html_mod
import io
import json
import math
import random
import re
import statistics
import sys
import time
import traceback
import urllib.parse
from pathlib import Path
from xml.etree import ElementTree

import requests

# ---------------------------------------------------------------- constants

# VERSIONING, and a temporary rule. Normally x.y.z means x a new
# source or panel, y a source update, z wording or format. While the
# site is under construction the x is FROZEN at 5 and only y and z
# move - the panels are changing weekly and an x that tracked every
# new one would say nothing. The "Under Construction" label on the
# masthead and this freeze come off together.
PIPELINE_VERSION = "5.6.0"
ROOT = Path(__file__).resolve().parents[1]
DATA_PATH = ROOT / "docs" / "data.json"
# The hourly store lives in its OWN file: a malformed hourly write can
# never corrupt the weekly tracker's data, the grid panel can be absent
# while everything else renders, and the store versions independently
# of HISTORY_SCHEMA / ANCHOR_EPOCH. Fetched by the page only when the
# grid panel needs it.
HOURLY_PATH = ROOT / "docs" / "hourly.json"
# 2 (7 Aug 2026): temp_ai added - population-weighted island air
# temperature per hour, from Open-Meteo, on the SAME local clock as
# the EirGrid series. Without it the grid layer cannot be computed at
# all: hourly heat needs an hourly temperature, and the only hourly
# temperature this pipeline previously touched was a trailing 60 days
# reduced to daily ODH26 before anything was persisted. The store
# holds temperature rather than degree hours so that hourly HDD (base
# 15.5), ODH26 and the Carnot source temperature all derive from one
# retained series instead of three.
# 3 (7 Aug 2026): the store is written as flat arrays against a base
# hour instead of one key per value. At schema 2 the file reached
# 1,025 kB rewritten daily, and the repeated 13-character keys were
# most of it. Readers must tolerate BOTH shapes - expand_hourly()
# accepts a schema 1/2 document so the first run after this change
# inherits its own history rather than refilling from empty.
# 4 (8 Aug 2026): price_ai added - SEMOpx day-ahead, EUR/MWh, for
# B.2.3 (the coincidence premium). Fills FORWARD only; see
# feed_semopx and semopx_history_probe for why.
HOURLY_SCHEMA = 4
HOURLY_MONTHS = 13          # rolling window the store keeps
HOURLY_CHUNKS_PER_RUN = 16  # walk-back budget; 14 fills 13 months
# Open-Meteo serves long hourly spans in one request, so the
# temperature series walks in far bigger chunks than EirGrid's
# ~28-day window; 4 x 120 days covers 13 months with room over.
HOURLY_TEMP_CHUNK_DAYS = 120
HOURLY_TEMP_CHUNKS_PER_RUN = 5
# Raised 400 -> 1150 (27 Jul 2026) so the back-look can reach the
# tariff-confidence floor with a full trailing HDD year beneath its
# earliest week.
SERIES_KEEP_DAYS = 1150
# Earliest week the back-look will price: all four tariff anchors are
# verified from this date (Power NI +4% and Firmus -7.86% both
# effective 1 Oct 2025; Electric Ireland electricity frozen through
# the winter, gas -4% from Sep 2025).
# 8 Aug 2026: extended BACK eight weeks to complete a year on record.
# Why this date is reachable and 1 Oct 2025 was the previous floor:
#  - ROI needs no new anchor at all. Electric Ireland held prices from
#    Oct 2022 to 1 Jul 2026, and SEAI/Eurostat semester S2 2025 runs
#    1 Jul - 31 Dec 2025, so Aug-Sep 2025 is inside the same period
#    the Oct 2025 row already describes.
#  - NI needs one prior row, and both regulated suppliers publish it:
#    SSE's maximum average price was set on 1 Apr 2024 and unchanged
#    until 1 Oct 2025; Power NI's tariff ran from 1 Dec 2024.
#  - Grid carbon for these weeks is NOT in the daily EirGrid feed
#    (50-day retention). It is taken from the hourly store, whose
#    floor advances daily - see daily_ci_from_hourly().
HISTORY_START = "2025-08-06"
# The date from which the record has been observed as it happened.
# Weeks before it are reconstructed from published tariffs and the
# hourly store, and are counted and labelled separately - the 52-live
# -weeks milestone is not reached by backfilling.
LIVE_FROM = "2025-10-01"
# Weeks retained. Was 60, which a 52-week record plus forward growth
# would have hit inside two months - silently dropping the weeks
# backfilled here. 120 leaves room for the 24-month view.
HISTORY_MAX = 120
# History schema (splits handover, 1 Aug 2026). Bump triggers a
# one-time restatement: stored weeks recompute through derive_hero
# with their STORED per-entry inputs (ef_electricity injected, oil
# and fx from the retained series) - existing fields must re-round
# identically, only new fields appear. Unlike the UK, the Irish
# recompute window is the price feeds at 1,150-day retention, so all
# weeks are restatable and stranding risk is ~nil - measured, not
# assumed, by the drift check below.
# 4 (7 Aug 2026): forces a re-run of the schema-3 migration, which
# used setdefault() for the ni/roi blocks and therefore left frozen
# weeks with jurisdiction blocks predating the per-fuel addition -
# so the windowed energy bars worked all-island but fell back to the
# live week for NI and ROI.
# 5 (7 Aug 2026): the per-jurisdiction sub-blocks gain the heat/cold
# splits the island entry has carried since schema 2. Without them
# every windowed view under the NI or ROI toggle failed the front
# end's split guard, so the 4w/12w/12m cards and what-if silently
# dropped their "(heat ... cooling ... saves ...)" breakdowns while
# all-island showed them - a jurisdiction-only gap that looked like a
# copy bug and was a missing field.
HISTORY_SCHEMA = 5
# Anchor epoch. Bump whenever a change to ANCHORS alters what a past
# week WOULD have computed - as distinct from HISTORY_SCHEMA, which
# tracks new FIELDS. A schema bump adds fields and leaves stored
# values alone; an epoch bump rewrites every recomputable week onto
# the new basis, using that week's own stored prices, tariffs, fx and
# emission factor. Without it a basis change leaves the series half on
# one footing and half on another, and the heat/cold splits stop
# reconciling with the totals they were derived from.
#   1 - launch basis
#   2 - 6 Aug 2026: data-centre line repriced to its cooling share
#       (SEAI NHS R1) and hot water re-anchored 18.3% -> 22.4%
#   3 - 7 Aug 2026: cooling service factors re-anchored to the SEAI
#       National Heat Study supporting data (useful-to-final ratios)
#   4 - 7 Aug 2026: heat-pump stock split out of the electricity line
#       so the bars can show heat-pump electricity and the ambient heat
#       it harvests, on census anchors
# 5 (8 Aug 2026): NI tariff basis rebuilt (see TARIFF_HISTORY) and
# the back-look floor moved to 2025-08-06.
# 6 (8 Aug 2026): ROI domestic moved onto the same all-in basis as NI
# - Eurostat band prices, standing charges included.
# 7 (8 Aug 2026): Irish anchors moved to semester 2 2025 (credit-free,
# see IE_PUBLISHED_P_PER_KWH) and the euro conversion moved from a
# hard-coded rate to the ECB semester mean computed from the feed.
# 8 (8 Aug 2026): non-domestic rates step by semester instead of being
# held at one, so the services share of each week is priced at the
# semester that week falls in. Every stored week reprices.
ANCHOR_EPOCH = 8
# The heat/cold split fields carried by both the island entry and,
# from history schema 5, each jurisdiction sub-block.
JUR_SPLIT_KEYS = ("heat_gwh", "cold_gwh",
                  "bill_heat_eur_m", "bill_cold_eur_m",
                  "bill_heat_gbp_m", "bill_cold_gbp_m",
                  "emissions_heat_kt", "emissions_cold_kt")

# ---------------------------------------------------------------- Irish
# anchors: published in sterling, converted at a FETCHED rate.
#
# WHY SEMESTER 2 2025 AND NOT AN EARLIER ONE. The Irish domestic
# electricity series carries government credits as negative taxes -
# UREGNI says so explicitly, EUR1,500 of them since 2022 with the last
# EUR125 in Jan/Feb 2025. That is why Ireland reads 31.3 p/kWh in S2
# 2024, 27.5 in S1 2025 and 35.2 in S2 2025: a 28% jump in one
# semester caused by a credit ending, not by a price moving. This site
# prices the REAL cost of heat, so the credit-bearing semesters are
# unusable and S2 2025 - the first clean one - is the anchor. It also
# happens to be the semester the back-look starts in.
#
# WHY THE RATE IS FETCHED. UREGNI publishes Ireland in sterling,
# having converted Eurostat's euro at the semester average. Recovering
# the euro figure therefore needs that same average, and it scales
# EVERY Irish anchor on the site. It was a hard-coded 0.84 until now,
# which was my estimate for the wrong semester - S2 2025 ran nearer
# 0.87, so the estimate was overstating every ROI anchor by about 4%.
# It is now the ECB semester mean, computed from the daily reference
# rates this pipeline already retains and logged on every run.
IE_SEMESTER = "2025-S2"

# ------------------------------------------------- non-domestic by semester
# Published REMM band prices, pence per kWh, incl CCL excl VAT.
# Electricity is consumption-weighted across every band BELOW Large +
# Very Large, using each semester's own published consumption shares -
# services buildings are not the industrial tail, and the shares move
# enough between semesters (very small 6.1 -> 6.9 -> 5.1% of I&C GWh)
# to be worth taking per semester rather than once. Gas is band I1,
# where services buildings sit; the REMM price bands do not map onto
# the network bands the gas consumption split is published in.
#
# WHY THIS STEPS AND DOMESTIC DOES NOT. There are no regulated
# non-domestic announcements to give finer timing, so the semester IS
# the resolution. Domestic keeps dated steps because the regulator and
# the incumbents publish them.
#
# CONVERSION. Each Irish figure converts at the ECB mean for ITS OWN
# semester - UREGNI sterlings Eurostat's euro at that semester's
# average, so that is the rate that recovers the original. Not the
# week's semester: the figure belongs to the semester it was published
# for, whichever week is being priced.
#
# TAIL. REMM lags about nine months, so weeks past the last published
# semester hold at it rather than extrapolating - flagged, not guessed.
#
# Sources: 2024 AREMM (S2 2024), Q3 2025 QREMM (S1 2025), Q1 2026
# QREMM (S2 2025). Domestic credits do not contaminate these: the
# Irish credits were an electricity-only, household-only measure.
NONDOM_SEMESTERS = {
    #             NI elec  NI gas   IE elec  IE gas
    "2024-S2": (23.457,   8.665,   23.727,  8.579),
    "2025-S1": (22.706,   8.576,   23.221,  8.280),
    "2025-S2": (23.808,   7.995,   22.692,  9.345),
}


def _nondom(a, week_ctx):
    """The week's own non-domestic rates if it carries them, else the
    live anchors. Historic weeks carry them so the services share is
    priced at the semester the week falls in, not at today's."""
    nd = (week_ctx or {}).get("nondom")
    return nd if nd else {"eur": a["nondom_eur_per_kwh"],
                          "gbp": a["nondom_gbp_per_kwh"]}


def semester_of(date_iso):
    """Which semester a week belongs to, by WEEK ENDING. A week
    straddling 30 June or 31 December lands wholly in the semester it
    ends in - two weeks a year, disclosed rather than split."""
    return f"{date_iso[:4]}-S{1 if int(date_iso[5:7]) <= 6 else 2}"


def nondom_for(date_iso, fx_by_semester=None):
    """
    Non-domestic rates in force for a week: NI in sterling, ROI in
    euro, converted at each published figure's OWN semester rate.

    Returns (rates, semester_used). Weeks before the first published
    semester clamp to it; weeks after the last hold at it, which is
    the REMM lag rather than a claim that prices stopped moving.
    """
    keys = sorted(NONDOM_SEMESTERS)
    want = semester_of(date_iso)
    use = keys[0] if want < keys[0] else (keys[-1] if want > keys[-1]
                                          else want)
    if use not in NONDOM_SEMESTERS:
        use = max(k for k in keys if k <= want)
    ni_e, ni_g, ie_e, ie_g = NONDOM_SEMESTERS[use]
    rate = (fx_by_semester or {}).get(use) or IE_FX["rate"]
    return {
        "gbp": {"electricity": round(ni_e / 100, 4),
                "gas": round(ni_g / 100, 4)},
        "eur": {"electricity": round(ie_e / 100 / rate, 4),
                "gas": round(ie_g / 100 / rate, 4)},
    }, use
# Used ONLY if the semester mean cannot be computed. Logged as
# unverified when it fires, because a silent fallback here would move
# the whole ROI side without saying so.
IE_FX_FALLBACK = 0.87
IE_FX = {"rate": IE_FX_FALLBACK, "source": "fallback (UNVERIFIED)",
         "semester": IE_SEMESTER}

# Published Irish figures, pence per kWh, semester 2 2025, as UREGNI
# prints them. Domestic incl all taxes; non-domestic incl CCL excl VAT.
# Electricity non-domestic is consumption-weighted across the bands
# below Large + Very Large using UREGNI's own published consumption
# shares (5.1 / 32.9 / 19.1 / 26.8 / 16.1%); gas non-domestic is band
# I1, where services buildings sit.
IE_PUBLISHED_P_PER_KWH = {
    "domestic_electricity": 35.2,
    "domestic_gas": 11.3,
    "nondom_electricity": 22.68,
    "nondom_gas": 9.3,
}
# Level from the semester, timing from the announcements. Electric
# Ireland held from Oct 2022 to 1 Jul 2026, then +8% / +7.7%.
IE_STEPS = [("2026-07-01", {"domestic_electricity": 1.08,
                            "domestic_gas": 1.077})]


def ie_eur(key, date_iso=None):
    """Published Irish sterling figure -> EUR/kWh at the fetched
    semester rate, stepped by any announcement on or before the date."""
    v = IE_PUBLISHED_P_PER_KWH[key] / 100.0 / IE_FX["rate"]
    for frm, steps in IE_STEPS:
        if date_iso and date_iso >= frm and key in steps:
            v *= steps[key]
    return round(v, 4)


def semester_means(daily, min_days=110):
    """{date: rate} -> {'YYYY-S1'|'YYYY-S2': mean}. Semesters with
    fewer than min_days observations are DROPPED, not averaged thin -
    a part-semester mean is indistinguishable from a whole one and
    would silently mis-scale every Irish anchor."""
    buckets = {}
    for d, v in (daily or {}).items():
        try:
            y, m = int(d[:4]), int(d[5:7])
        except (ValueError, IndexError):
            continue
        buckets.setdefault(f"{y}-S{1 if m <= 6 else 2}", []).append(v)
    return {k: round(sum(v) / len(v), 5)
            for k, v in buckets.items() if len(v) >= min_days}
UA = {"User-Agent": "ioi-heatsplit/0.5 (contact@causewaygt.com)"}
TIMEOUT = 90
RETRIES = 3

# Best-effort context feeds - failure marks stale but never pages or
# fails the run; regressions stay visible via status badges and flags.
SOFT_FEEDS = {"gb_oil", "entsog_probe", "sem_mix",
              "eirgrid_probe"}

# Feeds known broken for reasons outside this pipeline - marked stale,
# logged, but neither paged nor allowed to fail the run.
EXPECTED_DOWN = {}
# eirgrid was expected-down 09-18 Jul 2026 (dashboard redesign); restored
# via the /api/chart/ endpoint captured by browser probe.

# Population weights for degree days - Causeway judgement figures (dagger).
# Challenge and input welcome at contact@causewaygt.com.
STATIONS = {
    #  name        lat      lon     weight  jurisdiction
    "Dublin":    (53.35,  -6.26,   0.40, "ROI"),
    "Belfast":   (54.60,  -5.93,   0.20, "NI"),
    "Cork":      (51.90,  -8.47,   0.12, "ROI"),
    "Galway":    (53.27,  -9.05,   0.08, "ROI"),
    "Limerick":  (52.66,  -8.63,   0.08, "ROI"),
    "Derry":     (54.99,  -7.31,   0.06, "NI"),
    "Waterford": (52.26,  -7.11,   0.06, "ROI"),
}
HDD_BASE_C = 15.5

# Fill after browser XHR probe (the one remaining probe)
EIRGRID_ENDPOINT = "https://www.smartgriddashboard.com/api/chart/"
# Captured by browser XHR probe, 18 Jul 2026. Response schema is the
# pre-redesign DashboardService shape unchanged: {"Rows":[{"EffectiveTime":
# "18-Jul-2026 00:15:00","FieldName":"SYSTEM_DEMAND","Region":"NI",
# "Value":573}, ...]} - 15-min actuals with nulls for future intervals,
# half-hourly DEMAND_FORECAST_VALUE rows appended.

NTFY_TOPIC = None

# Set by main() before the feed loop - lets feeds merge history across runs
PREVIOUS_FEEDS: dict = {}
PREVIOUS_DERIVED: dict = {}

# ------------------------------------------------- annual anchors
# Sourced figures cite their publication; every judgement figure is marked
# with a dagger and is a current Causeway Energies estimate - challenge and
# input welcome at contact@causewaygt.com.

ANCHORS = {
    "year": 2024,
    "roi": {
        "residential_heat_twh": 22.3,      # SEAI Energy in Ireland 2025
        "services_heat_twh": 8.5,          # dagger - from SEAI sector shares
        "fuel_shares": {"oil": 0.565, "gas": 0.251, "peat": 0.067,
                        "electricity": 0.08, "other": 0.037},
        #             oil/gas/peat SEAI 2024; electricity/other dagger
        "gas_indigenous": 0.18,            # SEAI H1-2025, Corrib, falling
        "elec_indigenous": 0.413,          # SEAI RES-E 2024
    },
    "ni": {
        "residential_heat_twh": 10.0,      # dagger - NISRA stock x intensity
        "services_heat_twh": 3.0,          # dagger
        "fuel_shares": {"oil": 0.62, "gas": 0.26, "peat": 0.0,
                        "electricity": 0.08, "other": 0.04},
        #             oil NISRA CHS 2024/25; gas/electricity/other dagger
        "gas_indigenous": 0.0,             # all NI gas arrives via Moffat
        "elec_indigenous": 0.46,           # dagger - DfE yr-to-Mar-2026 ~48%
    },
    "efficiency": {"oil": 0.82, "gas": 0.85, "peat": 0.60,
                   "electricity": 1.0, "other": 0.70},          # dagger
    "geothermal_spf": 4.0,                                       # dagger
    # Air-source heat pump model - all dagger. Carnot-fraction COP against
    # the HDD-derived, demand-weighted outdoor temperature; defrost derate
    # for humid Irish winters; DHW share at higher flow. Calibrated to the
    # GB Electrification of Heat field-trial median (~2.8-2.9) rather than
    # laboratory SCOP figures.
    # PROVENANCE (audit, 2 Aug 2026 - these become the v7 hourly COP
    # engine's foundation, so every parameter is declared here):
    #   flow_c 45.0 - mid-range radiator flow for a retrofit onto
    #     existing emitters (weather compensation 30-50 C in v7);
    #     dagger, convention not measurement.
    #   carnot_fraction 0.38 - typical field-observed fraction of
    #     Carnot for modern inverter ASHP; consistent with the SPF
    #     anchors below and with UK field-trial ranges 0.35-0.45.
    #     Dagger; the SPF anchor pins the annual, so this shapes the
    #     seasonal SHAPE, not the level.
    #   defrost_derate 0.90 - performance penalty in the 0-7 C humid
    #     band that dominates the island's winter; dagger.
    #   dhw_share 0.20 - hot water as a share of heat-pump duty in
    #     the modelled route; sits alongside the 22.4% DHW share of
    #     national heat input (SEAI National Heat Study re-anchoring,
    #     Aug 2026); dagger.
    "ashp": {"flow_c": 45.0, "carnot_fraction": 0.38,
             "defrost_derate": 0.90, "dhw_share": 0.20,
             "dhw_flow_c": 55.0, "dhw_source_c": 10.0},
    # Fuel emission factors, g CO2e per kWh of fuel INPUT (not
    # delivered): oil/gas/peat from the SEAI emission-factor series
    # for kerosene, natural gas and milled peat; "other" is a
    # biomass-weighted residual (dagger); electricity is the anchor
    # replaced at runtime by live all-island grid intensity.
    "ef_g_per_kwh": {"oil": 257, "gas": 205, "peat": 340,
                     "other": 100, "electricity": 280},          # dagger -
    # electricity factor replaced by live grid intensity once eirgrid returns
    "indigenous": {"oil": 0.0, "peat": 1.0, "other": 0.9},      # dagger
    # CROSS-CALIBRATED 18 Jul 2026 against the UK sibling (input
    # basis, buildings only): UK 430.6 TWh / 68m = 6.3 MWh/person;
    # island 43.8 / 7.1 = 6.2 - per-capita parity, ratio 0.98.
    # RE-ANCHORED 6 Aug 2026 to an Irish source. SEAI National Heat
    # Study Report 1 puts residential hot water at 25% of residential
    # heat demand (space heating 75%). Applied to this file's own
    # sector split - island residential 32.3 TWh, services 11.5 TWh -
    # with services hot water at 15% (dagger; SEAI does not publish
    # the services split), the island hot-water share is 22.4% and
    # the weather-driven space share 77.6%. This REPLACES the
    # UK-aligned 18.3% adopted in July, which was explicitly marked
    # as pending an SEAI-specific figure. Caption symmetry with the
    # UK sibling is now broken by design: both sites shape space heat
    # by degree days and carry hot water flat, but each uses its own
    # national hot-water share. Weekly = A x [0.224/52 + 0.776 x Hw/Hy].
    "space_heat_fraction": 0.776,   # = 1 - DHW share; see above
    "kerosene_kwh_per_litre": 10.35,   # industry standard figure
    # Tariff anchors, July 2026 pass. Sourced bands, dagger on the point:
    #  ROI electricity: standard 24h ~35c (Electric Ireland, May 2026);
    #    Eurostat H2-2025 all-in ~40c; anchor 36c.
    #  ROI gas: standard unit ~11-12c incl 9% VAT; anchor 11.5c.
    #  NI electricity: Power NI 1 Jul 2026 review - GBP1,093/yr at 3,200 kWh
    #    -> ~32.5p unit ex-standing; anchor 32.5p.
    #  NI gas: Ten Towns GBP972/yr at 12,000 kWh (1 Jul 2026) -> ~7.5p unit.
    # Standard-tariff basis - time-of-use/night rates materially lower for
    # heat-pump households; see heat_gap basis note.
    # Sector blend (UK-pattern, Irish placeholders - dagger until read
    # from the SEAI national energy balance / BER end-use work and
    # Eurostat band prices nrg_pc_203 (gas) / nrg_pc_205 (electricity);
    # Eurostat publishes on ~half-year lag, stated in methodology).
    # dom_share: domestic fraction of each fuel's buildings-heat input.
    # Oil is priced identically across sectors (tanker market), so only
    # gas and electricity blend. Cooling is wholly non-domestic.
    "dom_share": {
        "roi": {"gas": 0.60, "electricity": 0.55},
        "ni": {"gas": 0.55, "electricity": 0.55},
    },
    # NON-DOMESTIC RATES - Utility Regulator Retail Energy Market
    # Monitoring, semester 2 2025 (Q1 2026 QREMM, published 16 Jun 2026).
    #
    # WHAT WAS WRONG. Both jurisdictions carried LARGE-USER prices.
    # The NI pair (24.0p / 5.5p) was the GB QEP manufacturing average
    # lifted verbatim from the UK sibling - wrong jurisdiction AND
    # wrong sector. This site's non-domestic scope is SERVICES
    # BUILDINGS - offices, retail, hospitality, public estate - which
    # sit at the small end of the distribution where prices run near
    # domestic, not at the industrial end where they are little more
    # than half that. We had been pricing offices at smelter rates.
    #
    # ELECTRICITY - consumption-weighted across the published bands,
    # excluding only Large + Very Large. That top band is seventeen NI
    # connections and 683 GWh of unambiguously heavy industry and data
    # centres; every band below it is where services buildings live.
    # Weighting is UREGNI's published I&C consumption share per band,
    # applied to both jurisdictions because Ireland's own split is not
    # published alongside - a dagger, but a much smaller one than the
    # band choice, and the two ladders track each other closely.
    # Semester 2 2025 (Q1 2026 QREMM), with UREGNI's own published
    # consumption shares - no proxy weighting needed any more:
    #   band (MWh/yr)      NI p/kWh   IE p/kWh   % of I&C GWh
    #   very small <20        28.6       26.2       5.1
    #   small 20-500          26.3       25.4      32.9
    #   small/med 500-2k      23.3       22.2      19.1
    #   medium 2k-20k         20.2       19.0      26.8
    #   large+VL >20k         17.8       16.1      16.1   <- excluded
    #   weighted (services)   23.8       22.7      83.9
    # Ireland has crossed BELOW NI since 2024 on both fuels.
    # Ireland converted back to euro at 0.84, the semester average -
    # UREGNI converts Eurostat's euro figures to sterling for the
    # comparison, so this recovers the original.
    #
    # GAS - band I1 (<278,000 kWh/yr), which is where services
    # buildings overwhelmingly sit. NOT consumption-weighted, because
    # the REMM price bands (I1/I2/I3&I4) do not map onto the network
    # EUC bands the consumption split is published in, and NI I&C gas
    # consumption is ~68% daily-metered heavy industry which would
    # drag a weighted figure to an industrial number. I1 for
    # reference: NI 8.67, Ireland 8.58; I2 7.55 / 7.76; I3&I4 5.82 /
    # 3.23. Dagger on the band choice.
    #
    # THE BASIS. UREGNI derives these the way Eurostat does - volume
    # sold and revenue submitted per size band, revenue over volume -
    # so standing charges are INCLUDED by construction. That makes
    # them like-for-like with the SEAI/Eurostat semester prices on the
    # ROI side, which is what closes the cross-jurisdiction basis gap
    # for the services share. I&C figures include CCL and EXCLUDE VAT,
    # which is the convention this site already applies to services.
    #
    # VINTAGE - and a live consequence. Held at S2 2025, the latest
    # published semester (Q1 2026 QREMM, 16 Jun 2026), and they do NOT move week to week. REMM lags
    # about nine months. Since the domestic rates ARE stepped through
    # to 2026 and NI gas fell about 15% over that span, NI
    # non-domestic gas (8.67p, 2024) now prints ABOVE NI domestic gas
    # (7.70p, Jul 2026). That is a vintage artefact, not a claim that
    # small business gas is cheaper than household gas - at a common
    # vintage the ordering is right (domestic 9.88p incl VAT, 9.41p
    # excl, against 8.67p non-domestic). It is disclosed rather than
    # escalated away, because applying regulated domestic steps to
    # unregulated business contracts would be a dagger on a dagger.
    # It closes when REMM publishes newer semesters.
    # Level from the semester series and timing from tariff
    # announcements is the intended design; until that is built the
    # services share of every week is priced at a 2024 semester.
    # set per run by apply_ie_fx() from the fetched semester rate
    "nondom_eur_per_kwh": {"electricity": 0.284, "gas": 0.102},
    "nondom_gbp_per_kwh": {"electricity": 0.2381, "gas": 0.080},
    "retail_eur_per_kwh": {"gas": 0.115, "electricity": 0.36},
    "retail_gbp_per_kwh": {"gas": 0.075, "electricity": 0.325},
    # The cold economy - island cooling loads, electricity basis.
    #  dc: CSO metered consumption ~22% of ROI electricity (sourced);
    #    29% projected by 2028. Other loads are Causeway anchors, dagger:
    #  refrigeration: dairy (farm milk cooling + processing), meat, cold
    #    stores, retail - the food-export cold chain;
    #  process: pharma/semiconductor chilled-water and process cooling;
    #  comfort: commercial space cooling and ventilation (summer-peaked,
    #    shaped by live overheating degree-hours when available);
    #  ni_cooling: food processing, retail refrigeration, small DC.
    #  rejection: refrigeration/comfort reject compressor work PLUS the
    #    heat they pump - factors dagger; DC rejects ~all electricity.
    # CORRECTION 6 Aug 2026 (SEAI National Heat Study, Report 1):
    # the census previously carried TOTAL data-centre electricity as
    # cold-economy cooling load. It is not: SEAI finds cooling is
    # ~14% of data-centre electricity, the balance being IT load that
    # BECOMES heat rather than removing it. The DC line is now priced
    # at its cooling share; the heat it rejects is unchanged, so the
    # rejection and service factors rise to match (see below) and the
    # delivered cooling service is preserved. SEAI's absolute (0.4
    # TWh) is a 2019-base figure and is stale against today's fleet -
    # the transferable parameter is the SHARE, applied to the current
    # metered total.
    # Heat-pump stock, from the two censuses - the anchor that lets the
    # electricity line split into resistive and heat-pump, and the free
    # ambient heat appear on the out-bar.
    #   ROI: Census 2022 - 71,000 households with heat pumps, of which
    #        57,000 air source and 14,000 ground source. A FLOOR: NZEB
    #        has made heat pumps near-universal in new dwellings since
    #        2019 and four further years of SEAI grants have followed,
    #        so 2026 stock is materially higher (dagger uplift 1.5).
    #   NI:  Census 2021 Table 27 - air source 0.09% and geothermal
    #        0.08% of 768,808 households = ~692 and ~615. The 615
    #        confirms the 500-700 dagger this file has carried for the
    #        NI domestic ground-source count.
    #   Non-domestic tail: WGC2026 gives ROI 20,128 ground-source
    #        systems against the census's 14,000 households; the ~6,000
    #        difference is the non-domestic and communal tail. At the
    #        fleet's ~11 kW average the stock is overwhelmingly
    #        domestic-scale. SSRH, the ROI non-domestic register, had
    #        supported only ~EUR0.5m of heat-pump projects by 2023.
    #   Communal schemes serving apartment blocks are UNDER-COUNTED by
    #        both censuses (a shared plant reads as many households);
    #        district heat is ~1% of island heat, so the error is small
    #        but real - dagger.
    "heat_pumps": {
        "roi": {"households": 71000, "census_uplift": 1.5,
                "nondom_equivalent": 6000},
        "ni": {"households": 1307, "census_uplift": 1.5,
               "nondom_equivalent": 200},
        # delivered heat per heat-pumped dwelling, MWh/yr - below the
        # stock average because heat-pumped homes are newer and better
        # insulated; dagger.
        "delivered_mwh_per_dwelling": 9.0,
    },
    "cool": {"dc_share_of_roi_elec": 0.22, "dc_share_2028": 0.29,
             "dc_cooling_share": 0.14,     # SEAI NHS Report 1
             "roi_elec_twh": 31.0,
             "loads_twh": {"dc": None,          # computed: share x elec
                           # Cold-economy load census, TWh/yr of electricity. DC from CSO
    # data-centre metered consumption (22% of ROI electricity);
    # the remainder are Causeway estimates pending component-level
    # sourcing in the autumn pass - each a dagger:
    #   refrigeration 2.3 - food/dairy/pharma cold chain, scaled from
    #     the sector's share of industrial electricity
    #   process 0.8 - industrial process cooling excl. cold chain
    #   comfort 1.0 - comfort cooling AND ventilation in services
    #     buildings (the UK sibling's whole cooling scope, roughly)
    #   ni_all 1.2 - the NI cold economy, all categories
    "refrigeration": 2.3,
                           "process": 0.8,
                           "comfort": 1.0,
                           "ni_all": 1.2},
             # CENSUS VALIDATED 7 Aug 2026 against the SEAI
             # National Heat Study supporting data (final-energy
             # table). SEAI ROI cooling electricity: commercial
             # 2,868 + public 221 + agriculture 93 + industry 804
             # + data centres 1,055 (their 2026 projection) =
             # 5,041 GWh. Our ROI lines (dc 955 + refrigeration
             # 2,300 + process 800 + comfort 1,000) = 5,055 GWh -
             # 14 GWh apart on 5 TWh, with industry/process
             # matching to 4 GWh. The earlier suspicion that
             # refrigeration was light came from comparing SEAI's
             # USEFUL demand against our ELECTRICITY; the
             # like-for-like comparison holds. The open question is
             # now the internal split between refrigeration and
             # comfort within the commercial total, not the total.
             # AUDIT 18 Jul 2026: census scope is the full cold economy
             # (DC + cold chain + process + comfort) - deliberately
             # wider than comfort-only national cooling lines (e.g. the
             # UK tracker's ECUK-anchored figure); the two are not
             # ratio-comparable and the hero basis says so. Component
             # source hardening scheduled for the autumn cycle.
             # dc 7.1: with the line repriced to cooling electricity
             # only, the heat rejected is unchanged (the whole DC draw
             # ends as heat), so heat rejected per unit COOLING
             # electricity = 1/0.14 ~ 7.1. Physical output preserved
             # across the 6 Aug 2026 correction; only the purchased
             # side moved.
             "rejection_factor": {"dc": 7.1, "refrigeration": 2.5,
                                  # Heat-rejection factors: kWh of heat rejected per kWh of
    # electricity drawn - approximately (COP + 1) for vapour
    # compression, near 1.0 for DC where most draw is IT load
    # rejected as heat. All dagger, all Causeway judgement.
    "process": 2.0, "comfort": 3.0,
                                  "ni_all": 1.8},
             # District heat as a share of national heat - order 1% on the
    # island against ~2% GB and 50%+ in Denmark; dagger, and the
    # number the geothermal panel's district-heat argument rests on.
    "dh_share_of_national_heat": 0.01,
             # dagger: electricity saving on cooling load moved to
             # ground-coupled systems (free cooling / high-EER ATES)
             "ground_cooling_saving": 0.70,
             # dagger: cooling service delivered per unit electricity
             # (seasonal system EER). DC ~1.0: its service is heat
             # removal, roughly its electricity; vapour-compression
             # loads deliver a multiple of theirs.
             # Service factors: kWh of COOLING SERVICE delivered per kWh of
    # electricity - the delivered-basis multiplier that lets the
    # out-bar legitimately exceed the in-bar. Dagger throughout.
    # dc 6.1: cooling service (heat removed from IT equipment) per
    # unit cooling electricity - an effective seasonal EER, high
    # because Ireland's climate permits extensive free cooling.
    # Raised from 1.0 with the 6 Aug 2026 repricing.
    # RE-ANCHORED 7 Aug 2026 to the SEAI National Heat Study
    # supporting data, which publishes useful cooling demand AND
    # final cooling consumption per sector - their ratio IS the
    # service factor. SEAI's ROI implied ratios: commercial 2.07,
    # public 2.57, industry 1.00, agriculture 1.00, aggregate 1.86
    # (excluding data centres). We adopt 2.07 for the commercial-type
    # loads - refrigeration and comfort - and hold process at 2.2
    # dagger rather than copying SEAI's 1.00, which is a modelling
    # pass-through (a COP of exactly 1.0 for process chilling is not
    # a physical claim). NI carries the commercial ratio. The data
    # centre factor stays 6.1 on its own provenance: Irish free
    # cooling genuinely delivers an effective EER above 6, and SEAI
    # models data centres in a separate sheet at 14% cooling share,
    # which our census already follows.
    "cooling_service_factor": {"dc": 6.1, "refrigeration": 2.07,
                                        "process": 2.2, "comfort": 2.07,
                                        "ni_all": 2.07}},
}

# Policy events rendered as chart annotations - date, jurisdiction, label.
EVENTS = [
    {"date": "2025-10-08", "jur": "ROI",
     "label": "Carbon tax to \u20ac71/t \u2013 motor fuels"},
    {"date": "2026-03-16", "jur": "UK",
     "label": "UK \u00a350m oil support; CMA review"},
    {"date": "2026-04-01", "jur": "ROI",
     "label": "NORA levy paused (two months)"},
    {"date": "2026-05-01", "jur": "ROI",
     "label": "Heating-fuel carbon increase postponed"},
    {"date": "2026-07-13", "jur": "ROI",
     "label": "Heat Bill advanced \u2013 district heating framework"},
    {"date": "2026-10-14", "jur": "ROI",
     "label": "Carbon tax \u20ac63.50\u2192\u20ac71/t \u2013 heating fuels (due)"},
]

# Value-level caveats, distinct from fetch status - machine-carried nuance.
FEED_FLAGS = {
    "ccni_oil": ["Verification pending that parsed values are the NI "
                 "average series, not a council-area series"],
    "oil_bulletin": ["Bulletin heading is gas oil; treated as ROI "
                     "heating-oil (kerosene) price level - Causeway "
                     "judgement"],
    "hdd": ["Population weights are Causeway estimates"],
    "gb_oil": ["BoilerJuice site average of lowest quotes, not a survey "
               "average - basis differs from CCNI; ONS kj5u discontinued "
               "Jan 2025"],
}




# ------------------------------------------------- geothermal register
# NI >60 kW register: Causeway research pass, June 2026, updated July 2026
# Martinstown (325 kW, 2015) is ROI - confirmed by S. Todd, counted in
# the WGC ROI totals, not carried here.
# (Ryan Daly pers. comm. - Randalstown 44 kW and Strabane 18 kW confirmed
# sub-threshold, logged as exclusions). ROI anchors: WGC2026 Country Update
# (Ireland, Blake, Pasquali, Dunphy & Hunter Williams) - sourced.
# Corrections welcome at contact@causewaygt.com.
GEO = {
    "ni_register": [
        {"id": "R1", "site": "McClay Library, QUB", "year": 2009,
         "kw": None, "duty": "cooling", "type": "closed vertical",
         "status": "candidate - unconfirmed", "confirmed": False},
        {"id": "R2", "site": "Lyric Theatre, Belfast", "year": 2011,
         "kw": 120, "duty": "cooling", "type": "open single-well",
         "status": "operational - minor issues", "confirmed": True,
         "note": "120 kW demand-inferred, not metered"},
        {"id": "R3", "site": "Giant's Causeway Visitor Centre", "year": 2012,
         "kw": 72, "duty": "heating", "type": "horizontal mat",
         "status": "operational", "confirmed": True},
        {"id": "R4", "site": "Girdwood Community Hub, Belfast", "year": 2016,
         "kw": 108, "duty": "heating", "type": "hybrid vertical + slinky",
         "status": "operational - impaired (~60% of 180 kW design)",
         "confirmed": True},
        {"id": "R5", "site": "QUB School of Biological Sciences", "year": 2018,
         "kw": 0, "duty": "cooling", "type": "open doublet",
         "status": "never commissioned (design flow set-point error)",
         "confirmed": True},
        {"id": "R7", "site": "QUB Business School Student Hub", "year": 2023,
         "kw": 280, "duty": "heating + DHW",
         "type": "closed vertical, 40 x 125 m, Sherwood Sandstone",
         "status": "operational", "confirmed": True},
        {"id": "R8", "site": "UU Jordanstown HPSC", "year": None,
         "kw": None, "duty": "heating", "type": "GSHP",
         "status": "unconfirmed", "confirmed": False},
    ],
    "ni_exclusions": [
        {"site": "Randalstown", "kw": 44, "detail": "2 x 22 kW",
         "source": "Ryan Daly, Jul 2026"},
        {"site": "Strabane", "kw": 18, "detail": "3 x 6 kW Ecogeo Lite "
         "(+3 x 9 kW ASHP out of scope)", "source": "Ryan Daly, Jul 2026"},
    ],
    "ni_domestic": {"low": 500, "high": 700,
                    "note": ("MCS ~386-450 certified plus pre-certification "
                             "era estimate - Causeway triangulation, "
                             "dagger")},
    "roi": {"capacity_mwth": 225, "heat_gwh": 293, "cooling_gwh": 11.9,
            "units": 20128, "new_2024_mwth": 7.4, "proj_2028_mwth": 261,
            "deep_plants": 0,
            # sector shares per WGC2026 text (approximate - sum > 100 in
            # the source; presented as reported)
            "sector_share_pct": {"residential": 85, "commercial": 14,
                                 "industrial": 4},
            "gshp_share_of_hp_market_pct": 4,
            "source": ("WGC2026 Country Update: Ireland - Ireland, Blake, "
                       "Pasquali, Dunphy & Hunter Williams, June 2026")},
    "per_capita_w": {"roi": 42, "ni": 3,
                     "note": "installed Wth per person - NI dagger"},
    "population_m": {"roi": 5.3, "ni": 1.92},   # dagger, mid-2026
    "eflh_h": 2000,   # equivalent full-load heating hours - dagger
    # European reference points, installed GSHP Wth per person -
    # shallow basis, EGC 2025 capacities over mid-2020s populations,
    # dagger: Sweden 8.12 GWth/10.5m; NL 2.49 GWth/17.8m (ATES-heavy);
    # France 2.29 GWth/68m.
    "reference_w_pp": {"Sweden": 773, "Netherlands": 140, "France": 34},
    # Comparator installed capacity, MWth - EGC 2025 country updates
    # (data year 2024), replacing the older WGC2023-lineage values on
    # 27 Jul 2026 audit: Sweden had understated ~18% (8,120 shallow,
    # 690k units, plus Lund's 47 MW heat-pump-coupled deep), the
    # Netherlands ~16% (doublet fleet 367 MWth deep, ~28 doublets,
    # mostly horticulture), France was ~8% high (older inflated
    # shallow estimate; true split is less shallow, more Paris Basin
    # Dogger district heating). ROI's own 225 stays on its WGC2026
    # citation - EGC 2025 carries 224, same lineage, rounding-level.
    "reference_mwth": {
        "Sweden": {"shallow": 8120, "deep": 47},
        "Netherlands": {"shallow": 2486, "deep": 367},
        "France": {"shallow": 2293, "deep": 724}},
    # EGEC Geothermal Market Report 2025, Key Findings - sourced.
    # Note: EGEC counts units SOLD in 2025; the WGC2026 paper reports
    # capacity COMMISSIONED in 2024 (+7.4 MWth). Different measures and
    # years - both carried, not reconciled.
    "egec_2025": {
        "eu_ghp_units_m": 2.55, "eu_ghp_gwth": 39.2,
        "eu_ghp_heat_twh": 88, "eu_people_served_m": 10.6,
        "eu_sales_2025": 123000, "eu_sales_growth_pct": 10,
        "ghp_sales_2025": {"Ireland": 1409, "Sweden": 26785,
                           "Netherlands": 22148, "Germany": 19125,
                           "United Kingdom": 4070},
        "dhc_systems": 434, "dhc_gwth": 6.0,
        "dhc_rank": "second-largest renewable source for district heat",
        "source": "EGEC Geothermal Market Report 2025, Key Findings"},
    "ni_capacity_mwth_est": 6.6,   # >60 kW register + domestic - dagger
    "island_today_twh": 0.30,
    "pipeline": [
        "GEMINI (EUR 20m, PEACEPLUS): 3 shallow demos Sligo + Belfast "
        "(NIHE, NI Water); deep 2 km at Grangegorman, drilling late 2027",
        "GeoEnergy NI: Stormont shallow boreholes drilled 2024; CAFRE "
        "Greenmount deep doublet consented (LA03/2025/0443/F)",
        "GSI deep scientific boreholes: BHT 22.6-38 C at 1 km, five "
        "completed 2022-2026",
    ],
}


# ------------------------------------------------- why heat? (whole economy)
# The zoom-out panel: annual, all-island, whole-economy anchors for the
# three energy services. Sourced where a publication exists; allocations
# are Causeway derivations, kept deliberately round and dagger-marked.
#  services: ROI TFC ~143 TWh (SEAI Energy in Ireland 2025) + NI ~48
#    dagger. Heat includes industrial process heat (dagger allocation);
#    power = non-heat electricity.
#  spend: retail-basis, EUR bn - dagger throughout.
#  imports: allocation of ~179 TWh imported primary energy (81.2% of
#    220.7 TWh PES - Causeway island Sankey, 2024).
#  emissions: ROI ~32 Mt energy CO2 (SEAI) + NI ~11 dagger.
WHY_HEAT = {
    "tfc_twh": 191,
    "services_twh": {"heat": 64, "transport": 83, "power": 35},
    "spend_eur_bn": {"heat": 6.0, "transport": 15.0, "power": 10.5},
    "imports_twh": {"heat": 52, "transport": 79, "power": 28},
    "emissions_mt": {"heat": 16, "transport": 17, "power": 9},
    "basis": ("Annual, all-island, whole-economy - dagger throughout. "
              "Services: SEAI Energy in Ireland 2025 (ROI TFC ~143 TWh) "
              "+ NI dagger; heat includes industrial heat. Imports: "
              "allocation of 81.2%-imported primary energy, 220.7 TWh "
              "PES (Causeway island Sankey 2024). Emissions: SEAI + NI "
              "dagger. Challenge and input welcome at "
              "contact@causewaygt.com"),
}


# ------------------------------------------------ per-week tariff history
# No regulated cap exists in either jurisdiction's oil-heated world and
# none in ROI electricity - representative standard domestic unit rates
# by period, dagger throughout (the representativeness caveat the UK
# series does not carry). Backfilled weeks freeze at first computation:
# a wrong early row is permanent, so rows are added only once verified
# against SEAI/CSO/CRU archives. Single verified period at porting.
# NI TARIFF BASIS - rebuilt 8 Aug 2026. Read this before changing a
# number.
#
# WHAT CHANGED. The NI rows were previously a percentage chain off a
# single supplier - Firmus Energy (Ten Towns) - with the standing
# charge stripped out to leave a bare unit rate. Two problems. Firmus
# regulates about 75,756 customers while SSE Airtricity Gas Supply
# covers roughly 195,000 in Greater Belfast plus 3,200 in the West, so
# the smaller market was setting the NI gas bill. And the two
# suppliers are not structurally comparable: SSE's domestic tariff is
# banded with NO standing charge (its two published unit rates alone
# reproduce the typical bill exactly), while Firmus charges a unit
# rate plus a standing charge. Stripping "the standing charge" from
# one and not the other imports a tariff structure as a price.
#
# THE BASIS NOW. Effective all-in pence per kWh at the Utility
# Regulator's own stated consumption - 12,000 kWh for gas, 3,200 kWh
# for electricity - taken from the published annual bills, INCLUDING
# 5% VAT, and weighted across suppliers by regulated customer count
# (SSE 0.7235 / Firmus 0.2765). All-in rather than unit-only is also
# the right aggregate: the national bill is households x (standing +
# unit x consumption), and average NI domestic consumption sits near
# the Regulator's basis.
#
# EFFECT. NI gas rises about 12% for Oct 2025 and 14% for Apr 2026
# against the old rows, roughly half from restoring the standing
# charge and half from weighting in the larger and dearer supplier.
# NI electricity rises about 5%. These are corrections, not new
# estimates - but they move the NI headline, so revert here first if
# an NI figure looks wrong after this release.
#
# SOURCES. UREGNI tariff review briefing papers. SSE maximum average
# price 246.78 p/therm (eff 1 Apr 2024, unchanged to 30 Sep 2025) ->
# 225.87 (1 Oct 2025, -8.47%) -> 207.59 (1 Apr 2026, -8.10%), not
# reviewed in Jul 2026 - Annex 1 of each paper carries the series back
# to Apr 2015 if the record is ever extended further. Firmus bills
# GBP1,014 -> 934 (-7.86%, 1 Oct 2025) -> 840 (-10.1%, 1 Apr 2026) ->
# 972 (+15.7%, 1 Jul 2026); the three announced GBP steps reproduce
# those percentages to 0.04pp, which is the check that the chain is
# sound. Power NI GBP989 -> 1,029 (+4%, 1 Oct 2025) -> 1,093 (+6.2%,
# 1 Jul 2026), no change in Apr 2026.
#
# ROI DOMESTIC - rebuilt 8 Aug 2026 onto the same KIND of basis as
# NI. Previously a standard unit rate incl 9% VAT with no standing
# charge, which meant the two jurisdictions' domestic bills were not
# comparable at the component level: NI carried the standing charge
# and ROI did not.
#
# The like-for-like artefact is the Eurostat band price, which is
# computed as total revenue over volume for a consumption band and
# therefore includes standing charges by construction - the same
# property that made the REMM non-domestic bands work. Band D2
# (5,557-55,557 kWh gas) and medium domestic (2,500-4,999 kWh
# electricity), incl all taxes, semester 2 2024, from the 2024 AREMM
# data publication: Ireland 31.33 p/kWh electricity and 11.30 p/kWh
# gas, converted back to euro at 0.84, the semester average UREGNI
# used to sterling them in the first place.
#
# LEVEL from the semester, TIMING from the announcements - the same
# split the NI oil bridge uses. Electric Ireland held prices from Oct
# 2022 to 1 Jul 2026, so the S2 2024 level carries unchanged through
# the backfill row, Oct 2025 and Apr 2026, then steps +8% electricity
# and +7.7% gas on 1 Jul 2026.
#
# TWO DAGGERS, both worth knowing. (1) The band price is a market-wide
# average across every supplier and tariff, including discounts, while
# the ROI steps come from the incumbent's standard tariff - so
# non-incumbent movements between Dec 2024 and Jul 2026 are not
# captured, and cannot be until REMM publishes newer semesters. (2)
# The residual asymmetry with NI is now scope, not basis: NI is
# incumbent-weighted regulated bills, ROI is a market-wide average.
# Measured against the same tables, NI's regulated electricity runs
# about 7% ABOVE its own market band and its regulated gas about 18%
# BELOW - so the two do not bias in one direction and neither is a
# simple correction on the other.
#
# VAT CONVENTION (applies to both jurisdictions and to the UK sibling)
# Domestic rates INCLUDE VAT - 5% NI, 9% ROI. Non-domestic rates
# EXCLUDE it, because businesses recover input VAT and it is not a
# cost to them. That is Eurostat level 3 for households and level 2
# for non-households.
#
# NO CLAMP. tariffs_for() REFUSES any date before the first row rather
# than resolving it to that row. The clamp was live and wrong: once
# the price panel's floor came off it reached April 2024, and 68 weeks
# were quietly priced with August-2025 tariffs. Extending this table
# backwards is what restores them - UREGNI briefing papers carry dated
# annexes (SSE Airtricity to 2008, firmus to 2015) and SEAI publishes
# ROI cent/kWh to 2008 Q1.
TARIFF_HISTORY = [
    # (from_date, {eur: {electricity, gas}, gbp: {electricity, gas}})
    # BOTH are now all-in effective rates at a stated consumption,
    # including VAT and standing charges. gbp: UREGNI regulated bills,
    # gas weighted SSE/Firmus by customers. eur: Eurostat band prices
    # (S2 2024) stepped by the Electric Ireland announcements.
    ("2025-08-06", {"eur": None,
                    "gbp": {"electricity": 0.3091, "gas": 0.0884}}),
    ("2025-10-01", {"eur": None,
                    "gbp": {"electricity": 0.3216, "gas": 0.0809}}),
    ("2026-04-01", {"eur": None,
                    "gbp": {"electricity": 0.3216, "gas": 0.0739}}),
    ("2026-07-01", {"eur": None,
                    "gbp": {"electricity": 0.3416, "gas": 0.0770}}),
]


def tariffs_for(date_iso):
    """
    Resolve the tariff period in force on a date. Dagger.

    The sterling side is a table of regulated all-in rates. The euro
    side is DERIVED at call time from the published Irish figures and
    the fetched semester rate, because a hard-coded euro row would go
    stale silently the moment the exchange rate moved - and it did,
    by about 4%, between the semester I first guessed and the one the
    anchors actually belong to.
    """
    if date_iso < TARIFF_HISTORY[0][0]:
        # REFUSE, do not clamp. The price panel reached back to April
        # 2024 the moment its floor came off, and 68 of those weeks
        # were being priced with August-2025 gas and electricity
        # tariffs - in both jurisdictions, silently, exactly the fault
        # just fixed for carbon. A missing figure is visible; a wrong
        # one is not. Extending TARIFF_HISTORY backwards is what
        # restores those weeks: UREGNI's briefing papers carry dated
        # annexes (SSE Airtricity to 2008, firmus to 2015) and SEAI
        # publishes ROI cent/kWh to 2008 Q1.
        return None
    row = TARIFF_HISTORY[0][1]
    for frm, rates in TARIFF_HISTORY:
        if date_iso >= frm:
            row = rates
    return {"gbp": row["gbp"],
            "eur": {"electricity": ie_eur("domestic_electricity", date_iso),
                    "gas": ie_eur("domestic_gas", date_iso)}}


def apply_ie_fx(feeds):
    """Set the euro conversion from the ECB semester mean and refresh
    the euro anchors that depend on it. Called once per run, before
    anything prices a week."""
    sem = ((feeds.get("ecb_fx") or {}).get("eur_gbp_semester") or {})
    rate = sem.get(IE_SEMESTER)
    if rate:
        IE_FX.update(rate=rate,
                     source=f"ECB semester mean {IE_SEMESTER}")
    else:
        IE_FX.update(rate=IE_FX_FALLBACK,
                     source="fallback (UNVERIFIED)")
        log(f"fx: WARNING {IE_SEMESTER} semester mean unavailable - "
            f"every Irish anchor is on the {IE_FX_FALLBACK} fallback")
    nd, used = nondom_for(today_utc().isoformat(), sem)
    ANCHORS["nondom_eur_per_kwh"] = nd["eur"]
    ANCHORS["nondom_gbp_per_kwh"] = nd["gbp"]
    log(f"fx: EUR/GBP {IE_FX['rate']} ({IE_FX['source']}); "
        f"ROI domestic {tariffs_for(HISTORY_START)['eur']}")
    log(f"nondom: live week on semester {used} "
        f"(latest published; REMM lags ~9 months) - "
        f"NI {nd['gbp']}, ROI {nd['eur']}")
    return IE_FX["rate"]


# ---------------------------------------------------------------- utilities

def log(*a):
    print(f"[{dt.datetime.now(dt.timezone.utc):%H:%M:%S}]", *a, flush=True)


def http_get(url, *, params=None, timeout=TIMEOUT, retries=RETRIES, headers=None):
    """GET with retries + exponential backoff + jitter. Raises on final failure."""
    last = None
    for attempt in range(1, retries + 1):
        try:
            r = requests.get(url, params=params, timeout=timeout,
                             headers=headers or UA)
            if r.status_code in (429, 500, 502, 503, 504):
                raise requests.HTTPError(f"{r.status_code} {r.reason}", response=r)
            r.raise_for_status()
            return r
        except requests.RequestException as e:
            last = e
            if attempt == retries:
                break
            wait = 2 ** attempt + random.uniform(0, 1.5)
            log(f"retry {attempt}/{retries - 1} after "
                f"{e.__class__.__name__} - sleeping {wait:.1f}s: {url[:80]}")
            time.sleep(wait)
    raise last


def today_utc():
    return dt.datetime.now(dt.timezone.utc).date()


def clip_days(day_values: dict) -> dict:
    """Drop future-dated keys (NESO lesson)."""
    cut = today_utc().isoformat()
    return {k: v for k, v in day_values.items() if k <= cut}


def trim_series(day_values: dict) -> dict:
    keep = (today_utc() - dt.timedelta(days=SERIES_KEEP_DAYS)).isoformat()
    return dict(sorted((k, v) for k, v in day_values.items() if k >= keep))


def autodetect_scale_to_gwh(values):
    med = statistics.median(abs(v) for v in values if v is not None) if values else 0
    if med > 1e6:
        return 1e-6, "kWh->GWh"
    if med > 1e3:
        return 1e-3, "MWh->GWh"
    return 1.0, "GWh"


def recency_status(latest_day: str | None, fresh_within_days: int) -> str:
    if not latest_day:
        return "stale"
    age = (today_utc() - dt.date.fromisoformat(latest_day)).days
    return "ok" if age <= fresh_within_days else "lagging"


def ddmmyyyy_to_iso(s: str) -> str | None:
    m = re.match(r"(\d{1,2})[/.-](\d{1,2})[/.-](\d{4})", s.strip())
    if not m:
        return None
    return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"


def find_date_in_text(s: str) -> str | None:
    """dd/mm/yyyy anywhere inside a longer string (bulletin title cells)."""
    m = re.search(r"(\d{1,2})[/.-](\d{1,2})[/.-](\d{4})", s)
    if not m:
        return None
    return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"


def prev_series(feed: str, *keys) -> dict:
    """Previously stored series for cross-run history accumulation."""
    node = PREVIOUS_FEEDS.get(feed, {})
    for k in keys:
        node = node.get(k, {}) if isinstance(node, dict) else {}
    return dict(node) if isinstance(node, dict) else {}


def _num(s) -> float | None:
    """Parse a number that may use a decimal comma (SEMOpx CSV convention)."""
    try:
        return float(str(s).strip().replace(",", "."))
    except (TypeError, ValueError):
        return None


# ------------------------------------------------- pure parsers (unit tested)

def extract_chart_data_arrays(page_html: str) -> list:
    """Bracket-matched '"data":[[...]]' chart payloads, entity-unescaped."""
    text = html_mod.unescape(page_html)
    out = []
    for m in re.finditer(r'"data"\s*:\s*(\[\[)', text):
        i = m.start(1)
        depth = 0
        for j in range(i, min(len(text), i + 2_000_000)):
            if text[j] == "[":
                depth += 1
            elif text[j] == "]":
                depth -= 1
                if depth == 0:
                    try:
                        out.append(json.loads(text[i:j + 1]))
                    except json.JSONDecodeError:
                        pass
                    break
    return out


def parse_ccni_series(page_html: str) -> dict:
    """{"300l": {iso: gbp}, ...} from litre-labelled embedded charts."""
    series = {"300l": {}, "500l": {}, "900l": {}}
    for arr in extract_chart_data_arrays(page_html):
        if not arr or not isinstance(arr[0], list):
            continue
        header = [str(c).lower() for c in arr[0]]
        if not any("litre" in h for h in header):
            continue
        cols = {}
        for idx, h in enumerate(header):
            for litres in ("300", "500", "900"):
                if litres in h:
                    cols[idx] = f"{litres}l"
        for row in arr[1:]:
            if not row:
                continue
            d = ddmmyyyy_to_iso(str(row[0]))
            if not d:
                continue
            for idx, key in cols.items():
                try:
                    v = float(row[idx])
                except (TypeError, ValueError, IndexError):
                    continue
                series[key][d] = round(v, 2)
    return series


def resolve_oil_bulletin_url(page_html: str, with_tax: bool = True) -> str | None:
    """Weekly prices xlsx, with or without taxes - unquote before matching."""
    for m in re.finditer(r'href="([^"]+)"', page_html):
        u = m.group(1)
        decoded = urllib.parse.unquote(u).lower()
        if ".xlsx" not in decoded:
            continue
        has_without = "without" in decoded
        has_with = re.search(r"with[ _]tax", decoded) is not None
        hit = (has_with and not has_without) if with_tax \
            else (has_without and "tax" in decoded)
        if hit:
            return u if u.startswith("http") else "https://energy.ec.europa.eu" + u
    return None


def parse_bulletin_rows(rows) -> tuple:
    """
    One-week-snapshot layout confirmed live (14 Jul 2026): a title cell
    carries the bulletin date; a header row names products ('...Heating...'
    at some column); country rows carry values with no per-row dates.
    Returns (iso_date | None, ireland_heating_value | None).
    Tolerates per-row dates too, should the layout ever grow them.
    """
    bulletin_date, idx_heat, value = None, None, None
    for row in rows:
        cells = list(row)
        strs = [str(c) if c is not None else "" for c in cells]
        joined = " ".join(strs).lower()

        if bulletin_date is None:
            for c in cells:
                if isinstance(c, dt.datetime):
                    bulletin_date = c.date().isoformat()
                    break
                if isinstance(c, dt.date):
                    bulletin_date = c.isoformat()
                    break
                iso = find_date_in_text(str(c)) if c is not None else None
                if iso:
                    bulletin_date = iso
                    break

        if idx_heat is None and ("heating" in joined or "chauffage" in joined):
            for i, c in enumerate(strs):
                if "heating" in c.lower() or "chauffage" in c.lower():
                    idx_heat = i
                    break
            continue

        if idx_heat is not None and value is None and (
                "ireland" in joined
                or (strs and strs[0].strip().upper() in ("IE", "EI"))):
            row_date = None
            for c in cells:
                if isinstance(c, (dt.datetime, dt.date)):
                    row_date = (c.date() if isinstance(c, dt.datetime)
                                else c).isoformat()
                    break
            try:
                v = float(cells[idx_heat])
            except (TypeError, ValueError, IndexError):
                continue
            value = round(v, 2)
            if row_date:
                bulletin_date = row_date
    return bulletin_date, value


def parse_bulletin_history_rows(rows, colpat=r"^IE_.*heating"):
    """
    EC oil bulletin price-history workbook - WIDE format (verified from a
    live run dump, 16 Jul 2026): one row per date in column 0, countries
    spread across columns headed e.g. 'IE_price_with_tax_heating_oil'.
    Falls back to long-table / country-block scanning if no wide header
    is found. Returns {iso_date: value}.
    """
    pat = re.compile(colpat, re.I)
    buffered = []
    idx = None
    for row in rows:
        cells = list(row)
        buffered.append(cells)
        for i, c in enumerate(cells):
            if c is not None and pat.match(str(c).strip()):
                idx = i
                break
        if idx is not None:
            break
    series = {}
    if idx is not None:
        def take(cells):
            if not cells:
                return
            d = cells[0]
            if isinstance(d, dt.datetime):
                d = d.date()
            if not isinstance(d, dt.date):
                return
            try:
                v = float(cells[idx])
            except (TypeError, ValueError, IndexError):
                return
            if 300 <= v <= 3000:
                series[d.isoformat()] = round(v, 2)
        for cells in rows:          # continue the iterator
            take(list(cells))
        return series
    # fallback: long-table / country-block layouts
    return _parse_history_longtable(buffered)


def _parse_history_longtable(rows) -> dict:
    EU = {"belgium", "bulgaria", "czechia", "czech republic", "denmark",
          "germany", "estonia", "greece", "spain", "france", "croatia",
          "italy", "cyprus", "latvia", "lithuania", "luxembourg", "hungary",
          "malta", "netherlands", "austria", "poland", "portugal",
          "romania", "slovenia", "slovakia", "finland", "sweden",
          "united kingdom", "uk"}
    idx_heat, series, in_ireland = None, {}, False
    for row in rows:
        cells = list(row)
        strs = [str(c) if c is not None else "" for c in cells]
        joined = " ".join(strs).lower()
        if idx_heat is None and ("heating" in joined or "chauffage" in joined):
            for i, c in enumerate(strs):
                if "heating" in c.lower() or "chauffage" in c.lower():
                    idx_heat = i
                    break
            continue
        if idx_heat is None:
            continue
        is_ireland_row = ("ireland" in joined
                          or (strs and strs[0].strip().upper() in ("IE", "EI")))
        if is_ireland_row:
            in_ireland = True
        elif any(c in joined for c in EU):
            in_ireland = False
        if not (is_ireland_row or in_ireland):
            continue
        row_date = None
        for c in cells:
            if isinstance(c, dt.datetime):
                row_date = c.date().isoformat()
                break
            if isinstance(c, dt.date):
                row_date = c.isoformat()
                break
            iso = find_date_in_text(str(c)) if c is not None else None
            if iso:
                row_date = iso
                break
        if row_date is None:
            continue
        try:
            v = float(cells[idx_heat])
        except (TypeError, ValueError, IndexError):
            continue
        if 300 <= v <= 3000:
            series[row_date] = round(v, 2)
    return series


def semopx_hour_keys(stamps, trade_day):
    """
    Delivery stamps -> Irish local-clock hour keys, ANCHORED on the
    trade day rather than converted by an assumed timezone.

    Two guesses have now been wrong. The stamps carry a Z suffix, so
    the first pass treated them as UTC and added an hour in summer;
    the span still came out one hour below local midnight, which
    pointed at CET - and a third guess would have been a third
    coin-toss. The document does not need decoding: a trade day runs
    00:00 to 23:00 local by definition, and the resource name states
    which day it is. So the offset is measured from the document -
    first stamp against local midnight - and applied to the rest.

    Returns [] if the trade day is unknown or nothing parses, so a
    document that cannot be anchored contributes nothing rather than
    contributing something misaligned.
    """
    parsed = []
    for t in stamps:
        m = re.match(r"(20\d\d)-(\d\d)-(\d\d)T(\d\d)", str(t))
        parsed.append(dt.datetime(*(int(g) for g in m.groups()))
                      if m else None)
    real = [x for x in parsed if x]
    if not real or not trade_day:
        return []
    try:
        anchor = dt.datetime.strptime(trade_day, "%Y-%m-%d")
    except ValueError:
        return []
    offset = anchor - min(real)
    return [(x + offset).strftime("%Y-%m-%dT%H") if x else None
            for x in parsed]


def parse_semopx_csv(text: str, trade_day: str = None) -> dict:
    """
    SEMOpx MarketResult CSV, format confirmed live (14 Jul 2026):
    semicolon-delimited, decimal commas, sections -
        Auction;SEM-DA
        FX rates
        EUR;GBP;0,85506627
        Market;NI-DA
        Index prices;30;EUR
        <row of ISO delivery timestamps>
        <row of prices>
    Returns {"fx_eur_gbp", "day", "auction", "markets"}.
    """
    fx, day, auction = None, None, None
    markets: dict = {}
    series: dict = {}
    stamps: dict = {}
    market, currency, expect_series = None, None, False
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split(";")]
        key = parts[0].lower()

        if key == "auction" and len(parts) > 1:
            auction = parts[1]
            continue
        if key == "market" and len(parts) > 1:
            market = parts[1]
            expect_series = False
            continue
        if key.startswith("index prices"):
            currency = parts[-1].upper() if len(parts) >= 2 else "EUR"
            expect_series = True
            continue
        if key == "eur" and len(parts) >= 3 and parts[1].upper() == "GBP":
            fx = _num(parts[2])
            continue
        if expect_series:
            if re.match(r"20\d\d-\d\d-\d\dT", parts[0]):
                if day is None:
                    day = parts[0][:10]
                # The delivery stamps were being read and discarded.
                # They are the only thing that makes a price hourly,
                # and B.2.3 asks what price applied in ONE hour.
                stamps[(market, currency)] = list(parts)
                continue
            nums = [n for n in (_num(p) for p in parts) if n is not None]
            if nums and market and currency:
                markets.setdefault(market, {}).setdefault(
                    currency, []).extend(nums)
                ts = stamps.get((market, currency)) or []
                for k, v in zip(semopx_hour_keys(ts, trade_day or day),
                                nums):
                    if k:
                        series.setdefault(market, {}).setdefault(
                            currency, {})[k] = v
                expect_series = False
    return {"fx_eur_gbp": fx, "day": day, "auction": auction,
            "markets": markets, "series": series}


def parse_gni_series(series_list) -> dict:
    """
    Pure parser for the GNI gasconsumption JSON API. Input: list of series
    objects {name, location, group, ..., data: [[unix_ms, value], ...]}.
    Output: {location: {iso_date: value}}. Keys off `location`, tolerates
    missing fields, skips unparseable points.
    """
    out = {}
    for s in series_list or []:
        loc = (s or {}).get("location")
        if not loc:
            continue
        pts = {}
        for pair in s.get("data") or []:
            try:
                ms, val = pair[0], pair[1]
                d = dt.datetime.fromtimestamp(
                    ms / 1000.0, tz=dt.timezone.utc).date().isoformat()
                pts[d] = float(val)
            except (TypeError, ValueError, IndexError, OSError):
                continue
        if pts:
            out.setdefault(loc, {}).update(pts)
    return out




def parse_gb_oil_page(text: str) -> tuple:
    """
    BoilerJuice prices page - server-rendered sentence:
      'Our average heating oil price for today, Saturday 25th March 2017
       is 40.32 pence per litre (inc. VAT)'
    Returns (iso_date | None, pence_per_litre | None). Date falls back to
    None if unparseable - caller may substitute the run date.
    """
    m = re.search(
        r"average (?:kerosene|heating oil) price for today[^0-9]*?"
        r"(?:(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)\s+(\d{4}))?"
        r"[^0-9]*?is\s*(\d{1,3}\.\d{1,2})\s*pence per litre",
        text, re.I | re.S)
    if not m:
        # modern template (observed 18 Jul 2026): dated sentence gone,
        # but the current price is server-rendered beside the chart -
        # digits, optional markup, then 'pence per litre'. Undated;
        # the caller stamps the run date. Fossils cannot reach here:
        # they carry the dated 2021 sentence, matched above and then
        # rejected by the caller's freshness gate.
        m2 = re.search(
            r"(\d{2,3}\.\d{1,2})\s*(?:</?[a-z][^>]*>\s*|&nbsp;)*"
            r"pence per litre", text, re.I)
        if m2:
            v = float(m2.group(1))
            if 40.0 <= v <= 150.0:
                return None, v
        return None, None
    day, month_name, year, ppl = m.group(1), m.group(2), m.group(3), m.group(4)
    iso = None
    if day and month_name and year:
        months = {"january": 1, "february": 2, "march": 3, "april": 4,
                  "may": 5, "june": 6, "july": 7, "august": 8,
                  "september": 9, "october": 10, "november": 11,
                  "december": 12}
        mo = months.get(month_name.lower())
        if mo:
            iso = f"{int(year):04d}-{mo:02d}-{int(day):02d}"
    return iso, float(ppl)


# ---------------------------------------------------------------- feeds

def parse_eirgrid_rows(payload, field="SYSTEM_DEMAND", daily="gwh",
                       min_intervals=48):
    """
    /api/chart/ response -> {iso_date: value}. Rows for the requested
    field are aggregated per day: daily="gwh" averages MW and converts
    (mean x 24 / 1000); daily="mean" returns the plain daily mean (used
    for CO2 intensity, gCO2/kWh). Nulls and other fields are ignored;
    days with fewer than min_intervals observations are dropped as
    incomplete. Pure, unit tested.
    """
    acc = {}
    for row in (payload or {}).get("Rows", []):
        if row.get("FieldName") != field or row.get("Value") is None:
            continue
        et = str(row.get("EffectiveTime", ""))
        try:
            iso = dt.datetime.strptime(et.split()[0], "%d-%b-%Y")\
                    .date().isoformat()
        except (ValueError, IndexError):
            iso = find_date_in_text(et)
        if not iso:
            continue
        s, n = acc.get(iso, (0.0, 0))
        acc[iso] = (s + float(row["Value"]), n + 1)
    if daily == "mean":
        return {d: round(s / n, 1)
                for d, (s, n) in acc.items() if n >= min_intervals}
    return {d: round(s / n * 24.0 / 1000.0, 2)
            for d, (s, n) in acc.items() if n >= min_intervals}


def feed_eirgrid():
    """
    Smart Grid Dashboard via the /api/chart/ endpoint (probed 18 Jul
    2026). Daily electricity demand in GWh for ALL / NI / ROI, merged
    across runs. Today is always incomplete and excluded by the parser's
    48-interval rule; each run therefore backfills the last full days.
    """
    def dmy(d):
        return d.strftime("%d-%b-%Y")

    out = {"source": "EirGrid Smart Grid Dashboard (/api/chart/)",
           "demand_gwh_daily": {}}
    end = today_utc()
    start = end - dt.timedelta(days=8)
    for region, key in (("ALL", "island"), ("NI", "ni"), ("ROI", "roi")):
        try:
            payload = http_get(EIRGRID_ENDPOINT, params={
                "region": region, "chartType": "demand",
                "dateRange": "month", "dateFrom": dmy(start),
                "dateTo": dmy(end), "areas": "demandactual",
            }, timeout=90).json()
        except Exception as e:
            log(f"eirgrid: {region} span fetch failed "
                f"({e.__class__.__name__}: {e})")
            payload = {}
        got = parse_eirgrid_rows(payload)
        if not got:
            rows = (payload or {}).get("Rows", [])
            log(f"eirgrid: {region} span returned {len(rows)} rows, "
                f"0 complete days - falling back to per-day calls")
            got = {}
            for i in range(1, 8):
                d = end - dt.timedelta(days=i)
                try:
                    pl = http_get(EIRGRID_ENDPOINT, params={
                        "region": region, "chartType": "demand",
                        "dateRange": "day", "dateFrom": dmy(d),
                        "dateTo": dmy(d), "areas": "demandactual",
                    }, timeout=60).json()
                    got.update(parse_eirgrid_rows(pl))
                except Exception as e:
                    log(f"eirgrid: {region} {d} failed ({e})")
        series = prev_series("eirgrid", "demand_gwh_daily", key)
        series.update(got)
        out["demand_gwh_daily"][key] = trim_series(series)
        log(f"eirgrid: {region} {len(got)} days this run, "
            f"{len(out['demand_gwh_daily'][key])} retained")

    # CO2 intensity - schema confirmed in the 18 Jul 2026 run log
    # (CO2_INTENSITY rows, same shape). Daily mean gCO2/kWh, island,
    # merged across runs; soft.
    try:
        pl = http_get(EIRGRID_ENDPOINT, params={
            "region": "ALL", "chartType": "co2",
            "dateRange": "month", "dateFrom": dmy(end - dt.timedelta(days=150)),
            "dateTo": dmy(end), "areas": "co2intensity"},
            timeout=90).json()
        got = parse_eirgrid_rows(pl, field="CO2_INTENSITY",
                                 daily="mean", min_intervals=40)
        if not got:
            for i in range(1, 8):
                d = end - dt.timedelta(days=i)
                pl = http_get(EIRGRID_ENDPOINT, params={
                    "region": "ALL", "chartType": "co2",
                    "dateRange": "day", "dateFrom": dmy(d),
                    "dateTo": dmy(d), "areas": "co2intensity"},
                    timeout=60).json()
                got.update(parse_eirgrid_rows(
                    pl, field="CO2_INTENSITY", daily="mean",
                    min_intervals=40))
        ser = prev_series("eirgrid", "co2_intensity_g_per_kwh")
        ser.update(got)
        out["co2_intensity_g_per_kwh"] = trim_series(ser)
        log(f"eirgrid: co2 intensity {len(got)} days this run, "
            f"{len(out['co2_intensity_g_per_kwh'])} retained")
    except Exception as e:
        log(f"eirgrid: co2 parse failed ({e.__class__.__name__}: {e})")

    isl = out["demand_gwh_daily"].get("island") or {}
    out["latest_day"] = max(isl) if isl else None
    return out, recency_status(out["latest_day"], 4)


def odh26_from_hourly(payload, names, weights, base_c=26.0):
    """
    Population-weighted overheating degree-hours per day from Open-Meteo
    hourly payloads: sum over hours of max(0, T - 26), weighted across
    stations, keyed by UTC date. Pure, unit tested.
    """
    locs = payload if isinstance(payload, list) else [payload]
    per_day = {}
    for name, loc in zip(names, locs):
        w = weights.get(name, 0.0)
        hh = loc.get("hourly", {})
        for ts, t in zip(hh.get("time", []), hh.get("temperature_2m", [])):
            if t is None:
                continue
            d = ts[:10]
            per_day[d] = per_day.get(d, 0.0) + w * max(0.0, t - base_c)
    return {d: round(v, 2) for d, v in per_day.items()}


# MIDLANDS PROBE - RAN AND RETIRED, 14 Aug 2026. All seven weighted
# stations sit on or near the coast, so the island series is entirely
# maritime, and Ireland's interior runs colder on winter nights. The
# worry was that the weighted HDD is biased low, and with it the
# space-heat share of every week.
#
# Measured rather than argued: Athlone was fetched alongside the seven
# at the midland counties' share of island population (Longford,
# Westmeath, Offaly, Laois ~ 4%) over 1,151 days. Island HDD moved
# 4.65 -> 4.66, +0.28%; winter alone 8.90 -> 8.92, +0.27%. Immaterial,
# so the all-coastal set stands and the probe is retired rather than
# carried for ever.
#
# It also left a station in the fetch that STATIONS did not know
# about, which broke the ODH26 loop with a KeyError - a reason to
# retire a probe once it has answered rather than leave it running.
PROBE_STATIONS: dict = {}


def feed_hdd():
    """Open-Meteo, batched; forecast tail optional (degrades to lagging)."""
    names = list(STATIONS) + list(PROBE_STATIONS)
    _all = {**STATIONS, **PROBE_STATIONS}
    lats = ",".join(str(_all[n][0]) for n in names)
    lons = ",".join(str(_all[n][1]) for n in names)

    def unpack(payload):
        locs = payload if isinstance(payload, list) else [payload]
        per_station = {}
        for name, loc in zip(names, locs):
            d = loc.get("daily", {})
            per_station[name] = {
                day: t for day, t in zip(d.get("time", []),
                                         d.get("temperature_2m_mean", []))
                if t is not None
            }
        return per_station

    arch = unpack(http_get(
        "https://archive-api.open-meteo.com/v1/archive", params={
            "latitude": lats, "longitude": lons,
            "start_date": (today_utc()
                           - dt.timedelta(days=SERIES_KEEP_DAYS)).isoformat(),
            "end_date": today_utc().isoformat(),
            "daily": "temperature_2m_mean", "timezone": "UTC",
        }, timeout=120).json())

    tail_ok = True
    try:
        tail = unpack(http_get(
            "https://api.open-meteo.com/v1/forecast", params={
                "latitude": lats, "longitude": lons, "past_days": 10,
                "forecast_days": 1, "daily": "temperature_2m_mean",
                "timezone": "UTC",
            }).json())
        for n in names:
            for d, t in tail.get(n, {}).items():
                arch[n].setdefault(d, t)
    except Exception as e:
        tail_ok = False
        log(f"hdd: forecast tail unavailable ({e.__class__.__name__}) - "
            "continuing archive-only, expect 'lagging'")

    daily_by_station = {n: clip_days(v) for n, v in arch.items()}

    def weighted_temp(subset):
        """Population-weighted daily mean AIR TEMPERATURE.

        HDD is max(0, base - t), so it throws the temperature away
        above the base - and the COP engine needs the temperature in
        summer as much as in winter, because that is when hot water is
        the whole load. Same weights, same days, one line of extra
        state.
        """
        wsum = sum(STATIONS[n][2] for n in subset)
        days = set.intersection(*(set(daily_by_station[n]) for n in subset))
        return {d: round(sum(daily_by_station[n][d] * STATIONS[n][2]
                             for n in subset) / wsum, 2)
                for d in sorted(days)}

    def weighted(subset):
        wsum = sum(STATIONS[n][2] for n in subset)
        days = set.intersection(*(set(daily_by_station[n]) for n in subset))
        return {
            d: round(max(0.0, HDD_BASE_C - sum(
                daily_by_station[n][d] * STATIONS[n][2] for n in subset) / wsum), 2)
            for d in sorted(days)
        }

    weighted_names = [n for n in names if n in STATIONS]
    roi = [n for n in weighted_names if STATIONS[n][3] == "ROI"]
    ni = [n for n in weighted_names if STATIONS[n][3] == "NI"]
    out = {
        "hdd_island": trim_series(weighted(weighted_names)),
        "temp_island": trim_series(weighted_temp(weighted_names)),
        "temp_roi": trim_series(weighted_temp(roi)),
        "temp_ni": trim_series(weighted_temp(ni)),
        "hdd_roi": trim_series(weighted(roi)),
        "hdd_ni": trim_series(weighted(ni)),
        "base_c": HDD_BASE_C,
        "forecast_tail": tail_ok,
        "weights_note": ("Population weights are current Causeway Energies "
                         "estimates - challenge and input welcome at "
                         "contact@causewaygt.com"),
        "source": "ERA5 via Open-Meteo, population-weighted HDD",
    }

    # --- midlands probe (log-only, changes nothing; empty once answered)
    try:
        if "Athlone" in daily_by_station and daily_by_station["Athlone"]:
            share = PROBE_STATIONS["Athlone"][2]
            days = set(daily_by_station["Athlone"])
            for n in weighted_names:
                days &= set(daily_by_station[n])
            days = sorted(days)
            if len(days) > 300:
                wsum = sum(STATIONS[n][2] for n in weighted_names)
                base, withm = [], []
                for d in days:
                    t7 = sum(daily_by_station[n][d] * STATIONS[n][2]
                             for n in weighted_names) / wsum
                    t8 = ((1 - share) * t7
                          + share * daily_by_station["Athlone"][d])
                    base.append(max(0.0, HDD_BASE_C - t7))
                    withm.append(max(0.0, HDD_BASE_C - t8))
                a, b = sum(base) / len(base), sum(withm) / len(withm)
                pct = 100 * (b - a) / a if a else 0.0
                # Winter is what matters: a summer difference cannot
                # move a space-heat share that is already near zero.
                wi = [i for i, d in enumerate(days)
                      if d[5:7] in ("12", "01", "02")]
                wa = sum(base[i] for i in wi) / max(len(wi), 1)
                wb = sum(withm[i] for i in wi) / max(len(wi), 1)
                wpct = 100 * (wb - wa) / wa if wa else 0.0
                log(f"hdd: MIDLANDS PROBE over {len(days)} days - island "
                    f"HDD {a:.2f} without Athlone, {b:.2f} with it at "
                    f"{share:.0%} ({pct:+.2f}%); winter only {wa:.2f} -> "
                    f"{wb:.2f} ({wpct:+.2f}%)")
                log("hdd:   log-only. Under ~1% the all-coastal station "
                    "set is fine as it stands; more than that and the "
                    "weighted series needs a midlands member, which "
                    "would move every HDD-shaped figure on the site")
    except Exception as exc:
        log(f"hdd: midlands probe failed ({exc.__class__.__name__}) - "
            "log-only, the feed is unaffected")

    # ODH26 groundwork: hourly overheating-degree-hours (base 26 C),
    # population-weighted, trailing 60 days per run, merged across runs.
    # Collected for the future comfort/cooling metric; not yet displayed.
    # Soft: any failure leaves the hdd feed intact.
    try:
        hp = http_get(
            "https://archive-api.open-meteo.com/v1/archive", params={
                "latitude": lats, "longitude": lons,
                "start_date": (today_utc()
                               - dt.timedelta(days=60)).isoformat(),
                "end_date": today_utc().isoformat(),
                "hourly": "temperature_2m", "timezone": "UTC",
            }, timeout=180).json()
        odh_new = odh26_from_hourly(
            # weighted_names, not names: a probe station is fetched
            # but is not in STATIONS, and passing it here raised a
            # KeyError that silently stopped ODH collection on
            # 14 Aug 2026. Anything downstream of the fetch reads the
            # WEIGHTED set.
            hp, weighted_names,
            {n: STATIONS[n][2] for n in weighted_names})
        odh = prev_series("hdd", "odh26_island")
        odh.update(odh_new)
        out["odh26_island"] = trim_series(odh)
        log(f"hdd: odh26 {len(odh_new)} days this run, "
            f"{len(out['odh26_island'])} retained, "
            f"nonzero {sum(1 for v in out['odh26_island'].values() if v>0)}")
    except Exception as e:
        log(f"hdd: odh26 groundwork skipped ({e.__class__.__name__}: {e})")
    latest = max(out["hdd_island"] or {"": None})
    out["latest_day"] = latest or None
    return out, recency_status(out["latest_day"], 3 if tail_ok else 7)


def feed_ecb_fx():
    r = http_get("https://www.ecb.europa.eu/stats/eurofxref/eurofxref-daily.xml")
    root = ElementTree.fromstring(r.content)
    ns = {"e": "http://www.ecb.int/vocabulary/2002-08-01/eurofxref"}
    cube_day = root.find(".//e:Cube[@time]", ns)
    rate = None
    for c in cube_day.findall("e:Cube", ns):
        if c.get("currency") == "GBP":
            rate = float(c.get("rate"))
    if rate is None:
        avail = [c.get("currency") for c in cube_day.findall("e:Cube", ns)]
        log("ecb_fx: GBP missing - available currencies:", avail)
        raise ValueError("GBP not in ECB daily cube")
    out = {"eur_gbp": rate, "gbp_eur": round(1 / rate, 5),
           "rate_date": cube_day.get("time"), "latest_day": cube_day.get("time"),
           "source": "ECB euro foreign exchange reference rates (daily)"}
    # 90-day history for per-week pricing of the back-look. Soft.
    try:
        r90 = http_get("https://www.ecb.europa.eu/stats/eurofxref/"
                       "eurofxref-hist-90d.xml")
        root90 = ElementTree.fromstring(r90.content)
        ser = prev_series("ecb_fx", "eur_gbp_daily")
        for cube in root90.findall(".//e:Cube[@time]", ns):
            for c in cube.findall("e:Cube", ns):
                if c.get("currency") == "GBP":
                    ser[cube.get("time")] = float(c.get("rate"))
        sem = dict(prev_series("ecb_fx", "eur_gbp_semester"))
        sem.update(semester_means(ser))
        if not ser or min(ser) > HISTORY_START or IE_SEMESTER not in sem:
            # Deep backfill to the back-look floor. Also fires when the
            # Irish anchor semester is not yet covered - the daily
            # series starts at HISTORY_START, which can sit inside the
            # semester and leave the mean short of its day count.
            import zipfile
            zr = http_get("https://www.ecb.europa.eu/stats/eurofxref/"
                          "eurofxref-hist.zip", timeout=120)
            with zipfile.ZipFile(io.BytesIO(zr.content)) as z:
                txt = z.read(z.namelist()[0]).decode("utf8")
            lines = [l for l in txt.splitlines() if l.strip()]
            cols = lines[0].split(",")
            gi = cols.index("GBP")
            added = 0
            # The zip is the FULL series back to 1999, so semester
            # means come off it whole rather than off the trimmed
            # daily window - which is the point of fetching it here.
            whole = {}
            for line in lines[1:]:
                parts = line.split(",")
                d = parts[0].strip()
                try:
                    whole[d] = float(parts[gi])
                except (ValueError, IndexError):
                    pass
                if d < HISTORY_START or d in ser:
                    continue
                try:
                    ser[d] = float(parts[gi])
                    added += 1
                except (ValueError, IndexError):
                    continue
            sem.update(semester_means(whole))
            log(f"ecb_fx: deep history backfill +{added} days, "
                f"{len(whole)} days seen, {len(sem)} semester means")
        out["eur_gbp_daily"] = trim_series(ser)
        out["eur_gbp_semester"] = sem
        log(f"ecb_fx: history {len(out['eur_gbp_daily'])} days retained, "
            f"{len(sem)} semester means; {IE_SEMESTER}="
            f"{sem.get(IE_SEMESTER, 'MISSING')}")
    except Exception as e:
        out["eur_gbp_daily"] = prev_series("ecb_fx", "eur_gbp_daily")
        out["eur_gbp_semester"] = prev_series("ecb_fx", "eur_gbp_semester")
        log(f"ecb_fx: history skipped ({e.__class__.__name__})")
    return out, recency_status(out["latest_day"], 5)


def feed_gni_ckan():
    """data.gov.ie CKAN - GNI daily demand by sector, CC BY 4.0, quarterly."""
    pkg = http_get("https://data.gov.ie/api/3/action/package_search",
                   params={"q": "daily gas demand", "rows": 10}).json()
    results = pkg.get("result", {}).get("results", [])
    resource_url = None
    for ds in results:
        if "gas networks ireland" not in json.dumps(
                ds.get("organization", {})).lower() \
           and "gas" not in ds.get("title", "").lower():
            continue
        for res in ds.get("resources", []):
            if res.get("format", "").upper() == "CSV" \
               and "demand" in (res.get("name", "") + ds.get("title", "")).lower():
                resource_url = res.get("url")
                break
        if resource_url:
            break
    if not resource_url:
        log("gni_ckan: no CSV resource matched - datasets found:",
            [d.get("title") for d in results])
        raise ValueError("CKAN resolution failed")
    log("gni_ckan: resolved", resource_url)

    csv_text = http_get(resource_url).text
    lines = [ln for ln in csv_text.splitlines() if ln.strip()]
    header = [h.strip().strip('"') for h in lines[0].split(",")]
    log("gni_ckan: header:", header)

    def col(*needles):
        for i, h in enumerate(header):
            hl = h.lower()
            if all(n in hl for n in needles):
                return i
        return None

    cols = {
        "ndm": col("ndm") if col("ndm") is not None else col("non", "daily"),
        "dm": None,
        "ldm": col("ldm"),
        "power": col("power"),
        "total_roi": col("total"),
    }
    for i, h in enumerate(header):
        if h.lower().startswith("daily metered"):
            cols["dm"] = i
            break
    if cols["ndm"] is None:
        raise ValueError(f"gni_ckan: NDM column not found in {header}")

    series = {k: {} for k in cols if cols[k] is not None}
    i_date = col("date") if col("date") is not None else 0
    for ln in lines[1:]:
        parts = [p.strip().strip('"') for p in ln.split(",")]
        raw_d = parts[i_date][:10]
        d = ddmmyyyy_to_iso(raw_d) or raw_d.replace("/", "-")
        try:
            d = dt.date.fromisoformat(d).isoformat()
        except ValueError:
            continue
        for k, i in cols.items():
            if i is None:
                continue
            try:
                series[k][d] = float(parts[i])
            except (ValueError, IndexError):
                continue

    scale, unit_note = autodetect_scale_to_gwh(list(series["ndm"].values()))
    out = {"unit_detection": unit_note,
           "latest_day": max(series["ndm"]) if series["ndm"] else None,
           "source": ("Gas Networks Ireland via data.gov.ie, CC BY 4.0 - "
                      "quarterly refresh, calibration series")}
    for k, vals in series.items():
        out[f"{k}_gwh"] = clip_days(
            {d: round(v * scale, 2) for d, v in vals.items()})
    return out, recency_status(out["latest_day"], 100)


def feed_gni_live():
    """
    GNI Data Transparency JSON API - probed 14 Jul 2026, daily/hourly/
    monthly all 200. The CSV export route 503s and is not used. Window per
    call unknown: anchor four dates a month apart, log the observed span,
    merge across runs. Values arrive in kWh by default - unit autodetected.
    JURISDICTION FLAG: DM/NDM carry no ROI prefix while LDM and Power Gen
    do - whether unprefixed series include NI exits is unconfirmed; the
    space-heat regression stays on the confirmed-ROI gni_ckan series until
    resolved.
    """
    base = "https://www.gasnetworks.ie/api/v1/gasconsumption"
    raw = {}
    for back in range(0, 92, 7):   # API window ~8 trailing days
        time.sleep(0.3)
        anchor = (today_utc() - dt.timedelta(days=back)).isoformat()
        try:
            body = http_get(base, params={
                "date": anchor, "frequency": "daily", "unit": ""}).json()
        except Exception as e:
            log(f"gni_live: anchor {anchor} failed - {e.__class__.__name__}: {e}")
            continue
        if not isinstance(body, list):
            log("gni_live: unexpected shape for", anchor, "-", str(body)[:300])
            continue
        parsed = parse_gni_series(body)
        if back == 0:
            log("gni_live: locations seen:", sorted(parsed))
        ndm = parsed.get("NDM", {})
        log(f"gni_live: anchor {anchor} -> NDM {len(ndm)} pts "
            f"{min(ndm) if ndm else '-'}..{max(ndm) if ndm else '-'}")
        for loc, pts in parsed.items():
            raw.setdefault(loc, {}).update(pts)
    if not raw:
        raise ValueError("gni_live: no series parsed from any anchor")

    spans = {loc: (min(p), max(p), len(p)) for loc, p in raw.items()}
    log("gni_live: spans:", spans)

    # Jurisdiction signal (standing reminder, 24 Jul 2026): if Total LDM
    # minus ROI LDM is ~zero, "Total" is ROI-total naming and the
    # unprefixed DM/NDM series are very likely ROI-scoped; a material,
    # stable difference means the feed carries NI exits. Logged every
    # run so a convention change announces itself.
    tot, roi = raw.get("Total LDM", {}), raw.get("ROI LDM", {})
    common = sorted(set(tot) & set(roi))[-28:]
    if len(common) >= 7:
        diffs = [tot[d] - roi[d] for d in common]
        mean_d = sum(diffs) / len(diffs)
        mean_t = sum(tot[d] for d in common) / len(common)
        pct = 100 * mean_d / mean_t if mean_t else 0.0
        log(f"gni_live: jurisdiction check - Total LDM minus ROI LDM "
            f"mean {mean_d:.3f} ({pct:.2f}% of Total) over "
            f"{len(common)}d "
            + ("-> ~zero: unprefixed series read as ROI-scoped"
               if abs(pct) < 1.0 else
               "-> MATERIAL: feed appears to include NI exits - "
               "review the jurisdiction note"))

    ndm_vals = list(raw.get("NDM", {}).values())
    scale, unit_note = autodetect_scale_to_gwh(
        ndm_vals or [v for p in raw.values() for v in p.values()])

    def keyname(loc):
        return loc.lower().replace(" ", "_")

    out = {"unit_detection": unit_note,
           "jurisdiction_note": ("DM/NDM carry no ROI prefix while LDM and "
                                 "Power Gen do - whether unprefixed series "
                                 "include NI exits is unconfirmed; regression "
                                 "remains on the confirmed-ROI calibration "
                                 "series until resolved"),
           "source": ("Gas Networks Ireland Data Transparency - "
                      "gasconsumption API, daily by market sector")}
    latest = None
    for loc, pts in raw.items():
        merged = prev_series("gni_live", f"{keyname(loc)}_gwh")
        merged.update({d: round(v * scale, 3) for d, v in pts.items()})
        merged = trim_series(clip_days(merged))
        out[f"{keyname(loc)}_gwh"] = merged
        if merged:
            latest = max(latest or "", max(merged))
    out["latest_day"] = latest
    return out, recency_status(latest, 3)


def semopx_trade_day(item):
    """Trade day from a SEMOpx resource name, e.g.
    MarketResult_SEM-DA_PWR-MRC-D+1_20260806100000_... -> 2026-08-06.

    The first probe sliced this positionally and cut mid-timestamp,
    which turned every day tag into nonsense and made the verdict line
    say the opposite of what the data showed. Match the field, do not
    count characters into it."""
    m = re.search(r"_(20\d{6})\d{6}_", str((item or {}).get("ResourceName")
                                            or item or ""))
    return f"{m.group(1)[:4]}-{m.group(1)[4:6]}-{m.group(1)[6:]}" \
        if m else None


def semo_dispatch_probe():
    """
    Where does per-UNIT downward dispatch live, and how far back? Log
    only. Nothing is stored, nothing is parsed into the payload.

    WHY THIS IS URGENT AND THE FEED IS NOT. Fintan Devenney (Montel)
    confirmed on 9 Aug 2026 that per-unit dispatch-down volumes are
    published on SEMO and that building against SEMO's reports would
    be fine for live data - but that SEMO does NOT retain the full
    history, which is why Montel serve their own store back to 2018.
    So the series exists only from the day capture starts. Every day
    without a feed is a day that cannot be recovered later, which
    makes finding the report the most time-critical thing in B.2.2
    even though the panel itself does not depend on it.

    WHAT THE ANSWER HAS TO CONTAIN before a parser is worth writing:
      - which report ID carries per-unit volumes (SEMO publishes
        dozens; the naming is not self-describing)
      - the retention window, since that sets how often the feed must
        run to lose nothing
      - the resolution (half-hourly imbalance periods, most likely)
      - what identifies a unit. This is the one that shapes the
        register: SEMO will name resource codes, not wind farms, so
        the map needs code -> farm -> coordinates -> capacity and
        only the first hop comes from here.

    Two families are tried because SEMO runs two: the BM/balancing
    reporting on the main site, and the semopx static-report API this
    pipeline already uses for prices. Candidates are listed, not
    chosen - a guessed report ID is how the first semopx probe wasted
    a day on 400s.
    """
    hits = []

    # --- family 1: the static-report listing this pipeline knows
    base = "https://reports.semopx.com/api/v1/documents/static-reports"
    for label, params in (
            ("semopx listing, unfiltered", {"page_size": 50}),
            ("semopx listing, dispatch text", {"page_size": 50,
                                               "search_text": "dispatch"}),
    ):
        try:
            items = http_get(base, params=params).json().get("items", [])
            ids = sorted({str(it.get("DPuG_ID") or "?") for it in items})
            names = [str(it.get("ResourceName") or "")[:60]
                     for it in items[:3]]
            log(f"semo_probe: {label} -> {len(items)} items, "
                f"report IDs {ids[:12]}")
            for n in names:
                log(f"semo_probe:     e.g. {n}")
            hits.append((label, len(items), ids))
        except Exception as e:
            log(f"semo_probe: {label} -> {e.__class__.__name__}")

    # --- family 2: the SEMO market-data API. Report IDs are listed
    # rather than assumed; whichever returns is the evidence.
    semo = "https://reports.sem-o.com/api/v1/documents/static-reports"
    for label, params in (
            ("sem-o listing, unfiltered", {"page_size": 50}),
            ("sem-o listing, page 2", {"page_size": 50, "page": 2}),
    ):
        try:
            items = http_get(semo, params=params).json().get("items", [])
            ids = sorted({str(it.get("DPuG_ID") or "?") for it in items})
            log(f"semo_probe: {label} -> {len(items)} items, "
                f"report IDs {ids[:12]}")
            for it in items[:3]:
                log(f"semo_probe:     e.g. "
                    f"{str(it.get('ResourceName') or '')[:60]}")
            hits.append((label, len(items), ids))
        except Exception as e:
            log(f"semo_probe: {label} -> {e.__class__.__name__}")

    # A response is not an answer. The first live run (13 Aug 2026)
    # returned 50 items from every trial - all one report ID, all
    # IDC_Statistic documents from 2019 - and the probe congratulated
    # itself on finding "the catalogue". This endpoint lists
    # DOCUMENTS, ascending by date, not report types, and search_text
    # is ignored. So the test is whether more than one report ID came
    # back, not whether rows did.
    ids = sorted({i for _, _, got in hits for i in got})
    if len(ids) > 1:
        log(f"semo_probe: {len(ids)} report IDs seen {ids[:12]} - "
            "look for balancing-market dispatch or unit-level "
            "availability, and write the feed against whichever names "
            "resource codes and half-hourly periods")
    else:
        log(f"semo_probe: INCONCLUSIVE - every trial returned the same "
            f"{'single report ' + ids[0] if ids else 'nothing'}, so "
            "this endpoint lists documents rather than report types "
            "and cannot enumerate the catalogue. STOP PROBING: read "
            "the report catalogue by hand at sem-o.com/market-data, "
            "then write the feed against the ID found. Two rounds of "
            "guessing have produced two wrong answers")
    log("semo_probe: reminder - SEMO does not retain full history, so "
        "the captured series starts the day the feed ships. The B.2.2 "
        "panel does not wait on it: the annual regional report and "
        "EirGrid's 30-minute jurisdiction series are already published")


def semopx_history_probe():
    """
    Can a historic SEMOpx trade day be resolved? Log-only.

    price_ai fills forward, which leaves B.2.3 unable to price the
    binding hour already found. Before writing a backfill this asks
    the listing API whether it will hand over an older day at all, and
    which parameter does it - by trying candidates and reporting what
    each returns, rather than picking one and hoping. Parsers here are
    written against evidence from live logs, never guessed; the same
    rule applies to query parameters.
    """
    base = "https://reports.semopx.com/api/v1/documents/static-reports"
    want = (today_utc() - dt.timedelta(days=120)).isoformat()
    trials = [
        ("page_size only, deep page", {"DPuG_ID": "EA-001",
                                       "page_size": 100, "page": 5,
                                       "sort_by": "Date",
                                       "order_by": "DESC"}),
        ("Date filter", {"DPuG_ID": "EA-001", "page_size": 20,
                         "Date": want}),
        ("date range", {"DPuG_ID": "EA-001", "page_size": 20,
                        "Date_from": want, "Date_to": want}),
        ("publish range", {"DPuG_ID": "EA-001", "page_size": 20,
                           "PublishDateFrom": want,
                           "PublishDateTo": want}),
    ]
    for label, params in trials:
        try:
            items = http_get(base, params=params).json().get("items", [])
            da = [it for it in items
                  if re.search(r"_SEM-DA_", str(it.get("ResourceName") or ""))]
            days = sorted(x for x in (semopx_trade_day(it) for it in da) if x)
            log(f"semopx_probe: {label} -> {len(items)} items, "
                f"{len(da)} DA, days "
                f"{days[0] if days else '-'}..{days[-1] if days else '-'} "
                f"[{'REACHES ' + days[0] if days and days[0] <= want else 'recent window only'}]")
        except Exception as e:
            log(f"semopx_probe: {label} -> {e.__class__.__name__}")
    log(f"semopx_probe: target was {want} (120 days back); if nothing "
        "reaches it, price_ai fills forward and B.2.3 waits or moves "
        "to an hour the store has priced")


SEMOPX_BACKFILL_PAGES = 6      # listing pages walked per run
SEMOPX_BACKFILL_DAYS = 12      # documents fetched per run


def semopx_backfill(prev_hourly, base, want_from):
    """
    Walk the SEMOpx listing backwards and fetch trade days the store
    has not priced. Bounded per run - the same converging-walk pattern
    the hourly chunks and the temperature archive already use, rather
    than 400 requests in one build.

    Paging is the route, not a date filter: the probe showed `Date`
    returns nothing while a deep DESC page reaches months back and an
    unsorted listing returns the oldest documents in the archive. So
    the archive holds the days; they just have to be walked to.
    """
    have_days = {k[:10] for k in prev_hourly}
    added, seen, fetched = {}, set(), 0
    for page in range(1, SEMOPX_BACKFILL_PAGES + 1):
        if fetched >= SEMOPX_BACKFILL_DAYS:
            break
        try:
            items = http_get(base, params={
                "DPuG_ID": "EA-001", "page_size": 100, "page": page,
                "sort_by": "Date", "order_by": "DESC"}).json().get("items", [])
        except Exception as e:
            log(f"semopx: backfill page {page} {e.__class__.__name__}")
            break
        if not items:
            break
        for it in items:
            name = str(it.get("ResourceName") or "")
            if "_SEM-DA_" not in name:
                continue
            day = semopx_trade_day(it)
            if not day or day in seen or day in have_days or day < want_from:
                continue
            seen.add(day)
            if fetched >= SEMOPX_BACKFILL_DAYS:
                break
            try:
                doc = parse_semopx_csv(
                    http_get(f"https://reports.semopx.com/documents/{name}").text,
                    trade_day=day)
            except Exception as e:
                log(f"semopx: backfill {day} {e.__class__.__name__}")
                continue
            fresh = {}
            for cur in (doc.get("series") or {}).values():
                for k, v in (cur.get("EUR") or {}).items():
                    fresh.setdefault(k, []).append(v)
            for k, vals in fresh.items():
                added[k] = round(statistics.mean(vals), 2)
            fetched += 1
            time.sleep(0.3)
    log(f"semopx: backfill walked {min(page, SEMOPX_BACKFILL_PAGES)} page(s), "
        f"fetched {fetched} trade day(s), +{len(added)} priced hours")
    return added


def feed_semopx():
    """
    SEMOpx DAM results. DPuG_ID=EA-001 listing confirmed live but mixes DA
    with IDA1/2/3 - select _SEM-DA_ resources explicitly, widening the page
    if the first window holds none. Document is semicolon-CSV with decimal
    commas - see parse_semopx_csv().
    """
    base = "https://reports.semopx.com/api/v1/documents/static-reports"

    def da_items(items):
        return [it for it in items
                if re.search(r"_SEM-DA_", str(it.get("ResourceName") or ""))]

    chosen = None
    for page_size in (20, 100):
        items = http_get(base, params={
            "DPuG_ID": "EA-001", "page_size": page_size,
            "sort_by": "Date", "order_by": "DESC"}).json().get("items", [])
        hits = da_items(items)
        tags = sorted({m.group(0) for it in items
                       for m in [re.search(r"SEM-[A-Z0-9]+",
                                           str(it.get("ResourceName") or ""))]
                       if m})
        log(f"semopx: page_size {page_size} - {len(items)} items, "
            f"{len(hits)} DA, auction tags seen: {tags}")
        if hits:
            chosen = hits[0]
            break
    if not chosen:
        raise ValueError("semopx: no _SEM-DA_ resource in EA-001 listing")

    resource = chosen.get("ResourceName") or chosen.get("_id")
    log("semopx: resolved", resource)
    body = http_get(f"https://reports.semopx.com/documents/{resource}")
    parsed = parse_semopx_csv(body.text,
                              trade_day=semopx_trade_day(chosen))

    if not parsed["markets"]:
        log("semopx: CSV parse empty - first 800 chars:", body.text[:800])
        raise ValueError("semopx CSV parse failed - inspect log")

    def avg(currency):
        vals = [v for mk, cur in parsed["markets"].items()
                for c, series in cur.items() if c == currency
                for v in series]
        return round(statistics.mean(vals), 2) if vals else None

    # All-island hourly EUR price: NI-DA and ROI-DA are the same
    # energy market in two currencies and settle at the same euro
    # price, so the euro series of whichever market carries it is the
    # island price. Mean where both are present rather than picking.
    hourly = dict(prev_series("semopx", "dam_hourly_eur_mwh"))
    fresh = {}
    for mk, cur in (parsed.get("series") or {}).items():
        for k, v in (cur.get("EUR") or {}).items():
            fresh.setdefault(k, []).append(v)
    for k, vals in fresh.items():
        hourly[k] = round(statistics.mean(vals), 2)
    # Bounded walk backwards. Soft: the daily document above is the
    # feed's job, this is opportunistic depth for B.2.3.
    try:
        floor = (today_utc() - dt.timedelta(days=HOURLY_MONTHS * 31)).isoformat()
        hourly.update(semopx_backfill(hourly, base, floor))
    except Exception as e:
        log(f"semopx: backfill skipped ({e.__class__.__name__})")
    out = {
        "dam_hourly_eur_mwh": trim_series(hourly),
        "dam_hourly_added": len(fresh),
        "dam_avg_eur_mwh": avg("EUR"),
        "dam_avg_gbp_mwh": avg("GBP"),
        "markets": {mk: {c: round(statistics.mean(v), 2)
                         for c, v in cur.items() if v}
                    for mk, cur in parsed["markets"].items()},
        "sem_fx_eur_gbp": parsed["fx_eur_gbp"],
        "auction": parsed["auction"],
        "trade_day": parsed["day"], "latest_day": parsed["day"],
        "source": ("SEMOpx day-ahead market results - dual currency "
                   "(EUR/GBP), incl. SEM trading-day FX rate"),
    }
    log("semopx: markets parsed:", list(parsed["markets"]))
    return out, recency_status(parsed["day"], 4)


def feed_oil_bulletin():
    """
    EU Weekly Oil Bulletin - Ireland heating gas oil, EUR/1000 L, with AND
    without taxes. Snapshot files; history accumulates in data.json across
    runs. Both sides of the border burn the same C2 kerosene; the series is
    treated as the ROI heating-oil price level (see FEED_FLAGS).
    """
    page = http_get(
        "https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en"
    ).text

    import openpyxl

    def fetch_ireland(with_tax):
        url = resolve_oil_bulletin_url(page, with_tax=with_tax)
        if not url:
            seen = [urllib.parse.unquote(u)[:110] for u in
                    re.findall(r'href="([^"]+)"', page)
                    if ".xlsx" in urllib.parse.unquote(u).lower()]
            log(f"oil_bulletin: no '{'with' if with_tax else 'without'} "
                "taxes' xlsx resolved - decoded links:", seen)
            return None, None
        log("oil_bulletin: resolved", urllib.parse.unquote(url)[:120])
        wb = openpyxl.load_workbook(
            io.BytesIO(http_get(url, timeout=180).content),
            read_only=True, data_only=True)
        for ws in wb.worksheets:
            d, v = parse_bulletin_rows(ws.iter_rows(values_only=True))
            if v is not None:
                return d, v
        for ws in wb.worksheets:
            head = [r for _, r in zip(range(6), ws.iter_rows(values_only=True))]
            log(f"oil_bulletin: sheet '{ws.title}' first rows:", head)
        return None, None

    d_wt, v_wt = fetch_ireland(True)
    d_nt, v_nt = fetch_ireland(False)
    if v_wt is None:
        raise ValueError("oil bulletin with-taxes parse failed - see log")
    for label, v in (("with", v_wt), ("without", v_nt)):
        if v is not None and not 300 <= v <= 3000:
            log(f"oil_bulletin: WARNING - {label}-taxes {v} EUR/1000L "
                "outside plausible range, check column selection")
    if d_wt is None:
        log("oil_bulletin: no bulletin date found - using today, verify")
        d_wt = today_utc().isoformat()
    log(f"oil_bulletin: Ireland heating {v_wt} EUR/1000L with taxes, "
        f"{v_nt} without, at {d_wt}")

    series = prev_series("oil_bulletin", "roi_heating_gasoil_eur_per_1000l")
    series[d_wt] = v_wt

    # weekly-history workbook backfills the chart to full depth; the
    # snapshot value above stays authoritative for its own date
    hist_url = None
    for m in re.finditer(r'href="([^"]+)"', page):
        u = m.group(1)
        d = urllib.parse.unquote(u).lower()
        if ".xlsx" in d and ("histor" in d or "time series" in d):
            hist_url = u if u.startswith("http") \
                else "https://energy.ec.europa.eu" + u
            break
    if not hist_url:
        # stable UUID cited in every weekly newsletter since 2024 - the
        # landing page does not carry this link, the newsletter does
        hist_url = ("https://energy.ec.europa.eu/document/download/"
                    "906e60ca-8b6a-44e7-8589-652854d2fd3f_en?filename="
                    "Weekly_Oil_Bulletin_Prices_History_maticni_4web.xlsx")
        log("oil_bulletin: history link not on page - using stable "
            "newsletter URL")
    if hist_url:
        log("oil_bulletin: history resolved",
            urllib.parse.unquote(hist_url)[:110])
        try:
            hwb = openpyxl.load_workbook(
                io.BytesIO(http_get(hist_url, timeout=240).content),
                read_only=True, data_only=True)
            hist, hist_nt = {}, {}
            for ws in hwb.worksheets:
                title = ws.title.lower()
                if "price" not in title:
                    continue
                got = parse_bulletin_history_rows(ws.iter_rows(values_only=True))
                if not got:
                    continue
                if "wo" in title.split() or "without" in title:
                    if len(got) > len(hist_nt):
                        hist_nt = got
                elif len(got) > len(hist):
                    hist = got
            if hist:
                log(f"oil_bulletin: history {len(hist)} Ireland weeks "
                    f"with tax, {len(hist_nt)} ex tax, "
                    f"{min(hist)}..{max(hist)}")
                merged = dict(hist)
                merged.update(series)   # snapshot/current values win
                series = merged
                globals()["_hist_nt_pending"] = hist_nt
            else:
                for ws in hwb.worksheets[:3]:
                    head = [r for _, r in zip(range(5),
                                              ws.iter_rows(values_only=True))]
                    log(f"oil_bulletin: history sheet '{ws.title}' rows:", head)
                log("oil_bulletin: history parse empty - dumps above, "
                    "snapshot-only this run")
        except Exception as e:
            log(f"oil_bulletin: history fetch/parse failed "
                f"({e.__class__.__name__}: {e}) - snapshot-only this run")
    series = trim_series(clip_days(series))
    series_nt = prev_series("oil_bulletin",
                            "roi_heating_gasoil_eur_per_1000l_ex_tax")
    hist_nt = globals().pop("_hist_nt_pending", None)
    if hist_nt:
        merged_nt = dict(hist_nt)
        merged_nt.update(series_nt)
        series_nt = merged_nt
    if v_nt is not None:
        series_nt[d_nt or d_wt] = v_nt
    series_nt = trim_series(clip_days(series_nt))
    out = {
        "roi_heating_gasoil_eur_per_1000l": series,
        "roi_heating_gasoil_eur_per_1000l_ex_tax": series_nt,
        "latest_value": v_wt,
        "latest_value_ex_tax": v_nt,
        "latest_day": max(series),
        "source": ("European Commission Weekly Oil Bulletin - prices with "
                   "and without taxes"),
    }
    return out, recency_status(out["latest_day"], 10)


def feed_ccni_oil():
    """
    Consumer Council NI - daily checker page (weekly page confirmed
    chart-free). Parsed series merges with previous runs so history extends
    beyond the page's rolling window. See FEED_FLAGS for the average-vs-
    council-area verification item.
    """
    url = "https://www.consumercouncil.org.uk/home-heating/price-checker/daily"
    page = http_get(url).text
    # per-chart diagnostics: the page can embed more than one litre-labelled
    # chart, and 14/15 Jul runs stored different values for the same date -
    # log every candidate so the series identity question is answerable
    for i, arr in enumerate(extract_chart_data_arrays(page)):
        if arr and isinstance(arr[0], list) \
                and any("litre" in str(c).lower() for c in arr[0]):
            body = [r for r in arr[1:] if r]
            log(f"ccni_oil: chart {i} header={arr[0]} n={len(body)} "
                f"first={body[0] if body else None} "
                f"last={body[-1] if body else None}")
    parsed = parse_ccni_series(page)
    n = sum(len(v) for v in parsed.values())
    if not n:
        arrays = extract_chart_data_arrays(page)
        log("ccni_oil: no litre-labelled chart; "
            f"{len(arrays)} chart array(s) found, first headers:",
            [a[0] for a in arrays[:3] if a])
        raise ValueError("ccni_oil: no series parsed - inspect log")
    log(f"ccni_oil: {n} datapoints across "
        f"{[k for k, v in parsed.items() if v]}")

    merged, conflicts = {}, 0
    for k, new in parsed.items():
        old = prev_series("ccni_oil", "series_gbp", "daily", k)
        for d, v in new.items():
            if d in old and old[d] and abs(v - old[d]) / old[d] > 0.05:
                conflicts += 1
                if conflicts <= 3:
                    log(f"ccni_oil: SERIES BREAK {k} {d}: stored "
                        f"{old[d]} -> page {v}")
        old.update(new)
        merged[k] = trim_series(clip_days(old))
    if conflicts:
        log(f"ccni_oil: {conflicts} same-date value conflicts vs stored "
            "history - series identity unstable, new values kept")
    out = {"series_gbp": {"daily": merged},
           "series_conflicts_this_run": conflicts}

    all_days = [d for s in merged.values() for d in s]
    out["latest_day"] = max(all_days) if all_days else None
    out["source"] = ("Consumer Council for Northern Ireland home heating oil "
                     "price checker - daily (Mon-Fri), NI average, "
                     "300/500/900 L")
    return out, recency_status(out["latest_day"], 7)



# ---------------------------------------------------------------- probes
# Structured probes (soft): fetch and log response shape only, so a run
# log establishes what each source actually serves before anything
# depends on it - the pattern that adopted the EirGrid CO2 series.
# Candidates from Paul Deane's source list, 27 Jul 2026.

# The ENTSOG probe is log-only and its finding is analysed
# fortnightly, but it polled six points every run - 4.5 of a 7-minute
# build on 13 Aug 2026, and 5.5 on a bad day. Its own retention is 25
# days, so a twice-weekly poll loses nothing, and the standing NI-exit
# finding it exists to witness moves at the pace of a monthly gas
# balance rather than a daily one.
ENTSOG_POLL_WEEKDAYS = (0, 3)      # Monday and Thursday


def feed_entsog_probe():
    """ENTSOG Transparency Platform - GB<->IE / GB<->GB(NI)
    interconnection points. Goal: observe Moffat / SNIP physical flows
    as (a) an independent NI gas measurement and (b) a cross-check of
    the gni_live jurisdiction finding (Total LDM 25% above ROI LDM)."""
    # Round 3 (30 Jul 2026): schema discovered - filter
    # operatorpointdirections by tSOCountry / adjacentCountry, list
    # Ireland-facing points, then sample one day of Physical Flow.
    out = {"source": "ENTSOG Transparency Platform (probe, round 3)"}
    # Twice weekly. On other days the retained series is returned
    # untouched, so the feed still reports and nothing downstream
    # notices - it simply does not spend four minutes re-measuring a
    # standing finding.
    if today_utc().weekday() not in ENTSOG_POLL_WEEKDAYS:
        keep = {k: prev_series("entsog_probe", k)
                for k in (PREVIOUS_FEEDS.get("entsog_probe") or {})
                if k.endswith("_gwh_daily")}
        prev_latest = ((PREVIOUS_FEEDS.get("entsog_probe") or {})
                       .get("latest_day"))
        if keep:
            out.update(keep)
            out["latest_day"] = prev_latest
            log(f"entsog_probe: skipped - polls "
                f"{'/'.join(dt.date(2026, 1, 5 + d).strftime('%a') for d in ENTSOG_POLL_WEEKDAYS)}"
                f", {len(keep)} series retained, latest {prev_latest}")
            return out, "lagging"
    pts = {}
    for q in ("tSOCountry=IE", "adjacentCountry=IE"):
        try:
            r = http_get("https://transparency.entsog.eu/api/v1/"
                         f"operatorpointdirections?{q}&limit=300",
                         timeout=90)
            rows = (r.json() or {}).get("operatorpointdirections", [])
            for row in rows:
                k = row.get("pointKey")
                if k:
                    pts[k] = (row.get("pointLabel"),
                              row.get("tSOCountry"),
                              row.get("adjacentCountry"),
                              row.get("directionKey"),
                              row.get("crossBorderPointType"))
            log(f"entsog_probe: {q} -> {len(rows)} rows")
        except Exception as e:
            log(f"entsog_probe: {q} {e.__class__.__name__}: {e}")
    for k, v in sorted(pts.items())[:12]:
        log(f"entsog_probe: point {k}: label={v[0]} tso={v[1]} "
            f"adj={v[2]} dir={v[3]} type={v[4]}")
    out["ie_points"] = sorted(pts)

    # Round 4a: the SNIP (Scotland->NI) is UK->UK, invisible to both
    # IE filters - label sweep of a large page for NI-relevant points.
    try:
        r = http_get("https://transparency.entsog.eu/api/v1/"
                     "operatorpointdirections?tSOCountry=UK"
                     "&limit=3000", timeout=120)
        rows = (r.json() or {}).get("operatorpointdirections", [])
        WANT = ("twynholm", "ballylumford", "brighouse", "belfast",
                "larne", "islandmagee", "scotland", "snip",
                "moffat", "south north", "gormanston", "corrib",
                "inch", "bellanaboy")
        hits = {}
        for row in rows:
            lbl = (row.get("pointLabel") or "").lower()
            if any(w in lbl for w in WANT):
                hits[row.get("pointKey")] = (row.get("pointLabel"),
                                             row.get("directionKey"),
                                             row.get("adjacentCountry"),
                                             row.get("adjacentZones"))
        log(f"entsog_probe: UK sweep {len(rows)} rows, "
            f"NI-relevant hits {len(hits)}")
        for k, v in sorted(hits.items())[:10]:
            log(f"entsog_probe: NI point {k}: label={v[0]} dir={v[1]} "
                f"adj={v[2]} zones={v[3]}")
        out["ni_candidate_points"] = sorted(hits)
    except Exception as e:
        log(f"entsog_probe: UK sweep {e.__class__.__name__}: {e}")

    # Round 6 (30 Jul 2026): Moffat combined, Twynholm and Greater
    # Belfast answer with sane values; the island-side mirror points
    # (Moffat IE/NI, South North) return non-JSON error documents -
    # their TSOs appear not to publish Physical Flow, so the UK-side
    # points are the observables. Retain daily series for the points
    # that answer: this is NI's first observed gas record, banking
    # from today. Jurisdiction confrontation runs on SNIP flow, the
    # measurable NI-bound artery (South North, ROI->NI, unmeasured).
    SHORT = {"ITP-00090": ("moffat_combined", "Moffat (combined)"),
             "ITP-00077": ("twynholm_snip", "Twynholm (SNIP)"),
             "DIS-00015": ("greater_belfast", "Greater Belfast"),
             "ITP-00495": ("moffat_ie", "Moffat (IE)"),
             "ITP-00496": ("moffat_ni", "Moffat (NI)"),
             "ITP-00222": ("south_north", "South North CSEP")}
    frm = (today_utc() - dt.timedelta(days=10)).isoformat()
    to = today_utc().isoformat()
    means = {}
    for pk, (slug, lbl) in SHORT.items():
        got = {}
        for dk in ("exit", "entry"):
            try:
                r2 = http_get(
                    "https://transparency.entsog.eu/api/v1/"
                    "operationaldata"
                    f"?pointKey={pk}&indicator=Physical%20Flow"
                    f"&periodType=day&directionKey={dk}"
                    f"&from={frm}&to={to}&limit=30", timeout=90)
                ctype = r2.headers.get("content-type", "?")
                if "json" not in ctype:
                    log(f"entsog_probe: {lbl} [{dk}] non-JSON "
                        f"({r2.status_code}, {ctype[:30]})")
                    continue
                j2 = r2.json() or {}
                key = next((k for k in j2
                            if isinstance(j2[k], list)), None)
                for x in (j2.get(key) or []):
                    try:
                        d = str(x.get("periodFrom", ""))[:10]
                        v = float(x["value"]) / 1e6
                        if d:
                            got[d] = got.get(d, 0.0) + v
                    except (TypeError, ValueError, KeyError):
                        continue
                if got:
                    break
            except Exception as e:
                log(f"entsog_probe: {lbl} [{dk}] "
                    f"{e.__class__.__name__}")
        if got:
            ser = prev_series("entsog_probe", f"{slug}_gwh_daily")
            ser.update(got)
            out[f"{slug}_gwh_daily"] = trim_series(ser)
            means[slug] = sum(got.values()) / len(got)
            log(f"entsog_probe: {lbl}: {len(got)} days, mean "
                f"{means[slug]:.1f} GWh/d, series "
                f"{len(out[f'{slug}_gwh_daily'])}d retained")
    snip = means.get("twynholm_snip")
    if snip is not None:
        sn = means.get("south_north")
        ni_bound = snip + (sn or 0.0)
        note = ("SNIP + South North" if sn is not None
                else "SNIP alone; South North (ROI->NI) unmeasured")
        log(f"entsog_probe: jurisdiction confrontation - NI-bound "
            f"{ni_bound:.1f} GWh/d ({note}) vs gni_live Total-ROI "
            f"LDM difference ~35.8 GWh/d "
            + ("-> magnitudes MATCH: independent witness for the "
               "MATERIAL verdict"
               if abs(ni_bound - 35.8) < 12 else
               "-> gap remains: scope needs a second look"))
    out["latest_day"] = today_utc().isoformat()
    return out, "ok"



def feed_eirgrid_probe():
    """Discovery dispatch for the v7 hourly engine (A'.2). One run,
    four chart types x three regions, answering the questions that
    stand between us and the 13-month hourly store:

      1. HISTORY DEPTH - does the endpoint serve a year, or only a
         rolling window? This is the single unknown that decides
         whether the store backfills or fills forward.
      2. WIND and SOLAR - both are on the endpoint we already use for
         demand and co2; we have simply never asked. Solar matters:
         AIRAA shows 3,260 MW in Ireland for 2026 rising to 6,880 by
         2031, and Energy-Charts cannot see it at all (which is the
         likeliest cause of the sem_mix indigenous-share shortfall).
      3. GRANULARITY - 15-minute rows aggregate to hourly MEANS, never
         samples; confirm the interval actually returned.
      4. REGIONALITY - is wind/solar published per jurisdiction as
         demand is, or all-island only?

    Soft, log-only. Nothing downstream depends on it.
    """
    # ROUND 2 (7 Aug 2026). Round 1's 400s were a parameter error, not
    # a missing series: `areas` must match the chart (demand ->
    # demandactual, co2 -> co2intensity), and I omitted it for wind,
    # solar and co2. Round 1 did establish: 15-minute interval,
    # 2,845/2,880 rows valued (98.8%), all three regions served for
    # demand, and a month returned regardless of the span requested -
    # so depth is the walk-back question tested below.
    out = {"source": "EirGrid Smart Grid Dashboard (probe, round 3)"}
    end = today_utc()

    def dmy(d):
        return d.strftime("%d-%b-%Y").replace(" 0", " ")

    def call(chart, region, areas, frm, to, rng="month"):
        r = http_get(EIRGRID_ENDPOINT, params={
            "region": region, "chartType": chart, "dateRange": rng,
            "dateFrom": dmy(frm), "dateTo": dmy(to), "areas": areas,
        }, timeout=120).json()
        rows = (r or {}).get("Rows", []) or []
        st = [x.get("EffectiveTime") for x in rows if x.get("EffectiveTime")]
        vals = [x.get("Value") for x in rows if x.get("Value") is not None]
        return rows, st, vals

    # --- A: correct areas names for wind and solar
    CAND = {"wind": ["windactual", "windforecast", "generationactual"],
            "solar": ["solaractual", "solarforecast"],
            "co2": ["co2intensity"], "demand": ["demandactual"]}
    good = {}
    for chart, names in CAND.items():
        for areas in names:
            try:
                rows, st, vals = call(chart, "ALL", areas,
                                      end - dt.timedelta(days=8), end)
                if rows:
                    good[chart] = areas
                    log(f"eirgrid_probe: {chart}/ALL areas={areas} OK - "
                        f"{len(rows)} rows, {len(vals)} valued, "
                        f"{st[0]} .. {st[-1]}")
                    break
                log(f"eirgrid_probe: {chart} areas={areas} 0 rows")
            except Exception as e:
                log(f"eirgrid_probe: {chart} areas={areas} "
                    f"{e.__class__.__name__}")

    # --- B: regionality of whichever series answered
    for chart, areas in good.items():
        if chart in ("demand", "co2"):
            continue
        for region in ("ROI", "NI"):
            try:
                rows, st, vals = call(chart, region, areas,
                                      end - dt.timedelta(days=8), end)
                log(f"eirgrid_probe: {chart}/{region} {len(rows)} rows, "
                    f"{len(vals)} valued")
            except Exception as e:
                log(f"eirgrid_probe: {chart}/{region} "
                    f"{e.__class__.__name__}")

    # --- C: THE DEPTH QUESTION. Can past months be walked? If a
    # request centred on a historic month returns that month's data,
    # the 13-month store backfills by chunked walking. If it returns
    # the current month (or nothing), the store fills forward instead
    # and the seasonal exhibit waits for the calendar.
    areas = good.get("demand", "demandactual")
    for back in (2, 6, 12):
        m_end = end - dt.timedelta(days=30 * back)
        m_start = m_end - dt.timedelta(days=27)
        try:
            rows, st, vals = call("demand", "ALL", areas, m_start, m_end)
            first = st[0] if st else "-"
            last = st[-1] if st else "-"
            hit = ("HISTORIC DATA RETURNED - walk-back works"
                   if st and str(m_end.year) in str(first)
                   and first[:6].lower() in
                   dmy(m_start).lower()[:6] + dmy(m_end).lower()[:6]
                   else "check span against request")
            log(f"eirgrid_probe: WALK-BACK {back}mo "
                f"(asked {dmy(m_start)}..{dmy(m_end)}) -> "
                f"{len(rows)} rows, got {first} .. {last} [{hit}]")
        except Exception as e:
            log(f"eirgrid_probe: WALK-BACK {back}mo "
                f"{e.__class__.__name__}: {e}")

    # --- C2: THE CARBON DEPTH QUESTION (round 3, 8 Aug 2026).
    #
    # This one decides the shape of the 24-month view. The hourly
    # store holds 13 months and the daily feed 391 days, so every week
    # before about mid-July 2025 has no own-week grid intensity from
    # anything retained. If co2intensity walks back two years, the
    # extension keeps the back-look's central claim - each week priced
    # at its own week's carbon. If it does not, those weeks need an
    # annual or monthly intensity and the window has to say so.
    #
    # RETRIED AND REPEATED, deliberately. On 8 Aug 2026 this endpoint
    # returned 0 rows for co2intensity and demandactual within seconds
    # of serving wind and solar, and the main feed pulled its 30 days
    # on the same run. A single zero is therefore not evidence of
    # depth - it is as likely to be the endpoint being flaky for that
    # chart at that moment. Nothing here is called EMPTY until two
    # independent attempts agree.
    co2_areas = good.get("co2", "co2intensity")
    depth = {}
    for back in (12, 18, 24):
        m_end = end - dt.timedelta(days=30 * back)
        m_start = m_end - dt.timedelta(days=27)
        got = []
        for attempt in (1, 2):
            try:
                rows, st, vals = call("co2", "ALL", co2_areas,
                                      m_start, m_end)
                got.append((len(rows), len(vals),
                            st[0] if st else "-", st[-1] if st else "-"))
                if rows:
                    break
                time.sleep(1.5)
            except Exception as e:
                got.append((0, 0, f"{e.__class__.__name__}", "-"))
                time.sleep(1.5)
        best = max(got, key=lambda g: g[0])
        # A returned span that lands in the requested month is the
        # real test; the endpoint is known to serve the current month
        # regardless of what was asked, which would look like success.
        want = {dmy(m_start)[3:], dmy(m_end)[3:]}
        landed = any(w.lower() in str(best[2]).lower() for w in want)
        if best[0] and landed:
            verdict = "HISTORIC CARBON RETURNED"
        elif best[0]:
            verdict = "rows returned but span is NOT the month asked for"
        else:
            verdict = f"EMPTY on {len(got)} attempts - treat as unavailable"
        depth[f"{back}mo"] = verdict
        log(f"eirgrid_probe: CARBON DEPTH {back}mo "
            f"(asked {dmy(m_start)}..{dmy(m_end)}) -> "
            f"{best[0]} rows, {best[1]} valued, got {best[2]} .. {best[3]} "
            f"[{verdict}]")
    out["carbon_depth"] = depth
    reach = [k for k, v in depth.items() if v == "HISTORIC CARBON RETURNED"]
    log("eirgrid_probe: CARBON DEPTH verdict - "
        + (f"own-week carbon reaches {max(reach, key=lambda k: int(k[:-2]))}"
           if reach else
           "no historic carbon at any depth tested; a 24-month window "
           "would need an annual or monthly intensity and must say so"))

    # --- D: does dateRange=year serve more than a month?
    try:
        rows, st, vals = call("demand", "ALL", areas,
                              end - dt.timedelta(days=365), end,
                              rng="year")
        log(f"eirgrid_probe: dateRange=year -> {len(rows)} rows, "
            f"{st[0] if st else '-'} .. {st[-1] if st else '-'}")
    except Exception as e:
        log(f"eirgrid_probe: dateRange=year {e.__class__.__name__}: {e}")

    out["latest_day"] = end.isoformat()
    return out, "ok"


def feed_sem_mix():
    """Energy-Charts (Fraunhofer ISE) public_power for Ireland (SEM,
    all-island) - adopted 30 Jul 2026 after the probe confirmed 13
    half-hourly production types. Computes a daily indigenous share of
    electricity: wind + hydro RoR + peat + Others x other-indigenous
    anchor, over generation plus imports; gas-fired generation is
    credited at the gas indigenous anchor for cross-fuel consistency.
    Pumped storage (both directions) is recycled load - excluded."""
    out = {"source": ("energy-charts.info public_power IE - "
                      "SEM all-island, half-hourly")}
    end = today_utc()
    start = end - dt.timedelta(days=35)
    r = http_get("https://api.energy-charts.info/public_power"
                 f"?country=ie&start={start.isoformat()}"
                 f"&end={end.isoformat()}", timeout=90)
    j = r.json() or {}
    stamps = j.get("unix_seconds") or []
    types = {p.get("name"): p.get("data") or []
             for p in (j.get("production_types") or [])}
    a_ind = ANCHORS["indigenous"]
    W_IND = {"Wind onshore": 1.0, "Hydro Run-of-River": 1.0,
             "Fossil peat": a_ind.get("peat", 1.0),
             "Others": a_ind.get("other", 0.9),
             "Fossil gas": a_ind.get("gas", 0.0),
             "Fossil oil": a_ind.get("oil", 0.0)}
    GEN = ["Wind onshore", "Hydro Run-of-River", "Fossil peat",
           "Others", "Fossil gas", "Fossil oil"]
    days = {}
    for i, ts in enumerate(stamps):
        d = dt.datetime.fromtimestamp(ts, dt.timezone.utc)\
            .date().isoformat()
        rec = days.setdefault(d, {"ind": 0.0, "gen": 0.0,
                                  "imp": 0.0, "n": 0})
        vals = {t: (types.get(t) or [None])[i]
                if i < len(types.get(t) or []) else None for t in GEN}
        if any(v is None for v in vals.values()):
            continue
        gen = sum(vals.values())
        ind = sum(v * W_IND[t] for t, v in vals.items())
        x = (types.get("Cross border electricity trading")
             or [None])[i] if i < len(
                 types.get("Cross border electricity trading") or []) \
            else None
        imp = max(x, 0.0) if x is not None else 0.0
        rec["ind"] += ind
        rec["gen"] += gen
        rec["imp"] += imp
        rec["n"] += 1
    ser = prev_series("sem_mix", "indigenous_share_daily")
    for d, rec in days.items():
        if rec["n"] >= 40:   # >=20h of half-hours = a complete-ish day
            denom = rec["gen"] + rec["imp"]
            if denom > 0:
                ser[d] = round(100.0 * rec["ind"] / denom, 2)
    out["indigenous_share_daily"] = trim_series(ser)
    ds = sorted(out["indigenous_share_daily"])
    log(f"sem_mix: {len(days)} days fetched, "
        f"{len(ds)} complete retained"
        + (f", latest {ds[-1]} at "
           f"{out['indigenous_share_daily'][ds[-1]]}%" if ds else ""))
    if not ds:
        raise RuntimeError("sem_mix: no complete days")
    out["latest_day"] = ds[-1]
    return out, recency_status(out["latest_day"], 4)


def feed_gb_oil():
    """
    GB heating-oil context line, two strategies (SOFT feed):
      A. BoilerJuice /kerosene-prices/ server-rendered sentence - present
         on legacy edge renders, absent on the modern template; tried with
         two user agents since the edge appears to vary by client.
      B. DESNZ/gov.uk monthly petroleum products table - official, stable,
         resolved from the statistics landing page at runtime. First
         contact logs candidate links, sheet names and header rows so the
         parser can be pinned in one iteration.
    History accumulates across runs whichever strategy lands.
    """
    # --- strategy A: BoilerJuice sentence. Diagnosis 18 Jul 2026: the CDN
    # serves non-browser clients an archived template (observed: Oct 2021,
    # (c) 2004-2021) while browsers and search crawlers receive the live
    # page with today's sentence. Countermeasures: browser headers, a
    # cache-busting query so no cached variant matches, and a freshness
    # gate - a parsed date older than 7 days is a fossil, rejected and
    # logged, never stored.
    A_HEADERS = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/126.0.0.0 Safari/537.36",
        "Referer": "https://www.google.com/",
        "Accept": "text/html,application/xhtml+xml",
        "Accept-Language": "en-GB,en;q=0.9",
        "Cache-Control": "no-cache", "Pragma": "no-cache",
    }
    fresh_floor = (today_utc() - dt.timedelta(days=7)).isoformat()
    for url in ("https://www.boilerjuice.com/kerosene-prices/",
                "https://www.boilerjuice.com/heating-oil-prices/"):
        for attempt in range(2):
            try:
                page = http_get(url, headers=A_HEADERS, retries=1,
                                params={"cb": int(time.time()) + attempt}
                                ).text
            except Exception as e:
                log(f"gb_oil: A fetch failed on {url.rsplit('/', 2)[-2]} "
                    f"({e.__class__.__name__})")
                break
            d, ppl = parse_gb_oil_page(page)
            if ppl is None:
                log(f"gb_oil: A no sentence on {url.rsplit('/', 2)[-2]} "
                    f"attempt {attempt + 1} - page {len(page)} chars")
                if attempt == 0:
                    i = page.lower().find("verage")
                    if i >= 0:
                        log("gb_oil: A context around 'average':",
                            re.sub(r"\s+", " ",
                                   page[max(0, i - 80):i + 320]))
                    for pat in ("pence", "chart", "props",
                                "application/json"):
                        j = page.lower().find(pat)
                        if j >= 0:
                            log(f"gb_oil: A first '{pat}' at {j}:",
                                re.sub(r"\s+", " ", page[j:j + 240]))
                continue
            if d is None:
                d = today_utc().isoformat()
            if d < fresh_floor:
                log(f"gb_oil: A fossil template on "
                    f"{url.rsplit('/', 2)[-2]} - sentence dated {d}, "
                    f"{ppl} p/L - rejected, cache-busting again")
                continue
            log(f"gb_oil: A (BoilerJuice) {ppl} p/L at {d}")
            return _gb_oil_out(d, ppl,
                               "BoilerJuice UK daily average (lowest quote "
                               "per postcode district, incl 5% VAT)")
    log("gb_oil: A exhausted (fossil or sentence-less variants only) - "
        "falling through to DESNZ")

    # --- strategy B: DESNZ "Oil and petroleum products monthly
    # statistics" (QEP 4.1.1: monthly typical retail prices incl.
    # standard grade burning oil - GB home heating oil). Canonical page
    # pinned 18 Jul 2026; xlsx links resolved from it at runtime.
    # Diagnostics-first: sheet names and header rows are logged before
    # the parser commits, so a format change costs one log read.
    try:
        page = http_get(
            "https://www.gov.uk/government/statistical-data-sets/"
            "oil-and-petroleum-products-monthly-statistics").text
    except Exception as e:
        log(f"gb_oil: B page fetch failed ({e.__class__.__name__}: {e})")
        return _gb_oil_stale()
    links = re.findall(
        r'href="(https://assets\.publishing\.service\.gov\.uk/'
        r'[^"]+\.xlsx)"', page)
    pref = [u for u in links if "4.1.1" in u or "411" in u
            or "petroleum" in u.lower()] or links
    log(f"gb_oil: B found {len(links)} xlsx links; trying "
        f"{pref[0].rsplit('/', 1)[-1] if pref else 'none'}")
    if not pref:
        return _gb_oil_stale()
    try:
        import openpyxl
        blob = http_get(pref[0], timeout=120).content
        wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True,
                                    read_only=True)
    except Exception as e:
        log(f"gb_oil: B download/open failed "
            f"({e.__class__.__name__}: {e})")
        return _gb_oil_stale()
    series = {}
    for ws in wb.worksheets:
        got = parse_desnz_burning_oil_rows(
            ws.iter_rows(values_only=True))
        if got:
            log(f"gb_oil: B parsed {len(got)} months from sheet "
                f"'{ws.title}', latest {max(got)} = "
                f"{got[max(got)]} p/L")
            series = got
            break
        else:
            hdr = [list(r)[:8] for _i, r in
                   zip(range(3), ws.iter_rows(values_only=True))]
            log(f"gb_oil: B sheet '{ws.title}' no burning-oil parse; "
                f"first rows: {hdr}")
    if not series:
        return _gb_oil_stale()
    prev = prev_series("gb_oil", "gb_ppl_daily")
    prev.update(series)
    latest = max(series)
    return ({"gb_ppl_daily": trim_series(clip_days(prev)),
             "latest_day": latest,
             "latest_ppl": series[latest],
             "source": ("DESNZ QEP 4.1.1 - typical retail price, "
                        "standard grade burning oil, monthly, incl. "
                        "taxes"),
             },
            recency_status(latest, 100))


def parse_desnz_burning_oil_rows(rows) -> dict:
    """
    QEP 4.1.1-style sheet -> {iso_date (mid-month): pence_per_litre}.
    Finds a header row containing a burning-oil/kerosene column, then
    reads month rows (datetime cells or 'January 2026' text) with a
    plausible price (25-250 p/L). Pure, unit tested against a fixture;
    real-format confirmation comes from the first run's diagnostics.
    """
    MONTHS = {m.lower(): i + 1 for i, m in enumerate(
        ["January", "February", "March", "April", "May", "June", "July",
         "August", "September", "October", "November", "December"])}
    idx = None
    series = {}
    for row in rows:
        cells = list(row)
        strs = [str(c) if c is not None else "" for c in cells]
        if idx is None:
            for i, s in enumerate(strs):
                sl = s.lower()
                if "burning oil" in sl or "kerosene" in sl:
                    idx = i
                    break
            continue
        row_date = None
        for c in cells:
            if isinstance(c, dt.datetime):
                row_date = c.date().replace(day=15).isoformat()
                break
            if isinstance(c, dt.date):
                row_date = c.replace(day=15).isoformat()
                break
            if isinstance(c, str):
                parts = c.strip().split()
                if len(parts) == 2 and parts[0].lower() in MONTHS \
                        and parts[1].isdigit():
                    row_date = (f"{int(parts[1]):04d}-"
                                f"{MONTHS[parts[0].lower()]:02d}-15")
                    break
        if row_date is None:
            continue
        try:
            v = float(cells[idx])
        except (TypeError, ValueError, IndexError):
            continue
        if 25 <= v <= 250:
            series[row_date] = round(v, 2)
    return series


def _gb_oil_stale():
    series = prev_series("gb_oil", "gb_ppl_daily")
    return ({"gb_ppl_daily": series,
             "latest_day": max(series) if series else None,
             "source": "GB heating oil - awaiting source",
             }, "stale")


def _gb_oil_out(d, ppl, source):
    series = prev_series("gb_oil", "gb_ppl_daily")
    series[d] = ppl
    series = trim_series(clip_days(series))
    return ({"gb_ppl_daily": series,
             "latest_day": max(series) if series else None,
             "source": source},
            recency_status(max(series) if series else None, 40))


# ------------------------------------------------- analysis (pure functions)

def space_heat_split(gas_daily: dict, hdd_daily: dict):
    """
    Space-heat sensitivity of daily gas demand. Primary estimator is
    within-class (monthly) centring - within-month deviations of demand on
    deviations of HDD - which removes seasonal confounds (holidays, school
    terms, baseload drift) that bias the naive slope. The naive OLS is
    retained for reference. See tests/test_synthetic.py.
    """
    days = sorted(set(gas_daily) & set(hdd_daily))
    if len(days) < 30:
        return None
    x = [hdd_daily[d] for d in days]
    y = [gas_daily[d] for d in days]
    n = len(days)
    mx, my = statistics.mean(x), statistics.mean(y)

    def ols(xs, ys):
        mxx, myy = statistics.mean(xs), statistics.mean(ys)
        sxx = sum((xi - mxx) ** 2 for xi in xs)
        if sxx == 0:
            return None, None
        slope = sum((xi - mxx) * (yi - myy)
                    for xi, yi in zip(xs, ys)) / sxx
        return slope, myy - slope * mxx

    naive_slope, _ = ols(x, y)

    # within-class centring: subtract each month's own means
    from collections import defaultdict
    bym = defaultdict(list)
    for d, xi, yi in zip(days, x, y):
        bym[d[:7]].append((xi, yi))
    xd, yd = [], []
    for pts in bym.values():
        mxm = statistics.mean(p[0] for p in pts)
        mym = statistics.mean(p[1] for p in pts)
        for xi, yi in pts:
            xd.append(xi - mxm)
            yd.append(yi - mym)
    sxx = sum(v * v for v in xd)
    if sxx == 0:
        return None
    slope = sum(a * b for a, b in zip(xd, yd)) / sxx
    ss_res = sum((b - slope * a) ** 2 for a, b in zip(xd, yd))
    ss_tot = sum(b * b for b in yd) or 1e-9
    dof = max(1, len(xd) - len(bym) - 1)
    return {"slope_gwh_per_hdd": round(slope, 3),
            "baseload_gwh_per_day": round(my - slope * mx, 2),
            "r2_within_month": round(1 - ss_res / ss_tot, 3),
            "residual_se_gwh_per_day": round((ss_res / dof) ** 0.5, 2),
            "n_days": n,
            "method": "within-month centred OLS (within-class centring)",
            "naive_slope_gwh_per_hdd": round(naive_slope, 3)
            if naive_slope is not None else None}


def derive_gas_calibration(reg, hdd_roi, anchors=None):
    """
    Calibration disclosure: the regression-implied annual gas space heat
    (slope x trailing-year ROI degree days) against the anchor-implied
    figure (ROI buildings heat x gas share x space-heat fraction, input
    basis). Ratio published with a +/-10% gate; a miss is disclosed, not
    hidden - the two measures differ in scope (distribution-metered gas
    vs whole-anchor) and the ratio quantifies exactly that.
    """
    a = anchors or ANCHORS
    if not reg or not hdd_roi:
        return None
    days = sorted(hdd_roi)[-365:]
    if len(days) < 300:
        return None
    annual_hdd = sum(hdd_roi[d] for d in days)
    implied = reg["slope_gwh_per_hdd"] * annual_hdd
    j = a["roi"]
    anchor = ((j["residential_heat_twh"] + j["services_heat_twh"])
              * j["fuel_shares"]["gas"] * a["space_heat_fraction"] * 1000.0)
    ratio = implied / anchor if anchor else None
    return {"implied_annual_space_heat_gwh": round(implied, 0),
            "anchor_annual_space_heat_gwh": round(anchor, 0),
            "ratio": round(ratio, 2), "gate": "0.90-1.10",
            "within_gate": bool(0.90 <= ratio <= 1.10),
            "note": ("Scopes differ: the regression sees "
                     "distribution-metered gas; the anchor is the whole "
                     "buildings-gas estimate. The ratio measures the "
                     "difference and is published either way. A ratio "
                     "below one is expected - much of services-sector "
                     "gas is daily-metered and outside the NDM series "
                     "the regression sees.")}



# ------------------------------------------------------ weekly back-look
def ni_bridge_margin(feeds):
    """
    Calibrated NI-vs-bulletin margin, p/L: mean over the CCNI overlap
    of (observed NI 900L p/L) minus (bulletin ex-tax x week FX x 1.05
    VAT). Requires >=20 overlapping days. Bridged weeks reconstruct
    pre-CCNI NI prices as bridge + margin, dagger.
    """
    ccni = (((feeds.get("ccni_oil") or {}).get("series_gbp") or {})
            .get("daily") or {}).get("900l") or {}
    ext = ((feeds.get("oil_bulletin") or {})
           .get("roi_heating_gasoil_eur_per_1000l_ex_tax") or {})
    fxs = ((feeds.get("ecb_fx") or {}).get("eur_gbp_daily") or {})
    if not (ccni and ext and fxs):
        return None
    ext_days = sorted(ext)
    diffs = []
    for d in sorted(ccni):
        if d not in fxs:
            continue
        prior = [x for x in ext_days if x <= d]
        if not prior:
            continue
        est = ext[prior[-1]] / 1000.0 * fxs[d] * 100.0 * 1.05
        diffs.append(ccni[d] / 9.0 - est)
    if len(diffs) < 20:
        return None
    return round(sum(diffs) / len(diffs), 2)


def report_skips(skips, built):
    """
    Say which weeks were dropped and why, grouped by reason.

    A silent `continue` is the worst failure mode this pipeline has:
    the record simply comes out shorter, which reads as a smaller
    number rather than an error, and there is nothing in the log to
    contradict it. Sixteen weeks quietly missing from a 104-week
    extension would look exactly like a 88-week extension.
    """
    if not skips:
        log(f"history: {built} weeks built, none skipped")
        return
    by_reason = {}
    for w, r in skips:
        by_reason.setdefault(r, []).append(w)
    log(f"history: WARNING {len(skips)} week(s) skipped of "
        f"{built + len(skips)} attempted")
    for r, ws in sorted(by_reason.items()):
        ws = sorted(ws)
        span = ws[0] if len(ws) == 1 else f"{ws[0]}..{ws[-1]}"
        log(f"history:   {len(ws)} week(s) [{span}] - {r}")


def week_inputs(feeds, w_end, skips=None):
    """
    Per-week prices, fx, carbon and tariffs, or None if the week
    cannot be priced.

    `skips` is an optional list this appends (week, reason) to. It
    exists because returning None silently makes a short back-look
    look like a smaller number rather than an error - a week that
    cannot be built simply is not there, and nothing says so. The
    only silent decline is a week outside the window by design.
    """
    """Per-week pricing context for a calendar week ending w_end (Sun):
    NI oil (CCNI 900L weekly mean, p/L), ROI oil (bulletin week), fx
    (weekly mean of daily ECB), electricity EF (weekly CI mean or
    anchor), tariffs (period resolver). Returns None if a required
    input is absent - the week is not built."""
    tar = tariffs_for(w_end)
    if tar is None:
        if skips is not None:
            skips.append((w_end, "before the tariff table starts "
                                 f"({TARIFF_HISTORY[0][0]}) - extend "
                                 "TARIFF_HISTORY to price it"))
        return None

    def _skip(reason):
        if skips is not None:
            skips.append((w_end, reason))
        return None

    w_start = (dt.date.fromisoformat(w_end)
               - dt.timedelta(days=6)).isoformat()
    days = [(dt.date.fromisoformat(w_start) + dt.timedelta(days=i))
            .isoformat() for i in range(7)]
    if w_end < HISTORY_START:
        return None
    ccni = (((feeds.get("ccni_oil") or {}).get("series_gbp") or {})
            .get("daily") or {}).get("900l") or {}
    ni_vals = [ccni[d] / 9.0 for d in days if d in ccni]
    ni_src = "ccni"
    if not ni_vals:
        # pre-CCNI weeks: bridge from the bulletin ex-tax series
        # (same cargoes, 5% VAT) plus the overlap-calibrated margin
        m = ni_bridge_margin(feeds)
        ext = ((feeds.get("oil_bulletin") or {})
               .get("roi_heating_gasoil_eur_per_1000l_ex_tax") or {})
        fxs = ((feeds.get("ecb_fx") or {}).get("eur_gbp_daily") or {})
        e_days = [d for d in sorted(ext) if d <= w_end]
        f_vals = [fxs[d] for d in days if d in fxs]
        if m is None or not e_days or not f_vals:
            return _skip(
                "NI oil: no CCNI reading and the bulletin bridge could "
                "not be built" + (" (no overlap margin)" if m is None else "")
                + ("" if e_days else " (no ex-tax bulletin week)")
                + ("" if f_vals else " (no FX for the week)"))
        fx_w = sum(f_vals) / len(f_vals)
        ni_vals = [ext[e_days[-1]] / 1000.0 * fx_w * 100.0 * 1.05 + m]
        ni_src = "bridged (bulletin ex-tax + calibrated margin, dagger)"
    bull = ((feeds.get("oil_bulletin") or {})
            .get("roi_heating_gasoil_eur_per_1000l") or {})
    bull_nt = ((feeds.get("oil_bulletin") or {})
               .get("roi_heating_gasoil_eur_per_1000l_ex_tax") or {})
    b_days = [d for d in sorted(bull) if d <= w_end]
    if not b_days:
        return _skip("ROI oil: no EU bulletin week at or before this date")
    fxs = ((feeds.get("ecb_fx") or {}).get("eur_gbp_daily") or {})
    fx_vals = [fxs[d] for d in days if d in fxs]
    fx = (sum(fx_vals) / len(fx_vals)) if fx_vals \
        else (feeds.get("ecb_fx") or {}).get("eur_gbp") or 0.855
    co2 = ((feeds.get("eirgrid") or {})
           .get("co2_intensity_g_per_kwh") or {})
    ci_vals = [co2[d] for d in days if d in co2]
    ef = round(sum(ci_vals) / len(ci_vals), 1) if len(ci_vals) >= 4 \
        else None
    return {"ni_oil_ppl": round(sum(ni_vals) / len(ni_vals), 2),
            "roi_oil_eur_1000l": bull[b_days[-1]],
            "fx": round(fx, 5), "ef_electricity": ef,
            "ef_source": ("weekly grid CI" if ef else "anchor"),
            "ni_oil_source": ni_src,
            # Reconstructed weeks are priced from published tariffs and
            # the hourly store rather than observed as they happened.
            # They are perfectly good weeks; they are just not LIVE
            # ones, and the milestone counts live.
            "live": w_end >= LIVE_FROM,
            "nondom": nondom_for(w_end, ((feeds.get("ecb_fx") or {})
                                         .get("eur_gbp_semester")))[0],
            "tariffs": tar}


# Wire format for the history block. Content is HISTORY_SCHEMA; this
# is how it is written, and the two are deliberately orthogonal - a
# re-encoding is not a restatement and must not trigger one.
HISTORY_ENCODING = "columnar-1"


def compact_history(entries):
    """
    Array of week objects -> columnar. Every key is written once
    instead of once per week, which is most of the file: the entries
    are wide and shallow, so at 52 weeks the key strings outweigh the
    numbers they label. Recurses into the ni/roi/fuels sub-blocks.
    """
    if not entries:
        return {"encoding": HISTORY_ENCODING, "n": 0, "cols": {}}

    def cols_of(objs):
        keys = []
        for o in objs:
            for k in o:
                if k not in keys:
                    keys.append(k)
        out = {}
        for k in keys:
            vals = [o.get(k) for o in objs]
            if any(isinstance(v, dict) for v in vals):
                out[k] = cols_of([v if isinstance(v, dict) else {}
                                  for v in vals])
            else:
                out[k] = vals
        return out

    return {"encoding": HISTORY_ENCODING, "n": len(entries),
            "cols": cols_of(entries)}


def expand_history(doc):
    """
    Columnar OR the legacy list -> list of week objects.

    Both shapes are accepted, and that is not politeness. index.html
    publishes the moment Pages deploys while data.json only changes at
    the next build, so for up to a day the new front end reads the old
    payload - and the pipeline reads its own previous output the same
    way. Either side must tolerate either shape.
    """
    if isinstance(doc, list):
        return doc
    if not isinstance(doc, dict) or not doc.get("cols"):
        return []
    n = int(doc.get("n") or 0)

    def rows(cols):
        out = [{} for _ in range(n)]
        for k, v in cols.items():
            if isinstance(v, dict):
                sub = rows(v)
                for i in range(n):
                    out[i][k] = sub[i]
            else:
                # Nulls are written back, not skipped. ef_electricity
                # is legitimately None in a week with too few carbon
                # observations, and dropping the key instead of the
                # value would make the round trip inexact for no gain.
                for i in range(n):
                    out[i][k] = v[i] if i < len(v) else None
        return out

    return rows(doc["cols"])


def build_history(feeds, anchors=None):
    """UK-pattern weekly history: complete calendar weeks (Mon-Sun),
    hero combined four + what-if twins per entry, frozen after the two
    most recent, capped at 60, append-or-update by week_ending. The
    Irish hero is anchor x HDD, so depth is bounded by the PRICE feeds
    (CCNI from 2026-02-26), not gas. Records the GNI feed's empirical
    window each run (gas_window) per the porting handover."""
    def fuel_sub(b):
        """Compact per-fuel in/useful for the windowed energy bars.
        Keys stay short - this rides in every history entry, three
        times over (island, NI, ROI)."""
        bf = b.get("by_fuel") or {}
        cl = b.get("cooling") or {}
        out = {f: {"i": round(v.get("in_gwh", 0.0), 1),
                   "u": round(v.get("useful_gwh", 0.0), 1)}
               for f, v in bf.items()}
        out["cool"] = {"i": round(cl.get("elec_gwh", 0.0), 1),
                       "u": round(cl.get("served_gwh",
                                         cl.get("elec_gwh", 0.0)), 1)}
        return out

    def jur_sub(b):
        # (JUR_SPLIT_KEYS is module-level so the tests can name it)
        return {
            "purchased_gwh": b["combined"]["purchased_gwh"],
            "served_gwh": b["combined"]["served_gwh"],
            "indigenous_pct": b["combined"]["indigenous_share_pct"],
            "bill_eur_m": b["combined"]["bill_eur_m"],
            "bill_gbp_m": b["combined"]["bill_gbp_m"],
            "emissions_kt": b["combined"]["emissions_kt_co2"],
            "fuels": fuel_sub(b),
            "wf_purchased_gwh": b["what_if_combined"]["purchased_gwh"],
            "wf_indigenous_pct":
                b["what_if_combined"]["indigenous_share_pct"],
            "wf_bill_eur_m": b["what_if_combined"]["bill_eur_m"],
            "wf_bill_gbp_m": b["what_if_combined"]["bill_gbp_m"],
            "wf_emissions_kt":
                b["what_if_combined"]["emissions_kt_co2"],
            # schema 5: the same heat/cold splits the island entry
            # carries, so a jurisdiction window can break its cards
            # down instead of falling back to a bare total. Sterling
            # AND euro, because the NI view prices in sterling and
            # must not convert components in the browser.
            **{k: b["combined"][k] for k in JUR_SPLIT_KEYS},
            **{"wf_" + k: b["what_if_combined"][k]
               for k in JUR_SPLIT_KEYS},
        }

    skips = []
    prev = list(expand_history((PREVIOUS_DERIVED or {})
                               .get("history") or []))
    prev_schema = int((PREVIOUS_DERIVED or {})
                      .get("history_schema") or 1)
    prev_epoch = int((PREVIOUS_DERIVED or {})
                     .get("anchor_epoch") or 1)
    re_anchor = prev_epoch < ANCHOR_EPOCH
    if re_anchor and prev:
        log(f"history: anchor epoch {prev_epoch} -> {ANCHOR_EPOCH}, "
            f"re-anchoring {len(prev)} stored weeks onto the new "
            f"basis (own-week prices retained)")
    frozen = {e["week_ending"]: e for e in prev[:-2]} if len(prev) > 2 \
        else {}
    hdd = (feeds.get("hdd") or {}).get("hdd_island") or {}
    if not hdd:
        return prev
    today = today_utc()
    last_sun = today - dt.timedelta(days=(today.isoweekday() % 7) or 7)
    out = []
    for k in range(59, -1, -1):
        w_end = (last_sun - dt.timedelta(weeks=k)).isoformat()
        if w_end in frozen:
            e = frozen[w_end]
            needs = (prev_schema < HISTORY_SCHEMA or re_anchor
                     or "ni" not in e or "roi" not in e)
            if needs:
                # Restatement (schema policy, 1 Aug 2026 handover):
                # recompute through derive_hero with the entry's
                # STORED inputs - ef_electricity injected so weeks
                # whose grid CI has rolled out of retention keep the
                # factor they were built with; oil, fx and tariffs
                # come from the retained series and the resolver.
                # Existing fields must re-round identically (audited
                # below); only new fields appear. Sub-block splits:
                # DEFERRED - combined-level is the like-for-like
                # scope; extend when a panel displays them.
                ctx = week_inputs(feeds, w_end)
                if ctx is not None:
                    if e.get("ef_electricity"):
                        ctx = {**ctx,
                               "ef_electricity": e["ef_electricity"],
                               "ef_source": e.get("ef_source",
                                                  ctx["ef_source"])}
                    h2 = derive_hero(feeds, anchors,
                                     week_ctx={"week_ending": w_end,
                                               **ctx})
                    if h2 is not None:
                        drift = max(
                            abs(h2["combined"]["purchased_gwh"]
                                - e["purchased_gwh"]),
                            abs(h2["combined"]["bill_eur_m"]
                                - e["bill_eur_m"]),
                            abs(h2["combined"]["emissions_kt_co2"]
                                - e["emissions_kt"]))
                        if drift > 0.5 and not re_anchor:
                            log(f"history: restatement drift {w_end} "
                                f"{drift:.1f} - stored values kept "
                                f"frozen, review")
                        e = dict(e)
                        C2, W2 = h2["combined"], h2["what_if_combined"]
                        if re_anchor:
                            # basis change: rewrite the stored values
                            # too, so totals and splits stay on one
                            # footing across the whole series
                            e.update({
                                "purchased_gwh": C2["purchased_gwh"],
                                "served_gwh": C2["served_gwh"],
                                "indigenous_pct":
                                    C2["indigenous_share_pct"],
                                "bill_eur_m": C2["bill_eur_m"],
                                "bill_gbp_m": C2["bill_gbp_m"],
                                "emissions_kt": C2["emissions_kt_co2"],
                                "wf_purchased_gwh": W2["purchased_gwh"],
                                "wf_indigenous_pct":
                                    W2["indigenous_share_pct"],
                                "wf_bill_eur_m": W2["bill_eur_m"],
                                "wf_bill_gbp_m": W2["bill_gbp_m"],
                                "wf_emissions_kt":
                                    W2["emissions_kt_co2"],
                                "ni": jur_sub(h2["ni"]),
                                "roi": jur_sub(h2["roi"]),
                                "fuels": fuel_sub(h2),
                            })
                        for k in ("heat_gwh", "cold_gwh",
                                  "bill_heat_eur_m", "bill_cold_eur_m",
                                  "bill_heat_gbp_m", "bill_cold_gbp_m",
                                  "emissions_heat_kt",
                                  "emissions_cold_kt"):
                            e[k] = C2[k]
                            e["wf_" + k] = W2[k]
                        e.setdefault("fx_eur_gbp", ctx["fx"])
                        # Assign, never setdefault: a block that
                        # already exists from an older schema must be
                        # REFRESHED so it gains new sub-fields. This
                        # is safe and idempotent because everything
                        # is recomputed from the entry's own stored
                        # inputs.
                        e["fuels"] = fuel_sub(h2)
                        # NOT setdefault: these keys already exist
                        # from earlier schemas, so a defaulting write
                        # silently skipped the per-fuel addition.
                        # A schema migration refreshes the block.
                        e["ni"] = jur_sub(h2["ni"])
                        e["roi"] = jur_sub(h2["roi"])
                        e["live"] = w_end >= LIVE_FROM
                    else:
                        log(f"history: {w_end} unrecomputable - "
                            f"kept at prior schema (caption degrades)")
                else:
                    log(f"history: {w_end} outside recompute window - "
                        f"kept at prior schema (caption degrades)")
            out.append(e)
            continue
        ctx = week_inputs(feeds, w_end, skips)
        if ctx is None:
            continue
        h = derive_hero(feeds, anchors, week_ctx={"week_ending": w_end,
                                                  **ctx})
        if h is None:
            skips.append((w_end, "derive_hero declined - check the HDD "
                                 "series has 200+ days"))
            continue
        C, WFC = h["combined"], h["what_if_combined"]
        out.append({
            "week_ending": w_end,
            "purchased_gwh": h["combined"]["purchased_gwh"],
            "served_gwh": h["combined"]["served_gwh"],
            "indigenous_pct": h["combined"]["indigenous_share_pct"],
            "bill_eur_m": h["combined"]["bill_eur_m"],
            "bill_gbp_m": h["combined"]["bill_gbp_m"],
            "emissions_kt": h["combined"]["emissions_kt_co2"],
            "wf_purchased_gwh": h["what_if_combined"]["purchased_gwh"],
            "wf_indigenous_pct":
                h["what_if_combined"]["indigenous_share_pct"],
            "wf_bill_eur_m": h["what_if_combined"]["bill_eur_m"],
            "wf_bill_gbp_m": h["what_if_combined"]["bill_gbp_m"],
            "wf_emissions_kt":
                h["what_if_combined"]["emissions_kt_co2"],
            "hdd": h["hdd_week"],
            "ef_electricity": ctx["ef_electricity"] or None,
            "ef_source": ctx["ef_source"],
            "heat_gwh": C["heat_gwh"], "cold_gwh": C["cold_gwh"],
            "bill_heat_eur_m": C["bill_heat_eur_m"],
            "bill_cold_eur_m": C["bill_cold_eur_m"],
            "bill_heat_gbp_m": C["bill_heat_gbp_m"],
            "bill_cold_gbp_m": C["bill_cold_gbp_m"],
            "emissions_heat_kt": C["emissions_heat_kt"],
            "emissions_cold_kt": C["emissions_cold_kt"],
            "wf_heat_gwh": WFC["heat_gwh"],
            "wf_cold_gwh": WFC["cold_gwh"],
            "wf_bill_heat_eur_m": WFC["bill_heat_eur_m"],
            "wf_bill_cold_eur_m": WFC["bill_cold_eur_m"],
            "wf_bill_heat_gbp_m": WFC["bill_heat_gbp_m"],
            "wf_bill_cold_gbp_m": WFC["bill_cold_gbp_m"],
            "wf_emissions_heat_kt": WFC["emissions_heat_kt"],
            "wf_emissions_cold_kt": WFC["emissions_cold_kt"],
            "fx_eur_gbp": ctx["fx"],
            "fuels": fuel_sub(h),
            "ni_oil_source": ctx["ni_oil_source"],
            "live": ctx["live"],
            "ni": jur_sub(h["ni"]),
            "roi": jur_sub(h["roi"]),
        })
    report_skips(skips, len(out))
    return out[-HISTORY_MAX:]


def derive_hero(feeds, anchors=None, week_ctx=None):
    """
    Weekly hero four-stat + what-if, per jurisdiction and all-island.
    Each jurisdiction is shaped by its own HDD series (island fallback);
    the island block is the sum of the two, so the toggle views always
    reconcile. Scaffold estimator - pure, unit tested. Top-level keys
    carry the island values; per-jurisdiction blocks under "roi"/"ni".
    """
    a = anchors or ANCHORS
    hddf = feeds.get("hdd") or {}
    island_hdd = hddf.get("hdd_island") or {}
    if len(island_hdd) < 200:
        return None
    W = week_ctx or {}
    fx = W.get("fx") or (feeds.get("ecb_fx") or {}).get("eur_gbp") or 0.855
    shf = a["space_heat_fraction"]

    if W:
        # historic week: prices, fx, EF and tariffs from the week itself
        oil_eur_kwh = W["roi_oil_eur_1000l"] / 1000.0 \
            / a["kerosene_kwh_per_litre"]
        oil_gbp_kwh = W["ni_oil_ppl"] / 100.0 \
            / a["kerosene_kwh_per_litre"]
    else:
        oil_eur_kwh = None
        ob = (feeds.get("oil_bulletin") or {}).get("latest_value")
        if ob:
            oil_eur_kwh = ob / 1000.0 / a["kerosene_kwh_per_litre"]
        oil_gbp_kwh = None
        ccni = ((feeds.get("ccni_oil") or {}).get("series_gbp") or {}).get(
            "daily", {}).get("900l") or {}
        if ccni:
            oil_gbp_kwh = ccni[max(ccni)] / 900.0 \
                / a["kerosene_kwh_per_litre"]

    def hdd_stats(series):
        days = sorted(series)
        if W:
            w_end = W["week_ending"]
            wk = [d for d in days if d <= w_end][-7:]
            yr = [d for d in days if d <= w_end][-365:]
            if len(wk) < 7 or len(yr) < 200:
                return (0.0, 0.0, w_end)
            return (sum(series[d] for d in wk),
                    sum(series[d] for d in yr), w_end)
        wk = days[-7:]
        return (sum(series[d] for d in wk),
                sum(series[d] for d in days[-365:]), wk[-1])

    def jur_block(jur, cur, hdd_series):
        hdd_week, hdd_year, week_end = hdd_stats(hdd_series)
        if hdd_year <= 0:
            return None
        j = a[jur]
        heat_twh = j["residential_heat_twh"] + j["services_heat_twh"]

        def week_input_gwh(annual_twh):
            annual_gwh = annual_twh * 1000.0
            return annual_gwh * ((1 - shf) / 52.0
                                 + shf * hdd_week / hdd_year)

        # Heat-pump share of the electricity line. The stock is an
        # annual quantity; its delivered heat is shaped by the same
        # weekly profile as everything else, so the split holds
        # week to week. hp_useful is delivered heat; hp_elec is the
        # purchased part; the remainder is ambient heat harvested
        # from the environment - free, and never purchased.
        hpa = a.get("heat_pumps") or {}
        hpj = hpa.get(jur) or {}
        hp_annual_twh = (
            (hpj.get("households", 0) * hpj.get("census_uplift", 1.0)
             + hpj.get("nondom_equivalent", 0))
            * hpa.get("delivered_mwh_per_dwelling", 9.0) / 1e6)
        spf = j.get("ashp_climate_spf") or a["ashp"].get("climate_spf") \
            or 2.85
        hp_useful = week_input_gwh(hp_annual_twh)
        hp_elec = hp_useful / spf
        hp_ambient = hp_useful - hp_elec

        inp_t = useful_t = indig_t = kt_t = bill_t = 0.0
        by_fuel = {}
        for fuel, share in j["fuel_shares"].items():
            inp = week_input_gwh(heat_twh) * share
            eff = a["efficiency"][fuel]
            useful = inp * eff
            if fuel == "electricity":
                # split the line: resistive keeps the remainder, the
                # heat-pump part is carried separately, and the
                # ambient harvest is an out-bar-only entry (no input)
                hp_in = min(hp_elec, inp)
                by_fuel["heatpump"] = {"in_gwh": round(hp_in, 1),
                                       "useful_gwh": round(hp_in, 1)}
                by_fuel["ambient"] = {
                    "in_gwh": 0.0,
                    "useful_gwh": round(hp_in * (spf - 1.0), 1)}
                inp = inp - hp_in
                useful = inp * eff
            by_fuel[fuel] = {"in_gwh": round(inp, 1),
                             "useful_gwh": round(useful, 1)}
            indig = useful * (
                j["gas_indigenous"] if fuel == "gas" else
                j["elec_indigenous"] if fuel == "electricity" else
                a["indigenous"].get(fuel, 0.0))
            kt = inp * EF[fuel] / 1000.0
            if fuel == "oil":
                price = oil_eur_kwh if cur == "eur" else oil_gbp_kwh
                if price is None:
                    price = 0.11 if cur == "eur" else 0.09  # dagger fallback
            else:
                table = ((W.get("tariffs") or {}).get(cur)
                         or (a["retail_eur_per_kwh"] if cur == "eur"
                             else a["retail_gbp_per_kwh"]))
                price = table.get(fuel, table["gas"])
                if fuel in ("gas", "electricity"):
                    nd = _nondom(a, W)["eur" if cur == "eur"
                                        else "gbp"][fuel]
                    ds = a["dom_share"][jur][fuel]
                    price = ds * price + (1 - ds) * nd
            inp_t += inp
            useful_t += useful
            if fuel == "electricity":
                # heat-pump electricity and its ambient harvest join
                # the totals: purchased counts only the electricity,
                # delivered counts electricity x SPF
                inp_t += by_fuel["heatpump"]["in_gwh"]
                useful_t += (by_fuel["heatpump"]["useful_gwh"]
                             + by_fuel["ambient"]["useful_gwh"])
            indig_t += indig
            kt_t += kt
            bill_t += inp * price   # GWh x cur/kWh = millions of cur

        bill_eur = bill_t if cur == "eur" else bill_t / fx
        bill_gbp = bill_t * fx if cur == "eur" else bill_t

        dom_elec = ((W.get("tariffs") or {}).get(cur, {})
                    .get("electricity")
                    or (a["retail_eur_per_kwh"]["electricity"]
                        if cur == "eur"
                        else a["retail_gbp_per_kwh"]["electricity"]))
        nondom_elec = _nondom(a, W)["eur" if cur == "eur"
                                    else "gbp"]["electricity"]
        # cooling is wholly non-domestic (UK convention)
        elec_price = nondom_elec
        # what-if heat-pump electricity blends at the domestic share of
        # delivered heat (UK convention)
        heat_dom = j["residential_heat_twh"] / max(
            j["residential_heat_twh"] + j["services_heat_twh"], 1e-9)
        wf_elec_price = heat_dom * dom_elec + (1 - heat_dom) * nondom_elec

        # what-if: 20% of useful heat moves to geothermal heat pumps
        spf = a["geothermal_spf"]
        moved = useful_t * 0.20
        elec_in = moved / spf
        ambient = moved - elec_in
        scale = 0.80
        wf_bill_native = bill_t * scale + elec_in * wf_elec_price
        wf = {
            "heat_purchased_gwh": round(inp_t * scale + elec_in, 1),
            "indigenous_share_pct": round(100 * (
                indig_t * scale + ambient
                + elec_in * j["elec_indigenous"]) / max(useful_t, 1e-9), 1),
            "bill_eur_m": round(wf_bill_native if cur == "eur"
                                else wf_bill_native / fx, 1),
            "bill_gbp_m": round(wf_bill_native * fx if cur == "eur"
                                else wf_bill_native, 1),
            "emissions_kt_co2": round(
                kt_t * scale
                + elec_in * EF["electricity"] / 1000.0, 1),
            "geothermal_spf": spf,
        }
        # cooling: cold-economy census, weekly. Flat loads spread
        # evenly; the ROI comfort load follows the live ODH26 record
        # (30% dagger ventilation floor) once a season of it exists.
        cc = a["cool"]
        if jur == "roi":
            dc = (cc["roi_elec_twh"] * cc["dc_share_of_roi_elec"]
                  * cc["dc_cooling_share"])
            l = cc["loads_twh"]
            comfort_a = l["comfort"]
            flat_a = dc + l["refrigeration"] + l["process"]
        else:
            comfort_a, flat_a = 0.0, cc["loads_twh"]["ni_all"]
        sf = cc["cooling_service_factor"]
        if jur == "roi":
            l2 = dict(l); l2["dc"] = dc
            num = (l2["dc"] * sf["dc"] + l2["refrigeration"]
                   * sf["refrigeration"] + l2["process"] * sf["process"]
                   + l2["comfort"] * sf["comfort"])
            blend = num / (l2["dc"] + l2["refrigeration"]
                           + l2["process"] + l2["comfort"])
        else:
            blend = sf["ni_all"]
        cool_week = flat_a * 1000.0 / 52.0
        if comfort_a > 0:
            frac = odh_frac if odh_frac is not None else 1.0 / 52.0
            cool_week += comfort_a * 1000.0 * (0.3 / 52.0 + 0.7 * frac)
        cool_bill_native = cool_week * elec_price
        cool_kt = cool_week * EF["electricity"] / 1000.0
        cool_indig = cool_week * j["elec_indigenous"]
        cool_served = cool_week * blend
        cooling = {
            "elec_gwh": round(cool_week, 1),
            "served_gwh": round(cool_served, 1),
            "service_factor": round(blend, 2),
            "bill_eur_m": round(cool_bill_native if cur == "eur"
                                else cool_bill_native / fx, 1),
            "bill_gbp_m": round(cool_bill_native * fx if cur == "eur"
                                else cool_bill_native, 1),
            "emissions_kt_co2": round(cool_kt, 1),
            "comfort_shaped_by_odh": bool(comfort_a > 0
                                          and odh_frac is not None),
        }
        combined = {
            "purchased_gwh": round(inp_t + cool_week, 1),
            "served_gwh": round(useful_t + cool_served, 1),
            "bill_eur_m": round((bill_t if cur == "eur"
                                 else bill_t / fx)
                                + cooling["bill_eur_m"], 1),
            "bill_gbp_m": round((bill_t * fx if cur == "eur"
                                 else bill_t)
                                + cooling["bill_gbp_m"], 1),
            "emissions_kt_co2": round(kt_t + cool_kt, 1),
            # heat/cold splits (schema 2): cold = the cold-economy
            # electricity; heat = the fuel side. Pairs reconcile per
            # currency within rounding.
            "heat_gwh": round(inp_t, 1),
            "cold_gwh": round(cool_week, 1),
            "bill_heat_eur_m": round(bill_t if cur == "eur"
                                     else bill_t / fx, 1),
            "bill_cold_eur_m": cooling["bill_eur_m"],
            "bill_heat_gbp_m": round(bill_t * fx if cur == "eur"
                                     else bill_t, 1),
            "bill_cold_gbp_m": cooling["bill_gbp_m"],
            "emissions_heat_kt": round(kt_t, 1),
            "emissions_cold_kt": round(cool_kt, 1),
            # delivered basis: purchased electricity carries the grid
            # share; the balance of the cooling service is ambient
            # rejection, indigenous by convention
            "indigenous_share_pct": round(100 * (
                indig_t + cool_indig + (cool_served - cool_week))
                / max(useful_t + cool_served, 1e-9), 1),
        }

        # peak winter week for the for-scale line
        days = sorted(hdd_series)[-365:]
        pk, pk_end = 0.0, None
        for i in range(6, len(days)):
            w = sum(hdd_series[d] for d in days[i - 6:i + 1])
            if w > pk:
                pk, pk_end = w, days[i]
        peak = None
        if pk_end and hdd_year > 0:
            annual_gwh = heat_twh * 1000.0
            peak = {"week_ending": pk_end, "hdd": round(pk, 1),
                    "heat_purchased_gwh": round(
                        annual_gwh * ((1 - shf) / 52.0
                                      + shf * pk / hdd_year), 1)}
        # combined what-if: heat side as computed in wf; cooling side
        # moves 20% of load to ground-coupled systems at the dagger
        # saving factor - service unchanged, electricity cut, the
        # avoided share ambient and indigenous
        save = cc["ground_cooling_saving"]
        cool_wf_elec = cool_week * (1 - 0.20 * save)
        wf_cool_bill = cool_wf_elec * elec_price
        wf_heat_indig_abs = wf["indigenous_share_pct"] / 100.0 * useful_t
        wf_combined = {
            "purchased_gwh": round(wf["heat_purchased_gwh"]
                                   + cool_wf_elec, 1),
            "bill_eur_m": round(wf["bill_eur_m"]
                                + (wf_cool_bill if cur == "eur"
                                   else wf_cool_bill / fx), 1),
            "bill_gbp_m": round(wf["bill_gbp_m"]
                                + (wf_cool_bill * fx if cur == "eur"
                                   else wf_cool_bill), 1),
            "emissions_kt_co2": round(
                wf["emissions_kt_co2"]
                + cool_wf_elec * EF["electricity"] / 1000.0, 1),
            "indigenous_share_pct": round(100 * (
                wf_heat_indig_abs
                + cool_wf_elec * j["elec_indigenous"]
                + (cool_served - cool_wf_elec))
                / max(useful_t + cool_served, 1e-9), 1),
            # what-if splits: cold at (1 - R x ground_cooling_saving)
            # - deliberately NOT the UK's COP-20 arithmetic, per the
            # what-if parameters addendum to the cross-calibration
            "heat_gwh": round(wf["heat_purchased_gwh"], 1),
            "cold_gwh": round(cool_wf_elec, 1),
            "bill_heat_eur_m": round(wf["bill_eur_m"], 1),
            "bill_cold_eur_m": round(wf_cool_bill if cur == "eur"
                                     else wf_cool_bill / fx, 1),
            "bill_heat_gbp_m": round(wf["bill_gbp_m"], 1),
            "bill_cold_gbp_m": round(wf_cool_bill * fx if cur == "eur"
                                     else wf_cool_bill, 1),
            "emissions_heat_kt": round(wf["emissions_kt_co2"], 1),
            "emissions_cold_kt": round(
                cool_wf_elec * EF["electricity"] / 1000.0, 1),
        }
        return {
            "heat_purchased_gwh": round(inp_t, 1),
            "heat_delivered_gwh": round(useful_t, 1),
            "cooling": cooling,
            "combined": combined,
            "what_if_combined": wf_combined,
            "by_fuel": by_fuel,
            "peak_week": peak,
            "indigenous_share_pct": round(
                100 * indig_t / max(useful_t, 1e-9), 1),
            "bill_eur_m": round(bill_eur, 1),
            "bill_gbp_m": round(bill_gbp, 1),
            "emissions_kt_co2": round(kt_t, 1),
            "hdd_week": round(hdd_week, 1),
            "hdd_year": round(hdd_year, 1),
            "week_ending": week_end,
            "what_if_20pct_geothermal": wf,
            "_raw": {"useful": useful_t, "indig": indig_t},
        }

    # Weekly comfort-cooling share of the year. The fraction is only
    # meaningful against a FULL year of overheating: with a partial
    # series the denominator omits the rest of the season and the
    # summer weeks inflate (observed 18 Jul 2026: a 61-day series made
    # one July week ~14% of "annual"). Hero shaping therefore requires
    # >=300 days; until then comfort is flat here. The cold-economy
    # panel keeps its 60-day rule - its supply and demand shapes share
    # one window, so partial coverage stays internally consistent.
    # electricity emission factor: live all-island grid intensity
    # (trailing 14-day mean) once at least 7 days exist; anchor
    # otherwise. The 280 g anchor predates the EirGrid restoration;
    # live summer intensity runs ~210-220 g.
    # Electricity indigenous share: HELD AT ANCHOR. The sem_mix live
    # share failed its cross-examination on first adoption (30 Jul
    # 2026): 14-day mean 31.2% against anchors of 41.3/46% during a
    # fortnight whose live grid intensity (~179 g/kWh) implies ~half
    # of generation zero-carbon. Two identified biases: the IE feed
    # has no Solar category, and the cross-border sign convention is
    # unverified (exports may be inflating the denominator). The feed
    # keeps collecting; the diagnostic below logs the divergence each
    # run; the anchor is not replaced until the live share reconciles
    # with the CI evidence.
    ind_src = "anchor (SEAI RES-E / DfE)"
    if not W:
        sm = ((feeds.get("sem_mix") or {})
              .get("indigenous_share_daily") or {})
        sdays = sorted(sm)[-14:]
        if len(sdays) >= 7:
            live_ind = sum(sm[d] for d in sdays) / len(sdays)
            co2d = ((feeds.get("eirgrid") or {})
                    .get("co2_intensity_g_per_kwh") or {})
            cdd = sorted(co2d)[-14:]
            ci_note = ""
            if len(cdd) >= 7:
                ci = sum(co2d[d] for d in cdd) / len(cdd)
                # CI-implied zero-carbon share against a ~420 g/kWh
                # dagger fossil-fleet average - coarse, diagnostic only
                zc = max(0.0, min(1.0, 1.0 - ci / 420.0)) * 100.0
                ci_note = f"; CI-implied zero-carbon ~{zc:.0f}%"
            log(f"sem_mix diagnostic: 14d live share {live_ind:.1f}% "
                f"vs anchors ROI 41.3 / NI 46{ci_note} - held at "
                f"anchor pending solar/sign validation")
        log(f"hero: elec indigenous share - {ind_src}")

    EF = dict(a["ef_g_per_kwh"])
    if W:
        ef_src = W["ef_source"]
        if W.get("ef_electricity"):
            EF["electricity"] = W["ef_electricity"]
    else:
        co2 = ((feeds.get("eirgrid") or {})
               .get("co2_intensity_g_per_kwh") or {})
        cdays = sorted(co2)[-14:]
        ef_src = "anchor"
        if len(cdays) >= 7:
            EF["electricity"] = round(
                sum(co2[d] for d in cdays) / len(cdays), 1)
            ef_src = f"live grid intensity, {len(cdays)}-day mean"
        log(f"hero: electricity EF {EF['electricity']} g/kWh ({ef_src})")

    odh = {} if W else (hddf.get("odh26_island") or {})
    odh_frac = None
    odays = sorted(odh)[-365:]
    if len(odays) >= 300:
        oy = sum(odh[d] for d in odays)
        if oy > 0:
            odh_frac = sum(odh[d] for d in odays[-7:]) / oy

    roi = jur_block("roi", "eur", hddf.get("hdd_roi") or island_hdd)
    ni = jur_block("ni", "gbp", hddf.get("hdd_ni") or island_hdd)
    if not (roi and ni):
        return None

    def sum_blocks(x, y):
        useful = x["_raw"]["useful"] + y["_raw"]["useful"]
        indig = x["_raw"]["indig"] + y["_raw"]["indig"]
        bf = {}
        for src_b in (x, y):
            for f, v in (src_b.get("by_fuel") or {}).items():
                slot = bf.setdefault(f, {"in_gwh": 0.0, "useful_gwh": 0.0})
                slot["in_gwh"] = round(slot["in_gwh"] + v["in_gwh"], 1)
                slot["useful_gwh"] = round(
                    slot["useful_gwh"] + v["useful_gwh"], 1)
        pk = None
        if x.get("peak_week") and y.get("peak_week"):
            pk = {"week_ending": max(x["peak_week"]["week_ending"],
                                     y["peak_week"]["week_ending"]),
                  "heat_purchased_gwh": round(
                      x["peak_week"]["heat_purchased_gwh"]
                      + y["peak_week"]["heat_purchased_gwh"], 1)}
        def addd(key, fields):
            return {f: round(x[key][f] + y[key][f], 1) for f in fields}
        cooling = addd("cooling", ("elec_gwh", "served_gwh",
                                   "bill_eur_m", "bill_gbp_m",
                                   "emissions_kt_co2"))
        cooling["comfort_shaped_by_odh"] = (
            x["cooling"]["comfort_shaped_by_odh"]
            or y["cooling"]["comfort_shaped_by_odh"])
        combined = addd("combined", ("purchased_gwh", "served_gwh",
                                     "bill_eur_m", "bill_gbp_m",
                                     "emissions_kt_co2",
                                     "heat_gwh", "cold_gwh", "bill_heat_eur_m", "bill_cold_eur_m", "bill_heat_gbp_m", "bill_cold_gbp_m", "emissions_heat_kt", "emissions_cold_kt"))
        served = x["combined"]["served_gwh"] + y["combined"]["served_gwh"]
        combined["indigenous_share_pct"] = round(
            (x["combined"]["indigenous_share_pct"]
             * x["combined"]["served_gwh"]
             + y["combined"]["indigenous_share_pct"]
             * y["combined"]["served_gwh"]) / max(served, 1e-9), 1)
        wfc = addd("what_if_combined", ("purchased_gwh", "bill_eur_m",
                                        "bill_gbp_m", "emissions_kt_co2",
                                        "heat_gwh", "cold_gwh", "bill_heat_eur_m", "bill_cold_eur_m", "bill_heat_gbp_m", "bill_cold_gbp_m", "emissions_heat_kt", "emissions_cold_kt"))
        wfc["indigenous_share_pct"] = round(
            (x["what_if_combined"]["indigenous_share_pct"]
             * x["combined"]["served_gwh"]
             + y["what_if_combined"]["indigenous_share_pct"]
             * y["combined"]["served_gwh"]) / max(served, 1e-9), 1)
        out = {
            "cooling": cooling,
            "combined": combined,
            "what_if_combined": wfc,
            "by_fuel": bf,
            "peak_week": pk,
            "heat_purchased_gwh": round(
                x["heat_purchased_gwh"] + y["heat_purchased_gwh"], 1),
            "heat_delivered_gwh": round(
                x["heat_delivered_gwh"] + y["heat_delivered_gwh"], 1),
            "indigenous_share_pct": round(100 * indig / max(useful, 1e-9), 1),
            "bill_eur_m": round(x["bill_eur_m"] + y["bill_eur_m"], 1),
            "bill_gbp_m": round(x["bill_gbp_m"] + y["bill_gbp_m"], 1),
            "emissions_kt_co2": round(
                x["emissions_kt_co2"] + y["emissions_kt_co2"], 1),
        }
        wf = {}
        for k in ("heat_purchased_gwh", "bill_eur_m", "bill_gbp_m",
                  "emissions_kt_co2"):
            wf[k] = round(x["what_if_20pct_geothermal"][k]
                          + y["what_if_20pct_geothermal"][k], 1)
        wf["geothermal_spf"] = a["geothermal_spf"]
        wf["indigenous_share_pct"] = round(
            (x["what_if_20pct_geothermal"]["indigenous_share_pct"]
             * x["_raw"]["useful"]
             + y["what_if_20pct_geothermal"]["indigenous_share_pct"]
             * y["_raw"]["useful"]) / max(useful, 1e-9), 1)
        out["what_if_20pct_geothermal"] = wf
        return out

    island = sum_blocks(roi, ni)
    hdd_week, hdd_year, week_end = hdd_stats(island_hdd)
    for b in (roi, ni):
        b.pop("_raw", None)

    out = dict(island)
    out.update({
        "week_ending": week_end,
        "hdd_week": round(hdd_week, 1), "hdd_year": round(hdd_year, 1),
        "roi": roi, "ni": ni,
        "ef_electricity_g_per_kwh": EF["electricity"],
        "ef_electricity_source": ef_src,
        "elec_indigenous_source": ind_src,
        "basis": ("Scaffold estimator (dagger throughout) - annual anchors "
                  "shaped by each jurisdiction's weekly HDD; SEAI 2024, "
                  "DfE/NISRA, Causeway estimates. Oil, the island's "
                  "majority heating fuel, is modelled from annual "
                  "anchors, not metered - its weekly estimates carry a "
                  "materially wider band than gas. Hot water is carried as a flat "
                  "22.4% of annual (SEAI National Heat Study, Aug 2026 "
                  "re-anchoring); space heat follows the week's "
                  "degree days. Bills are sector-blended: services gas and "
                  "electricity, and all cooling, price at non-domestic "
                  "rates (dagger, pending Eurostat band prices); oil "
                  "prices identically across sectors. VAT convention: "
                  "domestic rates include VAT (5% NI, 9% ROI), "
                  "non-domestic rates exclude it, because businesses "
                  "recover input VAT - Eurostat level 3 and level 2 "
                  "respectively. NI gas and electricity are all-in "
                  "rates at the Utility Regulator's own consumption "
                  "basis, gas weighted across SSE Airtricity (Greater "
                  "Belfast and West) and Firmus (Ten Towns) by "
                  "regulated customer count; ROI is a standard unit "
                  "rate without standing charges, so the two "
                  "jurisdictions are internally consistent but not "
                  "like-for-like at component level (open item). The "
                  "Domestic rates in BOTH jurisdictions are "
                  "all-in effective rates at a stated consumption, "
                  "including VAT and standing charges - NI from the "
                  "Utility Regulator's regulated bills, ROI from the "
                  "Eurostat band price (S2 2024) stepped by the "
                  "supplier announcements. The residual difference is "
                  "scope rather than basis: NI is incumbent-weighted "
                  "regulated, ROI a market-wide average including "
                  "discounts, and the two do not bias in one "
                  "direction (dagger). Cooling is the "
                  "cold-economy "
                  "census (dagger loads beside the CSO data-centre "
                  "anchor), flat across the year with the comfort share "
                  "following live overheating degree-hours once a season "
                  "exists - a cold-economy scope, wider than comfort-only national "
                  "cooling lines and not one-to-one comparable with them; the "
                  "data-centre line counts cooling electricity (~14% of "
                  "the fleet's draw, SEAI), not the whole draw, and "
                  "delivered cooling uses the SEAI National Heat "
                  "Study's useful-to-final ratios. "
                  "Electricity emissions use the live all-island grid "
                  "intensity when available. Challenge and input "
                  "welcome at contact@causewaygt.com"),
        "anchors_used": a,
    })
    return out


def derive_ashp_spf(hdd_daily: dict, anchors=None):
    """
    Air-source heat pump seasonal performance from the HDD series itself.
    For heating days T_out = base - HDD, so the demand-weighted source
    temperature is base - sum(h^2)/sum(h) over the trailing year. COP is a
    Carnot fraction at 45C flow with a defrost derate, blended (harmonic,
    energy-weighted) with a DHW share at 55C. All parameters dagger - see
    ANCHORS["ashp"]. Returns None without a season of data.
    """
    a = (anchors or ANCHORS)
    p = a["ashp"]
    days = sorted(hdd_daily)[-365:]
    hs = [hdd_daily[d] for d in days if hdd_daily[d] > 0]
    if len(hs) < 60 or sum(hs) < 200:
        return None
    w = sum(hs)
    t_src = HDD_BASE_C - sum(h * h for h in hs) / w

    def cop(source_c, flow_c, derate=1.0):
        lift = flow_c - source_c
        if lift <= 5:
            lift = 5.0
        return p["carnot_fraction"] * (flow_c + 273.15) / lift * derate

    space = cop(t_src, p["flow_c"], p["defrost_derate"])
    dhw = cop(p["dhw_source_c"], p["dhw_flow_c"])
    sh = p["dhw_share"]
    spf = 1.0 / ((1 - sh) / space + sh / dhw)
    return {"spf": round(spf, 2),
            "demand_weighted_source_c": round(t_src, 1),
            "space_cop": round(space, 2), "dhw_cop": round(dhw, 2),
            "params": p}


# ------------------------------------------------------- Phase B.2.1
# The de-rated all-island dispatchable block, MW. AIRAA Appendix 3
# registered capacity x Table 5.18 availability factors, of which
# ~1,490 MW is run-hour-limited. Dagger. Observed all-island peak was
# 7,502 MW on 8 Jan 2025, which is the sanity rail: a computed demand
# far above that on a mild hour means the shaping is wrong, not that
# the island is short of plant.
# ROUTE TIERS - aligned with the UK sibling, 8 Aug 2026, so the two
# grid layers are read against the same ladder. Field-observed
# in-situ figures, not brochure SCOPs:
#   ASHP  2.80  Energy Systems Catapult, Electrification of Heat median
#   GSHP  3.24  Energy Systems Catapult, in-situ GSHP average
#   NET   5.00  networked geothermal on a shared ambient loop
# The Irish site keeps its own Carnot-fraction engine for the AIR
# route, because the whole point of that column is that air-source
# performance collapses in the hour that binds - a flat 2.80 would
# hide exactly the effect being measured. ASHP_SPF is therefore the
# seasonal comparator, not the hourly one, and the two are logged
# side by side. Ground and network are flat by construction: a
# borehole field does not care what the air is doing.
ASHP_SPF = 2.80
GSHP_SPF = 3.24
GEO_NETWORK_SCOP = 5.00

GRID_BLOCK_MW = 8595
GRID_RUN_HOUR_LIMITED_MW = 1490
OBSERVED_PEAK_MW = 7502
OBSERVED_PEAK_AT = "2025-01-08"


def hourly_heat_mw(store, anchors=None):
    """
    Island building-heat demand per hour, MW of USEFUL heat.

    Shaped the way the weekly hero shapes a week, one level finer:
    hot water flat, space heat following each hour's share of the
    store's own heating-degree total. Degree hours come from temp_ai,
    so this is the island's own weather rather than a profile.

    Returns {hour_key: MW} or {} if the store cannot support it.
    """
    a = anchors or ANCHORS
    temps = (expand_hourly(store) or {}).get("temp_ai") or {}
    if len(temps) < 24 * 300:
        return {}
    # useful heat, not input: the anchors are fuel input and each fuel
    # burns at its own efficiency, so convert before shaping.
    useful_twh = 0.0
    for jur in ("ni", "roi"):
        j = a[jur]
        heat = j["residential_heat_twh"] + j["services_heat_twh"]
        useful_twh += heat * sum(sh * a["efficiency"][f]
                                 for f, sh in j["fuel_shares"].items())
    n = len(temps)
    dh = {k: max(0.0, HDD_BASE_C - v) for k, v in temps.items()}
    total_dh = sum(dh.values())
    if total_dh <= 0:
        return {}
    # The store is ~13 months, so scale the annual anchor to its span
    # rather than assuming a calendar year.
    span_years = n / (365.25 * 24)
    dhw_mwh = useful_twh * (1 - a["space_heat_fraction"]) * 1e6 * span_years
    space_mwh = useful_twh * a["space_heat_fraction"] * 1e6 * span_years
    flat = dhw_mwh / n
    return {k: flat + space_mwh * dh[k] / total_dh for k in temps}


def derive_tightest_hour(store, feeds=None, anchors=None):
    """
    B.2.1 - the tightest hour. Log-only; nothing draws from it yet,
    by design: the three B.2 computations are run and logged before
    any panel is written, so the headline is chosen after the numbers
    are seen rather than before.

    Electrified heat is added to OBSERVED demand net of the resistive
    heating it displaces - that share is already in the demand series,
    so counting the heat pump without removing the immersion would
    double-count it.
    """
    a = anchors or ANCHORS
    heat = hourly_heat_mw(store, a)
    if not heat:
        return None
    ser = expand_hourly(store) or {}
    demand, temps = ser.get("demand_ai") or {}, ser.get("temp_ai") or {}
    hours = sorted(set(heat) & set(demand))
    if len(hours) < 24 * 300:
        return None

    p = ANCHORS["ashp"]

    def cop_at(t):
        lift = max(5.0, p["flow_c"] - t)
        return p["carnot_fraction"] * (p["flow_c"] + 273.15) / lift \
            * p["defrost_derate"]

    # resistive share of useful heat, from the fuel shares
    res = 0.0
    for jur in ("ni", "roi"):
        j = a[jur]
        h = j["residential_heat_twh"] + j["services_heat_twh"]
        res += h * j["fuel_shares"].get("electricity", 0.0) \
            * a["efficiency"].get("electricity", 1.0)
    tot = sum((a[j]["residential_heat_twh"] + a[j]["services_heat_twh"])
              * sum(sh * a["efficiency"][f]
                    for f, sh in a[j]["fuel_shares"].items())
              for j in ("ni", "roi"))
    res_share = (res / tot) if tot else 0.0

    wind = ser.get("wind_ai") or {}
    solar = ser.get("solar_ai") or {}

    def ceiling_at(key):
        """The fleet that is actually there in that hour.

        De-rated dispatchable capacity PLUS the wind and solar the
        island actually generated - the ceiling breathes with the
        weather, because the hour that binds is cold AND still AND
        dark, and a flat block would credit wind that was not blowing
        or ignore wind that was. This is also what makes the figure
        comparable with the UK sibling, whose ceiling is defined the
        same way.
        """
        return GRID_BLOCK_MW + (wind.get(key) or 0.0) + (solar.get(key) or 0.0)

    # --- how far can heat be electrified before the fleet fills?
    # Solving for the SHARE is the right question, and the one the UK
    # sibling asks. Fixing a share instead forces a choice between the
    # site's own 20% what-if and a 100% ceiling that appears nowhere
    # else, and the answer swings entirely on which is picked. The
    # netting is linear in the share - displaced resistive scales with
    # it - so share = headroom / added-at-100%, exactly.
    routes = ("air_source", "ground_source", "geothermal_network")
    best = {r: None for r in routes}
    rows = []
    for k in hours:
        q = heat[k]
        displaced = q * res_share            # already drawn as MW today
        full = {"air_source": q / cop_at(temps.get(k, 5.0)) - displaced,
                "ground_source": q / GSHP_SPF - displaced,
                "geothermal_network": q / GEO_NETWORK_SCOP - displaced}
        head = ceiling_at(k) - demand[k]
        for r in routes:
            if full[r] <= 0:
                continue
            share = max(0.0, head / full[r])
            if best[r] is None or share < best[r][0]:
                best[r] = (share, k)
        rows.append((demand[k] + full["air_source"], k, q,
                     full["air_source"] + displaced,
                     full["ground_source"] + displaced,
                     full["geothermal_network"] + displaced))
    worst = max(rows)
    total_mw, k, q, ashp, gshp, net = worst
    added = {"air_source": round(ashp - q * res_share),
             "ground_source": round(gshp - q * res_share),
             "geothermal_network": round(net - q * res_share)}
    # Headroom PER ROUTE. The first live run reported it for the air
    # route alone, which meant the actual result - that only the
    # network route fits under the block - had to be worked out by
    # hand from three other numbers. The route ordering IS the
    # finding; it should not need subtracting.
    totals = {r: round(demand[k]) + v for r, v in added.items()}
    ceil_k = round(ceiling_at(k))
    headroom = {r: ceil_k - t for r, t in totals.items()}
    fits_share = {r: (round(100 * best[r][0], 1) if best[r] else None)
                  for r in routes}
    fits_hour = {r: (best[r][1] if best[r] else None) for r in routes}
    fits = [r for r, h in headroom.items() if h >= 0]
    out = {
        "hour": k, "observed_mw": round(demand[k]),
        "total_mw": totals, "headroom_by_route_mw": headroom,
        "routes_that_fit": fits,
        # The heat system against the power system, same hour, same
        # units. This is the quotable one and it was computed all
        # along without ever being printed.
        "heat_vs_block_ratio": round(q / GRID_BLOCK_MW, 2),
        "air_c": temps.get(k), "useful_heat_mw": round(q),
        "added_mw": added,
        "hour_cop_air": round(cop_at(temps.get(k, 5.0)), 2),
        "spf": {"air_seasonal": ASHP_SPF, "ground": GSHP_SPF,
                "network": GEO_NETWORK_SCOP},
        "total_with_air_source_mw": round(total_mw),
        "block_mw": GRID_BLOCK_MW,
        "ceiling_mw": ceil_k,
        "wind_solar_mw": round(ceil_k - GRID_BLOCK_MW),
        "share_that_fits_pct": fits_share,
        "share_binding_hour": fits_hour,
        "headroom_mw": headroom["air_source"],
        "observed_peak_mw": OBSERVED_PEAK_MW,
        "hours_considered": len(hours),
        "basis": ("Hourly useful heat from temp_ai degree hours (hot "
                  "water flat, space heat degree-shaped), through a "
                  "Carnot-fraction COP at each hour's own air "
                  "temperature, NETTED of the resistive heating "
                  "already in observed demand, added to observed "
                  "all-island demand. Block is the de-rated "
                  "dispatchable capacity (dagger); no panel drawn."),
    }
    log(f"B.2.1 how far can heat be electrified inside today's fleet? "
        f"(ceiling = {GRID_BLOCK_MW} MW de-rated block + the wind and "
        f"solar actually generated)")
    for r in routes:
        pct, hr = fits_share[r], fits_hour[r]
        log(f"B.2.1   {r:<19} {pct:>6}% of island heat fits "
            f"[binds {hr}]"
            + ("  - ALL OF IT, with room over" if pct and pct >= 100 else ""))
    log(f"B.2.1 tightest hour for added load: {k} at {out['air_c']} C - "
        f"observed {out['observed_mw']} MW, ceiling {ceil_k} MW "
        f"({out['wind_solar_mw']} MW of it wind and solar)")
    for r in ("air_source", "ground_source", "geothermal_network"):
        h = headroom[r]
        log(f"B.2.1   at 100%: {r:<19} +{added[r]:>5} MW -> "
            f"{totals[r]:>6} MW, headroom {h:+6} MW  "
            f"[{'FITS' if h >= 0 else 'EXCEEDS THE CEILING'}]")
    log(f"B.2.1   routes that fit: "
        f"{', '.join(fits) if fits else 'NONE'}")
    log(f"B.2.1   useful heat in that hour {out['useful_heat_mw']} MWth "
        f"against an {GRID_BLOCK_MW} MW electrical block - the island's "
        f"heat system is {out['heat_vs_block_ratio']}x its power system")
    log(f"B.2.1   air COP in that hour {out['hour_cop_air']} against a "
        f"seasonal {ASHP_SPF} - the gap IS the argument; ground {GSHP_SPF}, "
        f"network {GEO_NETWORK_SCOP} are flat by construction")
    log(f"B.2.1   observed peak on record {OBSERVED_PEAK_MW} MW "
        f"({OBSERVED_PEAK_AT}) - this hour is "
        f"{round(100 * out['observed_mw'] / OBSERVED_PEAK_MW)}% of it, "
        f"which is the sanity rail: a binding hour far below that "
        f"would mean the shaping picked a mild hour")
    if min(headroom.values()) < 0:
        log("B.2.1   NOTE headroom is negative - a full electrification "
            "of heat exceeds the ceiling in this hour. That is a "
            "scenario result, not a forecast: nothing here phases the "
            "conversion or counts storage, diversity or demand response.")
    return out


def derive_geo_percap(anchors=None, geo=None):
    """
    Ground-source Wth per person - installed today vs the capacity the
    hero's 20% what-if implies. Requirement = 20% of annual delivered
    (useful) buildings heat / equivalent full-load hours, per person.
    Pure, unit tested; all sizing parameters dagger.
    """
    a = anchors or ANCHORS
    g = geo or GEO
    eflh = g["eflh_h"]
    pop = g["population_m"]

    def useful_twh(jur):
        j = a[jur]
        heat = j["residential_heat_twh"] + j["services_heat_twh"]
        return heat * sum(sh * a["efficiency"][f]
                          for f, sh in j["fuel_shares"].items())

    def block(jur, current_mwth):
        u = useful_twh(jur)
        need_w = 0.20 * u * 1e12 / eflh          # W of capacity
        p = pop[jur] * 1e6
        return {"current_w_pp": round(current_mwth * 1e6 / p, 1),
                "whatif_w_pp": round(need_w / p, 0),
                "current_mwth": round(current_mwth, 1),
                "whatif_mwth": round(need_w / 1e6, 0),
                "useful_twh": round(u, 1)}

    roi = block("roi", g["roi"]["capacity_mwth"])
    ni = block("ni", g["ni_capacity_mwth_est"])
    pop_i = (pop["roi"] + pop["ni"]) * 1e6
    cur_i = (g["roi"]["capacity_mwth"] + g["ni_capacity_mwth_est"]) * 1e6
    need_i = 0.20 * (roi["useful_twh"] + ni["useful_twh"]) * 1e12 / eflh
    island = {"current_w_pp": round(cur_i / pop_i, 1),
              "whatif_w_pp": round(need_i / pop_i, 0),
              "current_mwth": round(cur_i / 1e6, 1),
              "whatif_mwth": round(need_i / 1e6, 0),
              "useful_twh": round(roi["useful_twh"] + ni["useful_twh"], 1)}
    return {"roi": roi, "ni": ni, "island": island, "eflh_h": eflh,
            "basis": ("20% of delivered buildings heat at "
                      f"{eflh} equivalent full-load hours - all sizing "
                      "parameters dagger; current NI capacity dagger. "
                      "Comparator bars: EGC 2025 country updates, data "
                      "year 2024. Challenge and input welcome at "
                      "contact@causewaygt.com")}


def derive_cool(feeds, anchors=None):
    """
    The cold economy - island cooling loads against the shape of heat
    demand. Flat loads (data centres, refrigeration, process, NI) run
    all year; comfort cooling is shaped by live overheating degree-hours
    (ODH26) when at least a season of the series exists, else treated
    flat with a note. Heat rejection applies per-load factors: vapour-
    compression loads reject compressor work plus the heat they pump.
    With annual totals normalised, the stranded share is the part of
    supply produced while heat demand runs below it. Pure, unit tested.
    """
    a = (anchors or ANCHORS)
    c = a["cool"]
    hddf = feeds.get("hdd") or {}
    hdd = hddf.get("hdd_island") or {}
    days = sorted(hdd)[-365:]
    if len(days) < 200:
        return None
    hs = [hdd[d] for d in days]
    H = sum(hs)
    if H <= 0:
        return None
    n = len(hs)

    loads = dict(c["loads_twh"])
    # DC line = COOLING electricity only (total x cooling share), not
    # the whole data-centre draw - see the correction note in ANCHORS
    dc_total = c["roi_elec_twh"] * c["dc_share_of_roi_elec"]
    loads["dc"] = round(dc_total * c["dc_cooling_share"], 2)
    rf = c["rejection_factor"]
    elec_total = round(sum(loads.values()), 1)
    reject_total = round(sum(loads[k] * rf[k] for k in loads), 1)

    # supply shape: flat loads spread 1/n; comfort follows ODH26
    comfort = loads["comfort"]
    flat = elec_total - comfort
    odh = hddf.get("odh26_island") or {}
    odh_days = [d for d in days if d in odh]
    odh_used = False
    supply = [flat / elec_total / n] * n
    if len(odh_days) >= 60 and sum(odh.get(d, 0.0) for d in days) > 0:
        O = sum(odh.get(d, 0.0) for d in days)
        base = 0.3   # dagger: ventilation floor of the comfort load
        for i, d in enumerate(days):
            shaped = (base / n + (1 - base) * odh.get(d, 0.0) / O)
            supply[i] += comfort / elec_total * shaped
        odh_used = True
    else:
        supply = [s + comfort / elec_total / n for s in supply]

    demand = [h / H for h in hs]
    stranded = sum(max(0.0, s - d) for s, d in zip(supply, demand))

    res_twh = (a["roi"]["residential_heat_twh"]
               + a["ni"]["residential_heat_twh"])
    r1 = lambda x: round(x, 1)
    return {
        "loads_twh": loads,
        "cooling_elec_twh": elec_total,
        "heat_rejected_twh": reject_total,
        "dc_twh": loads["dc"],
        "dc_share_pct": r1(100 * c["dc_share_of_roi_elec"]),
        "dc_share_2028_pct": r1(100 * c["dc_share_2028"]),
        "reject_vs_island_residential_pct": r1(
            100 * reject_total / res_twh),
        "stranded_summer_pct": r1(100 * stranded),
        "comfort_shaped_by_odh": odh_used,
        "dh_share_pct": r1(100 * c["dh_share_of_national_heat"]),
        "basis": ("DC electricity share CSO; other loads and rejection "
                  "factors dagger (refrigeration and comfort reject "
                  "compressor work plus pumped heat). Stranded share "
                  "computed from this site's own degree-day record"
                  + (", comfort load shaped by live overheating "
                     "degree-hours" if odh_used else
                     "; comfort treated flat pending a season of "
                     "overheating degree-hours") +
                  ". Challenge and input welcome at "
                  "contact@causewaygt.com"),
    }


# ------------------------------------------------- DHW vs space mode
# Every route performs WORSE on hot water than on space heating, and
# the site's own anchors say a July week is almost all hot water: hot
# water is a flat 22.4% of annual input while space heat is 77.6%
# HDD-shaped, so the space term collapses in summer and DHW is nearly
# all that is left. Pricing the summer half at an annual efficiency
# would flatter the heat pumps and flatter the boiler at exactly the
# point where the lines converge - which is where the argument is won
# or lost. So each route carries a PAIR.
#
#   oil boiler   space 0.82 / DHW 0.71. SAP 2012 Table 4b, p.207,
#                "Seasonal efficiency for gas and oil boilers", gives a
#                WINTER and a SUMMER figure for every archetype. Across
#                the modern oil stock - standard 1998+ 80/68,
#                condensing 84/72, condensing combi 82/73 - summer runs
#                85 to 89% of winter, mean 0.866. On this site's 0.82
#                winter anchor that is 0.71.
#   gas boiler   space 0.85 / DHW 0.75. Same table: regular condensing
#                84/74, condensing combi 84/75, regular non-condensing
#                74/64; mean ratio 0.879 on 0.85 gives 0.75.
#
# CORRECTED 14 Aug 2026, and the correction is large. These were 0.55
# and 0.68, resting on BRE STP09/B01's remark about "very low hot water
# efficiency (under 40% gross)" - a line describing TWO SPECIFIC oil
# boilers in a methodology paper, which I treated as a fleet floor and
# cited to Simon as one. Table 4b IS the fleet answer, and its worst
# archetype - a single-burner range cooker boiler at 47/37 - is a ratio
# of 0.79, still far above the 0.67 I was using. The summer oil penalty
# is real and roughly half what the panel was showing.
#
# Two things deliberately NOT applied. Table 4b is SAP's fallback for
# boilers absent from the Product Characteristics Database, so it is
# conservative and skewed to older plant; a BER-weighted PCDB figure
# would beat it, and BER records boiler make and model. And Table 4c
# deducts 5 points from BOTH figures where a regular boiler has no
# interlock, which is common in older Irish installations (dagger).
#   ashp         space from the Carnot engine at that day's air
#                temperature; DHW 1.70 - the MCS 031 Issue 4.0 default
#                (18 Mar 2025), cut from 1.75 in line with SAP 10.2.
#   gshp         space 3.24 / DHW 2.24 - Energy Systems Catapult
#                in-situ for space, MCS HPSPE default for DHW.
#   network      space 5.00 / DHW derived at the SAME ratio as GSHP
#                (2.24/3.24 = 0.691), because both are ground-coupled
#                with a constant source: the DHW penalty is the flow
#                temperature lift alone, not a source-side collapse.
#
# NOT SPLIT: EoH did not meter DHW separately, so there is no field
# DHW SPF - the heat-pump DHW figures are MCS design defaults. Said
# plainly rather than dressed as measurement.
# ------------------------------------------------- statutory wedges
# What has to be stripped to get from a retail price to an EX-TAX one
# - product plus network plus margin. Deliberately NOT called
# "wholesale": the EU bulletin's ex-tax oil line is product,
# distribution and margin, not a wholesale quote, and the same is true
# of a gas or electricity unit rate with its tax removed. Calling it
# wholesale would overclaim.
#
# ORDER OF OPERATIONS. VAT is the OUTERMOST layer in both
# jurisdictions - it is charged on the carbon-tax-inclusive price. So
# ex-VAT = retail / (1 + vat), THEN subtract the carbon component.
# Reversing that would understate the wedge.
#
# ROI, from Revenue: kerosene carries no non-carbon excise, a carbon
# component of EUR 160.81 per 1,000 L (16.081 c/L at EUR 63.50/t CO2),
# the NORA levy at 2 c/L, and VAT at 13.5%. Natural gas carries NGCT
# at EUR 11.48/MWh GCV = 1.148 c/kWh and VAT at 9%. Electricity
# carries VAT at 9% and a per-customer PSO levy that is NOT per kWh,
# so it is out of scope for a unit-rate strip and noted rather than
# subtracted.
#
# NI: kerosene is fully duty-rebated to nil with no carbon price; gas
# and electricity carry no carbon price and no Climate Change Levy
# (domestic use is excluded). VAT is 5% on all three. So ex-tax is
# simply retail / 1.05 - exact, not estimated.
#
# THE STEP THAT IS COMING. Ireland's carbon increase on non-propellant
# fuels normally lands on 1 May; for 2026 it was postponed to 14
# October, when the charge moves from EUR 63.50 to EUR 71.00 a tonne.
# That is a discontinuity in the middle of the record, so it is a
# dated table rather than a constant.
# The carbon component scales linearly with the charge per tonne, so
# the whole table derives from one anchored pair: at EUR 63.50/t
# Revenue publishes 16.081 c/L on kerosene and EUR 11.48/MWh GCV =
# 1.148 c/kWh on gas. Checked against the published prior rate - at
# EUR 56.00/t the gas figure comes out 1.012 against Revenue's 1.013,
# which is rounding, not a different method.
#
# WHY IT REACHES BACK TO 2020. Removing the back-look floor took the
# price panel to 122 weeks and 2024-04-15 on its first run, and the
# HDD record keeps growing. A table starting in 2025 would have
# clamped 55 of those weeks to a rate that did not exist yet - the
# 2024-25 charge was EUR 56.00, not EUR 63.50 - and done it silently.
# So the table covers the whole Finance Act 2020 trajectory, and
# anything earlier than its first row is REFUSED rather than clamped.
_CARBON_PER_TONNE = [
    ("2020-05-01", 26.00), ("2021-05-01", 33.50), ("2022-05-01", 41.00),
    ("2023-05-01", 48.50), ("2024-05-01", 56.00), ("2025-05-01", 63.50),
    # Normally 1 May; for 2026 the non-propellant step was postponed
    # to 14 October, so this row is the exception that proves the rule.
    ("2026-10-14", 71.00), ("2027-05-01", 78.50), ("2028-05-01", 86.00),
    ("2029-05-01", 93.50), ("2030-05-01", 100.00),
]
CARBON_STEPS = [(d, r, round(16.081 * r / 63.50, 3),
                 round(1.148 * r / 63.50, 4)) for d, r in _CARBON_PER_TONNE]

# ROI gas and electricity were cut from 13.5% to 9% on 1 May 2022 as a
# cost-of-living measure, extended to 31 Dec 2030. Kerosene never got
# the cut and stays at 13.5% - which is why the tax wedge on oil is
# heavier than on gas inside the same jurisdiction. NI is a flat 5%
# throughout.
VAT_STEPS = {
    "roi": {"oil": [("2000-01-01", 0.135)],
            "gas": [("2000-01-01", 0.135), ("2022-05-01", 0.09)],
            "electricity": [("2000-01-01", 0.135), ("2022-05-01", 0.09)]},
    "ni": {f: [("2000-01-01", 0.05)]
           for f in ("oil", "gas", "electricity")},
}


def vat_for(jur, fuel, date_iso):
    """The VAT rate in force on a date. A constant would have been
    wrong the moment the panel reached back past 1 May 2022."""
    rate = VAT_STEPS[jur][fuel][0][1]
    for frm, r in VAT_STEPS[jur][fuel]:
        if date_iso >= frm:
            rate = r
    return rate


VAT = {"roi": {"oil": 0.135, "gas": 0.09, "electricity": 0.09},
       "ni": {"oil": 0.05, "gas": 0.05, "electricity": 0.05}}
NORA_LEVY_C_PER_L = 2.0


def carbon_for(date_iso):
    """
    The carbon component in force on a date, or None before the table
    starts.

    None rather than a clamp, deliberately. Clamping is how 55 weeks
    got priced at a rate that had not been legislated yet, and it left
    no trace - the row simply looked like the others. A refusal
    propagates to a missing ex-tax figure, which the panel can grey
    out and a reader can see.
    """
    if date_iso < CARBON_STEPS[0][0]:
        return None
    row = CARBON_STEPS[0]
    for r in CARBON_STEPS:
        if date_iso >= r[0]:
            row = r
    return {"eur_per_tonne": row[1], "kerosene_c_per_l": row[2],
            "gas_c_per_kwh": row[3]}


def ex_tax(retail, jur, fuel, date_iso, kwh_per_litre=None):
    """
    Retail -> ex-tax, in the same units in. VAT first, then carbon.

    For oil, `retail` is per litre and the carbon and NORA components
    are per litre. For gas and electricity, per kWh. Returns None if
    the strip would go negative, which would mean a wrong rate rather
    than a cheap fuel.
    """
    net = retail / (1 + vat_for(jur, fuel, date_iso))
    if jur == "roi" and fuel in ("oil", "gas"):
        c = carbon_for(date_iso)
        if c is None:
            return None          # before the table: refuse, don't guess
        net -= (c["kerosene_c_per_l"] + NORA_LEVY_C_PER_L
                if fuel == "oil" else c["gas_c_per_kwh"])
    return round(net, 4) if net > 0 else None


DHW_MODE = {
    "oil_boiler": 0.71,        # SAP Table 4b ratio on the 0.82 anchor
    "gas_boiler": 0.75,        # SAP Table 4b ratio on the 0.85 anchor
    "ashp": 1.70,
    "gshp": 2.24,
    "network": round(5.00 * 2.24 / 3.24, 2),
}
# Worst archetype in Table 4b: single-burner range cooker boiler with
# a permanent pilot, 47 winter / 37 summer. Nothing in the Irish stock
# should price below it.
DHW_MODE_FLOOR_OIL = 0.37

# ZEROED, 14 Aug 2026. The fraction of oil-home hot water made by an
# electric immersion rather than the boiler. The mechanism is real -
# Irish oil households do switch in summer rather than fire a boiler
# at sub-40% for a cylinder, and at COP 1 that is DEARER than the
# inefficient boiler, so it pushed the oil line up rather than down.
#
# But at 0.30 it was adding 5.8 c to the summer oil figure, roughly
# 46% of which was electricity rather than oil, on no metered evidence
# at all - the research found consumer guidance and owner forums and
# nothing measured. It was the largest unevidenced term on the panel
# and it sat under its most striking feature. The UK sibling has no
# equivalent, so carrying it here also broke parity between the two.
#
# Left as a named constant rather than deleted: the code path stays,
# and a metered Irish or NI study of summer immersion use is what
# would turn it back on.
OIL_IMMERSION_DHW_SHARE = 0.0    # dagger - zeroed pending evidence


def sector_blend(jur, fuel, date_iso, anchors=None, nondom=None):
    """
    The price a route actually pays, blended across sectors.

    Every route on this panel was priced at pure DOMESTIC tariffs,
    which is wrong twice over. About a quarter of the island's
    building heat is services rather than residential, and the hero
    bill has always blended that - so the cost panel disagreed with
    the bill on the same page. And a heat NETWORK operator is not a
    household at all: it buys electricity on a commercial contract and
    would never pay a domestic tariff, which is the UK sibling's own
    correction.

    So the blend is not uniform:
      oil            one price - no non-domestic oil rate exists, and
                     kerosene is sold to both sectors on the same terms
      gas, ashp,     blended at the services share of island heat
      gshp           input (~26%), matching the hero bill
      network        100% NON-DOMESTIC, whoever the end customer is

    Non-domestic rates already EXCLUDE VAT by convention (businesses
    recover it), so the blend mixes a VAT-inclusive domestic rate with
    a VAT-exclusive commercial one - which is the point, not an error:
    it is what the heat actually costs to buy.
    """
    a = anchors or ANCHORS
    dom = tariffs_for(date_iso)
    if dom is None:
        return None
    cur = "gbp" if jur == "ni" else "eur"
    d = dom[cur][fuel] * 100
    nd = (nondom or {}).get(cur, {}).get(fuel)
    nd = nd * 100 if nd else None
    if nd is None:
        return {"domestic": d, "nondom": None, "blend": d, "network": d}
    r = sum(a[j]["residential_heat_twh"] for j in ("ni", "roi"))
    sv = sum(a[j]["services_heat_twh"] for j in ("ni", "roi"))
    w = sv / (r + sv) if (r + sv) else 0.0
    return {"domestic": d, "nondom": nd,
            "blend": round((1 - w) * d + w * nd, 4),
            "network": nd, "services_share": round(w, 4)}


# ---------------------------------------------------- the COP engine
# Ported from the UK sibling so the two dashboards compute the electric
# routes the same way. The Irish panel had ONE air-source SPF for the
# whole record, which made every electric line a flat multiple of the
# electricity price - they could never spread apart in cold weather,
# which is the entire point of the chart.
#
# Method, in the UK's order: assert the SPF anchor, then CALIBRATE the
# Carnot fraction so the trailing-year heat-weighted SPF reproduces it.
# That is the right way round. Assuming a fraction and computing an SPF
# (which is what I did by hand for the ROI figure) lets the two drift
# apart; calibrating means the headline number is the anchor by
# construction and the weather only redistributes it.
FLOW_MILD_C, FLOW_COLD_C = 30.0, 50.0      # weather compensation ends
FLOW_MILD_AT, FLOW_COLD_AT = 15.0, -5.0
DHW_FLOW_C = 52.0        # dagger - cylinder flow, year-round
MIN_LIFT_K = 8.0         # dagger - floor on Carnot lift
AIR_APPROACH_C = 3.0     # evaporator approach below air temperature
GROUND_SOURCE_C = 8.0    # 11 C ground less 3 C brine approach
DEFROST_MAX = 0.12       # peak fractional COP loss, centred ~2 C

# THE JURISDICTIONAL SPLIT. Permo-Triassic HSA is an NI play and has no
# onshore ROI equivalent at scale, so the two network routes are
# different machines and cannot share a number.
#   NI  - the UK sibling's blended model: UTES and intermediate-doublet
#         together, source 19.6 C, SPF 5.0.
#   ROI - a 5G ambient loop on seasonal storage alone, charged from
#         comfort cooling and process rejection. Source 16 C is the
#         SEASONAL MEAN of a store charged to 25-30 C in September and
#         depleting through the heating season; measured Dutch ATES
#         recovery of 68-87% is what makes the mean that rather than
#         the charge temperature. SPF 4.0, not the 4.2 I first derived
#         - a distributed ambient loop carries more circulation
#         pumping than the borehole array the fraction came from.
NETWORK_MODEL = {
    "ni": {"source_c": 19.6, "spf": 5.0,
           "note": "UTES and intermediate-doublet blend (Permo-Triassic "
                   "HSA), as the UK sibling"},
    "roi": {"source_c": 16.0, "spf": 4.0,
            "note": "5G ambient loop on seasonal storage, charged from "
                    "comfort cooling and process heat rejection"},
}
SPF_ANCHORS = {"ashp": 2.80, "gshp": 3.24}


def flow_temp(t_out):
    """Weather-compensated SPACE flow. Hot water does not follow it."""
    f = min(1.0, max(0.0, (FLOW_MILD_AT - t_out)
                     / (FLOW_MILD_AT - FLOW_COLD_AT)))
    return FLOW_MILD_C + (FLOW_COLD_C - FLOW_MILD_C) * f


def defrost_factor(t_out):
    """Fraction of COP retained. Bell-shaped loss centred on 2 C, the
    humid frost band. Shape is a dagger; the level is re-anchored by
    the eta calibration afterwards, so only the shape matters here."""
    return 1.0 - DEFROST_MAX * math.exp(-((t_out - 2.0) / 3.0) ** 2)


def carnot_cop(t_flow, t_source):
    return (t_flow + 273.15) / max(MIN_LIFT_K, t_flow - t_source)


def route_cop(route, t_out, eta, jur="roi", dhw=False):
    """Point COP for one route on one day. `dhw` prices the hot-water
    leg at the cylinder flow rather than the compensated space flow -
    without that split, a mild summer day gives absurd COPs on a load
    that is entirely hot water."""
    tf = DHW_FLOW_C if dhw else flow_temp(t_out)
    if route == "ashp":
        return max(1.0, eta * carnot_cop(tf, t_out - AIR_APPROACH_C)
                   * defrost_factor(t_out))
    if route == "gshp":
        return max(1.0, eta * carnot_cop(tf, GROUND_SOURCE_C))
    if route == "network":
        return max(1.0, eta * carnot_cop(tf, NETWORK_MODEL[jur]["source_c"]))
    raise ValueError(route)


def calibrate_eta(route, days, jur="roi", anchors=None):
    """
    Solve the Carnot fraction so the HEAT-WEIGHTED SPF over the days
    given reproduces the route's anchor.

    Heat-weighted, not day-averaged: almost all the heat is drawn in
    the cold, so a mean over days would flatter every route. `days` is
    [(t_out, space_kwh, dhw_kwh), ...].
    """
    a = anchors or ANCHORS
    target = (NETWORK_MODEL[jur]["spf"] if route == "network"
              else SPF_ANCHORS[route])

    def spf(eta):
        heat = elec = 0.0
        for t, sp, dhw in days:
            if sp > 0:
                elec += sp / route_cop(route, t, eta, jur, dhw=False)
            if dhw > 0:
                elec += dhw / route_cop(route, t, eta, jur, dhw=True)
            heat += sp + dhw
        return (heat / elec) if elec > 0 else 0.0

    lo, hi = 0.05, 1.20
    for _ in range(60):
        mid = (lo + hi) / 2
        if spf(mid) < target:
            lo = mid
        else:
            hi = mid
    return round((lo + hi) / 2, 4)


def route_cost_useful(prices, ashp_spf, dhw_share, anchors=None):
    """
    Cost per useful kWh by route, in native minor units, with hot
    water and space heat priced in their own modes and blended at the
    week's own DHW share.

    prices: {"oil_per_kwh", "gas_per_kwh", "elec_per_kwh"} in minor
    units per kWh of INPUT. dhw_share: 0-1, the fraction of delivered
    heat that is hot water this week. Pure, unit tested.
    """
    a = anchors or ANCHORS
    w = min(1.0, max(0.0, dhw_share))
    oil, gas, elec = (prices["oil_per_kwh"], prices["gas_per_kwh"],
                      prices["elec_per_kwh"])
    # The network route buys commercially; the heat-pump routes are a
    # domestic/services blend. Falls back to the blended price when no
    # non-domestic rate is supplied, so callers need not change.
    elec_net = prices.get("elec_network_per_kwh", elec)

    cops = prices.get("cops")           # {route: (space_cop, dhw_cop)}

    def blend(space_perf, dhw_perf, price):
        return (1 - w) * price / space_perf + w * price / dhw_perf

    def blend_route(route, price, sp_fb, dhw_fb):
        """Per-day COPs when supplied, the flat SPF otherwise. The
        fallback is passed as its two performances rather than as a
        computed number, so it is not evaluated when the COPs are
        present - ashp_spf is legitimately None on the daily path."""
        c = (cops or {}).get(route)
        return blend(*(c if c else (sp_fb, dhw_fb)), price)

    # Oil hot water is part boiler, part immersion - and the immersion
    # runs on electricity at COP 1, so the leakage RAISES the cost
    # rather than lowering it.
    oil_dhw = ((1 - OIL_IMMERSION_DHW_SHARE) * oil / DHW_MODE["oil_boiler"]
               + OIL_IMMERSION_DHW_SHARE * elec / 1.0)
    return {
        "oil_boiler": round((1 - w) * oil / a["efficiency"]["oil"]
                            + w * oil_dhw, 2),
        "gas_boiler": round(blend(a["efficiency"]["gas"],
                                  DHW_MODE["gas_boiler"], gas), 2),
        "ashp": round(blend_route("ashp", elec, ashp_spf,
                                  DHW_MODE["ashp"]), 2),
        "gshp": round(blend_route("gshp", elec, GSHP_SPF,
                                  DHW_MODE["gshp"]), 2),
        "network": round(blend_route("network", elec_net, GEO_NETWORK_SCOP,
                                     DHW_MODE["network"]), 2),
        "dhw_share": round(w, 3),
    }


def week_dhw_share(hdd_daily, w_end, anchors=None):
    """
    What fraction of a week's delivered heat is hot water.

    The site's own shaping, one step further: hot water is flat across
    the year, space heat follows the week's share of the trailing
    year's heating degree days. So the DHW share is high in July and
    low in January, and it is computed rather than assumed.
    """
    a = anchors or ANCHORS
    days = sorted(d for d in hdd_daily if d <= w_end)
    if len(days) < 300:
        return None
    wk = [d for d in days if d > (dt.date.fromisoformat(w_end)
                                  - dt.timedelta(days=7)).isoformat()]
    year = days[-365:]
    hdd_wk = sum(hdd_daily[d] for d in wk)
    hdd_yr = sum(hdd_daily[d] for d in year)
    if hdd_yr <= 0 or len(wk) < 5:
        return None
    dhw = (1 - a["space_heat_fraction"]) / 52.0
    space = a["space_heat_fraction"] * hdd_wk / hdd_yr
    tot = dhw + space
    return (dhw / tot) if tot > 0 else None


def day_dhw_share(hdd_daily, day, anchors=None):
    """Fraction of a DAY's delivered heat that is hot water. Same
    shaping as the weekly version, one step finer."""
    a = anchors or ANCHORS
    days = sorted(d for d in hdd_daily if d <= day)
    # Needs a trailing year to know what share of the year's degree
    # days this one carries. Below ~200 the denominator is a season
    # rather than a year and the share is meaningless.
    if len(days) < 200:
        return None
    year = days[-365:]
    hdd_yr = sum(hdd_daily[d] for d in year)
    if hdd_yr <= 0:
        return None
    dhw = (1 - a["space_heat_fraction"]) / 365.0
    space = a["space_heat_fraction"] * hdd_daily.get(day, 0.0) / hdd_yr
    tot = dhw + space
    return (dhw / tot) if tot > 0 else None


def derive_heat_cost_series(feeds, anchors=None):
    """
    DAILY cost of a useful kWh by route - the Irish equivalent of the
    UK sibling's panel, with oil as the main series.

    Daily, not weekly, because the electric routes are priced at each
    day's own COP and the whole argument is what happens on the cold
    days rather than the average of the days. The oil price is the
    only weekly input (the EU bulletin publishes weekly), so it is
    STEP-HELD across the week rather than interpolated - a price that
    did not move should not appear to.
    """
    a = anchors or ANCHORS
    hddf = feeds.get("hdd") or {}
    hdd_i = hddf.get("hdd_island") or {}
    temp = {"roi": hddf.get("temp_roi") or {}, "ni": hddf.get("temp_ni") or {}}
    if len(hdd_i) < 300 or not temp["roi"]:
        return None
    bull = ((feeds.get("oil_bulletin") or {})
            .get("roi_heating_gasoil_eur_per_1000l") or {})
    bull_nt = ((feeds.get("oil_bulletin") or {})
               .get("roi_heating_gasoil_eur_per_1000l_ex_tax") or {})
    ccni = (((feeds.get("ccni_oil") or {}).get("series_gbp") or {})
            .get("daily", {}).get("900l") or {})
    if not bull:
        return None
    kwh_l = a["kerosene_kwh_per_litre"]

    # Calibrate once, over the trailing year, per jurisdiction.
    etas = {}
    for jur in ("roi", "ni"):
        t = temp[jur]
        cal = []
        for d in sorted(t)[-365:]:
            sh = day_dhw_share(hdd_i, d, a)
            if sh is None:
                continue
            cal.append((t[d], 1 - sh, sh))
        # A full year is what you want - the fraction is calibrated
        # against the whole seasonal swing - but a short record should
        # degrade rather than return nothing, and say so.
        if len(cal) < 150:
            log(f"heat cost: cannot calibrate {jur} - only {len(cal)} "
                "days carry a hot-water share; needs 150")
            return None
        if len(cal) < 330:
            log(f"heat cost: WARNING calibrating {jur} on {len(cal)} days, "
                "not a full year - the fraction is biased toward whichever "
                "season the record covers")
        etas[jur] = {r: calibrate_eta(r, cal, jur, a)
                     for r in ("ashp", "gshp", "network")}
    log("heat cost: calibrated Carnot fractions "
        + "; ".join(f"{j} " + ", ".join(f"{r} {v}" for r, v in e.items())
                    for j, e in etas.items())
        + f" (anchors ashp {SPF_ANCHORS['ashp']}, gshp "
          f"{SPF_ANCHORS['gshp']}, network ni "
          f"{NETWORK_MODEL['ni']['spf']} / roi "
          f"{NETWORK_MODEL['roi']['spf']})")
    spread = max(v for e in etas.values() for v in e.values()) / \
        min(v for e in etas.values() for v in e.values())
    if spread > 1.15:
        log(f"heat cost: WARNING calibrated fractions spread {spread:.2f}x "
            "- more than 15% apart means a source temperature and an SPF "
            "anchor that do not describe the same machine")

    def cops_for(jur, t_out):
        e = etas[jur]
        return {r: (route_cop(r, t_out, e[r], jur, dhw=False),
                    route_cop(r, t_out, e[r], jur, dhw=True))
                for r in ("ashp", "gshp", "network")}

    # THE MODE REGIME IS SEASONAL, NOT DAILY. A boiler in January does
    # not drop to summer cycling efficiency because one day was mild -
    # it is still running its space-heating circuit. The sub-40% hot
    # water efficiency arises from a REGIME: the boiler off for space
    # heat, cycling for a small cylinder load. So the mode blend runs
    # on a trailing 28-day share while the heat-pump COPs stay on the
    # day's own temperature, which is instantaneous physics.
    #
    # Before this the oil line sawtoothed 15 to 28 c between adjacent
    # days, which is not a price signal, it is a shaping artefact.
    MODE_SMOOTH_DAYS = 28
    shares = {}
    for d in sorted(temp["roi"]):
        sh = day_dhw_share(hdd_i, d, a)
        if sh is not None:
            shares[d] = sh
    sm_days = sorted(shares)
    smooth = {}
    for i, d in enumerate(sm_days):
        w = [shares[x] for x in sm_days[max(0, i - MODE_SMOOTH_DAYS + 1):i + 1]]
        smooth[d] = sum(w) / len(w)

    weeks = sorted(bull)
    out, unpriced = [], 0
    for day in sorted(temp["roi"]):
        if day < weeks[0]:
            continue
        share = day_dhw_share(hdd_i, day, a)
        if share is None:
            continue
        t = tariffs_for(day)
        if t is None:
            unpriced += 1
            continue
        # step-held: the most recent bulletin week at or before today
        wk = max((w for w in weeks if w <= day), default=None)
        if wk is None:
            continue
        nd, _ = nondom_for(day, ((feeds.get("ecb_fx") or {})
                                 .get("eur_gbp_semester")))
        # daily share for the caption, smoothed share for the money
        mode = smooth.get(day, share)
        row = {"day": day, "dhw_share": round(share, 3),
               "dhw_mode": round(mode, 3),
               "t_roi": temp["roi"].get(day)}
        oil_c_l = bull[wk] / 10.0
        oil_ex = (bull_nt[wk] / 10.0 if wk in bull_nt
                  else ex_tax(oil_c_l, "roi", "oil", day))
        gb = sector_blend("roi", "gas", day, a, nd)
        eb = sector_blend("roi", "electricity", day, a, nd)
        cr = cops_for("roi", temp["roi"][day])
        row["roi"] = route_cost_useful(
            {"oil_per_kwh": oil_c_l / kwh_l, "gas_per_kwh": gb["blend"],
             "elec_per_kwh": eb["blend"], "elec_network_per_kwh": eb["network"],
             "cops": cr}, None, mode, a)
        gd_x = ex_tax(gb["domestic"], "roi", "gas", day)
        en_x = ex_tax(eb["domestic"], "roi", "electricity", day)
        c = carbon_for(day)
        gn_x = (gb["nondom"] - c["gas_c_per_kwh"]) if c and gb["nondom"] else None
        w_sv = gb.get("services_share", 0.0)
        if oil_ex and gd_x and en_x and gn_x:
            row["roi_ex_tax"] = route_cost_useful(
                {"oil_per_kwh": oil_ex / kwh_l,
                 "gas_per_kwh": (1 - w_sv) * gd_x + w_sv * gn_x,
                 "elec_per_kwh": (1 - w_sv) * en_x + w_sv * eb["nondom"],
                 "elec_network_per_kwh": eb["nondom"], "cops": cr},
                None, mode, a)
        if day in ccni and temp["ni"].get(day) is not None:
            ppl = ccni[day] * 100 / 900
            gbn = sector_blend("ni", "gas", day, a, nd)
            ebn = sector_blend("ni", "electricity", day, a, nd)
            cn = cops_for("ni", temp["ni"][day])
            row["t_ni"] = temp["ni"][day]
            row["ni"] = route_cost_useful(
                {"oil_per_kwh": ppl / kwh_l, "gas_per_kwh": gbn["blend"],
                 "elec_per_kwh": ebn["blend"],
                 "elec_network_per_kwh": ebn["network"], "cops": cn},
                None, mode, a)
            row["ni_ex_tax"] = route_cost_useful(
                {"oil_per_kwh": ex_tax(ppl, "ni", "oil", day) / kwh_l,
                 "gas_per_kwh": ((1 - w_sv) * ex_tax(gbn["domestic"], "ni",
                                                     "gas", day)
                                 + w_sv * gbn["nondom"]),
                 "elec_per_kwh": ((1 - w_sv) * ex_tax(ebn["domestic"], "ni",
                                                      "electricity", day)
                                  + w_sv * ebn["nondom"]),
                 "elec_network_per_kwh": ebn["nondom"], "cops": cn},
                None, mode, a)
        out.append(row)
    if not out:
        return None
    last = out[-1]
    for j in ("roi", "ni"):
        if j in last:
            r = last[j]
            log(f"heat cost ({j} {last['day']}, {last.get('t_' + j)} C, "
                f"DHW {last['dhw_share']}): oil {r['oil_boiler']} / gas "
                f"{r['gas_boiler']} / ashp {r['ashp']} / gshp {r['gshp']} "
                f"/ network {r['network']} per useful kWh")
    if unpriced:
        log(f"heat cost: {unpriced} day(s) NOT priced - before the tariff "
            f"table starts ({TARIFF_HISTORY[0][0]})")
    log(f"heat cost: {len(out)} days priced "
        f"({sum(1 for r in out if 'ni' in r)} with NI), "
        f"{out[0]['day']}..{out[-1]['day']}; oil price step-held weekly")
    return out


def derive_heat_gap(feeds, anchors=None):
    """
    Cost of useful heat by route, per jurisdiction, standard tariffs -
    plus the break-even SPF against the incumbent oil boiler. Pure,
    unit tested. Native currency minor units (p or c) per useful kWh.
    """
    a = anchors or ANCHORS
    fx = (feeds.get("ecb_fx") or {}).get("eur_gbp") or 0.855

    ccni = ((feeds.get("ccni_oil") or {}).get("series_gbp") or {}).get(
        "daily", {}).get("900l") or {}
    oil_ni_ppl = ccni[max(ccni)] * 100 / 900 if ccni else None
    ob = (feeds.get("oil_bulletin") or {}).get("latest_value")
    oil_roi_cpl = ob * 100 / 1000 if ob else None
    kwh_l = a["kerosene_kwh_per_litre"]
    eff_oil, eff_gas = a["efficiency"]["oil"], a["efficiency"]["gas"]
    spf_geo = a["geothermal_spf"]
    hddf = feeds.get("hdd") or {}
    ashp_ni = derive_ashp_spf(hddf.get("hdd_ni") or {}, a)
    ashp_roi = derive_ashp_spf(hddf.get("hdd_roi") or {}, a)
    fallback = {"spf": 2.8}   # field-trial median, dagger
    ashp_ni = ashp_ni or fallback
    ashp_roi = ashp_roi or fallback

    def jur(oil_pl, elec, gas, ashp):
        if oil_pl is None:
            return None
        oil_useful = oil_pl / kwh_l / eff_oil
        r2 = lambda x: round(x, 2)
        return {
            "oil_boiler": r2(oil_useful),
            "gas_boiler": r2(gas * 100 / eff_gas),
            "ashp": r2(elec * 100 / ashp["spf"]),
            "ashp_spf": ashp["spf"],
            "ashp_model": {k: v for k, v in ashp.items() if k != "params"},
            "geothermal_spf40": r2(elec * 100 / spf_geo),
            "breakeven_spf_vs_oil": r2(elec * 100 / oil_useful),
            "breakeven_spf_vs_gas": r2(elec * 100 / (gas * 100 / eff_gas)),
            "inputs": {"oil_per_litre": round(oil_pl, 2),
                       "electricity_per_kwh": elec, "gas_per_kwh": gas},
        }

    ni = jur(oil_ni_ppl, a["retail_gbp_per_kwh"]["electricity"],
             a["retail_gbp_per_kwh"]["gas"], ashp_ni)
    roi = jur(oil_roi_cpl, a["retail_eur_per_kwh"]["electricity"],
              a["retail_eur_per_kwh"]["gas"], ashp_roi)
    if not (ni and roi):
        return None
    return {
        "ni": ni, "roi": roi, "fx_eur_gbp": fx,
        "geo_spf": spf_geo,
        "basis": ("Standard tariffs, July 2026 pass (Power NI/UR review, "
                  "ROI standard 24h rates) - dagger; time-of-use and night "
                  "tariffs materially lower for heat-pump households. Oil "
                  "prices live. ASHP SPF is modelled from each "
                  "jurisdiction's HDD-weighted climate (Carnot-fraction, "
                  "defrost derate, DHW share - all dagger), calibrated to "
                  "GB field-trial medians. Kerosene 10.35 kWh/L; boiler "
                  "efficiencies 82%/85% dagger. Challenge and input "
                  "welcome at contact@causewaygt.com"),
    }


# ---------------------------------------------------------------- assembly

FEEDS = {
    "eirgrid": feed_eirgrid,
    "hdd": feed_hdd,
    "ecb_fx": feed_ecb_fx,
    "gni_ckan": feed_gni_ckan,
    "semopx": feed_semopx,
    "oil_bulletin": feed_oil_bulletin,
    "gni_live": feed_gni_live,
    "ccni_oil": feed_ccni_oil,
    "gb_oil": feed_gb_oil,
    "entsog_probe": feed_entsog_probe,
    "sem_mix": feed_sem_mix,
    "eirgrid_probe": feed_eirgrid_probe,
}



# ------------------------------------------------------- hourly store
# EirGrid serves a ~30-day window of 15-minute rows ending at dateTo,
# for chartType x region, with `areas` matching the chart. Probe round
# 2 (7 Aug 2026) confirmed historic windows return correctly-dated
# data a year back, so the 13-month store backfills by chunked
# walking. dateRange=year returns nothing - month chunks are the
# mechanism. All-island scope, per the v7 grid-layer decision.
HOURLY_SERIES = {
    "demand_ai": ("demand", "ALL", "demandactual"),
    "wind_ai": ("wind", "ALL", "windactual"),
    "solar_ai": ("solar", "ALL", "solaractual"),
    "co2_ai": ("co2", "ALL", "co2intensity"),
}


def _hour_key(ts):
    """15-minute stamp -> UTC hour key. EirGrid stamps are local
    clock; the weekly layer already treats these as day-local, and the
    grid layer needs consistency with it rather than with UTC."""
    for f in ("%d-%b-%Y %H:%M:%S", "%d-%B-%Y %H:%M:%S"):
        try:
            return dt.datetime.strptime(ts, f).strftime("%Y-%m-%dT%H")
        except ValueError:
            continue
    return None


def hourly_from_rows(rows, min_quarters=3):
    """15-minute rows -> hourly MEANS (never samples). An hour needs
    at least min_quarters of its four values or it is dropped, so a
    partial hour cannot masquerade as a low one."""
    buckets = {}
    for r in rows or []:
        v = r.get("Value")
        if v is None:
            continue
        k = _hour_key(str(r.get("EffectiveTime") or ""))
        if k:
            buckets.setdefault(k, []).append(float(v))
    return {k: round(sum(v) / len(v), 2)
            for k, v in buckets.items() if len(v) >= min_quarters}


def fetch_hourly_chunk(chart, region, areas, end_day):
    """One ~30-day window ending end_day."""
    def dmy(d):
        return d.strftime("%d-%b-%Y").replace(" 0", " ")
    payload = http_get(EIRGRID_ENDPOINT, params={
        "region": region, "chartType": chart, "dateRange": "month",
        "dateFrom": dmy(end_day - dt.timedelta(days=27)),
        "dateTo": dmy(end_day), "areas": areas,
    }, timeout=120).json()
    return hourly_from_rows((payload or {}).get("Rows", []))


def weighted_hourly_temp(payload, names, weights):
    """
    Open-Meteo hourly payloads -> population-weighted island air
    temperature per hour, keyed 'YYYY-MM-DDTHH'.

    The divisor is the weight ACTUALLY PRESENT in each hour, not the
    full 1.0. A station missing an hour then leaves the island mean
    unbiased instead of dragging it toward zero, which a plain
    weighted sum would do silently and only in the hours where a feed
    gap already exists. Pure, unit tested.
    """
    locs = payload if isinstance(payload, list) else [payload]
    acc, wt = {}, {}
    for name, loc in zip(names, locs):
        w = weights.get(name, 0.0)
        if w <= 0:
            continue
        hh = (loc or {}).get("hourly", {}) or {}
        for ts, t in zip(hh.get("time", []) or [],
                         hh.get("temperature_2m", []) or []):
            if t is None:
                continue
            k = str(ts)[:13]
            acc[k] = acc.get(k, 0.0) + w * float(t)
            wt[k] = wt.get(k, 0.0) + w
    return {k: round(acc[k] / wt[k], 2) for k in acc if wt[k] > 0}


def fetch_hourly_temp(previous, floor_key, end_day):
    """
    Island hourly temperature across the store's window, merged onto
    whatever previous runs retained.

    TIMEZONE. Requested on Europe/Dublin clock, NOT UTC. _hour_key
    deliberately keys the EirGrid series on its own local stamps, and
    the daily HDD feed asks Open-Meteo for UTC. Joining those two
    would put temperature an hour out of step with demand from late
    March to late October - silently, and in exactly the direction
    that misaligns an evening peak with the temperature that caused
    it. Both sides of this store are therefore local clock.

    The autumn fold repeats one local hour; the later value wins, in
    both series, for one hour a year.
    """
    names = list(STATIONS)
    lats = ",".join(str(STATIONS[n][0]) for n in names)
    lons = ",".join(str(STATIONS[n][1]) for n in names)
    weights = {n: STATIONS[n][2] for n in names}
    have = dict(previous or {})

    ends, cursor = [], end_day
    while len(ends) < HOURLY_TEMP_CHUNKS_PER_RUN:
        ends.append(cursor)
        cursor = cursor - dt.timedelta(days=HOURLY_TEMP_CHUNK_DAYS)
        if cursor.strftime("%Y-%m-%dT00") < floor_key:
            break

    for i, e in enumerate(ends):
        start = e - dt.timedelta(days=HOURLY_TEMP_CHUNK_DAYS - 1)
        span_start = start.strftime("%Y-%m-%dT00")
        span_end = e.strftime("%Y-%m-%dT23")
        covered = sum(1 for k in have if span_start <= k <= span_end)
        # Expect only the part of the chunk AT OR ABOVE the retention
        # floor. The oldest chunk always straddles it - roughly a
        # quarter of its hours are discarded on write - so measured
        # against the whole chunk its coverage could never reach the
        # skip threshold, and it was re-fetched on every run forever
        # for a month of data already held.
        expect = sum(1 for h in range(HOURLY_TEMP_CHUNK_DAYS * 24)
                     if (start + dt.timedelta(hours=h)
                         ).strftime("%Y-%m-%dT%H") >= floor_key)
        expect = max(expect, 1)
        # Newest chunk always re-fetched (ERA5 revises); older chunks
        # skipped only once nearly whole, so gaps converge over runs
        # rather than becoming permanent.
        if i and covered >= expect * 0.98:
            continue
        try:
            payload = http_get(
                "https://archive-api.open-meteo.com/v1/archive", params={
                    "latitude": lats, "longitude": lons,
                    "start_date": start.isoformat(),
                    "end_date": e.isoformat(),
                    "hourly": "temperature_2m",
                    "timezone": "Europe/Dublin",
                    # 90, not 240. The first live run (7 Aug 2026)
                    # spent 256 s on a ReadTimeout and the identical
                    # retry then succeeded in three seconds, so a
                    # long timeout buys nothing here but wasted
                    # runner minutes - the retry budget is what
                    # actually absorbs the transient.
                }, timeout=90).json()
            got = weighted_hourly_temp(payload, names, weights)
            have.update(got)
            log(f"hourly: temp_ai archive {start.isoformat()}.."
                f"{e.isoformat()} {len(got)}h")
        except Exception as exc:
            log(f"hourly: temp_ai archive chunk to {e.isoformat()} "
                f"{exc.__class__.__name__} - gap left for next run")
        time.sleep(0.4)

    # ERA5 lags ~5 days; the forecast endpoint covers the tail. Archive
    # values are authoritative where both exist.
    try:
        tail = weighted_hourly_temp(http_get(
            "https://api.open-meteo.com/v1/forecast", params={
                "latitude": lats, "longitude": lons,
                "past_days": 10, "forecast_days": 1,
                "hourly": "temperature_2m",
                "timezone": "Europe/Dublin",
            }, timeout=120).json(), names, weights)
        added = 0
        for k, v in tail.items():
            if k not in have:
                have[k] = v
                added += 1
        log(f"hourly: temp_ai forecast tail +{added}h")
    except Exception as exc:
        log(f"hourly: temp_ai forecast tail unavailable "
            f"({exc.__class__.__name__}) - archive only")
    return have


def _naive_hour(key):
    """Hour key -> datetime, parsed as written. The keys are IRISH
    LOCAL CLOCK, so this is not a UTC instant and must never be
    treated as one; it exists only to give the keys a total order and
    a spacing, which is all the array encoding needs."""
    return dt.datetime.strptime(key, "%Y-%m-%dT%H")


def compact_hourly(series):
    """
    {name: {hour_key: value}} -> (t0, n, {name: [value|None] * n}).

    Position i is the key t0 + i hours, formatted the same way. On the
    spring transition the local clock skips an hour, which shows up
    here as one null a year in every series - the alternative, a
    stored key list, costs about 140 kB to avoid a gap that is already
    indistinguishable from a feed gap. The autumn fold repeats a local
    hour, but the source dict has already collapsed it to one entry,
    so offsets stay unique in both directions.
    """
    keys = set()
    for v in series.values():
        keys |= set(v or {})
    if not keys:
        return None, 0, {k: [] for k in series}
    t0 = min(keys)
    h0 = _naive_hour(t0)
    n = int((_naive_hour(max(keys)) - h0).total_seconds() // 3600) + 1
    out = {}
    for name, v in series.items():
        arr = [None] * n
        for k, val in (v or {}).items():
            i = int((_naive_hour(k) - h0).total_seconds() // 3600)
            if 0 <= i < n:
                arr[i] = val
        out[name] = arr
    return t0, n, out


def expand_hourly(doc):
    """
    Read a store document of ANY schema back into
    {name: {hour_key: value}}.

    Schema 1 and 2 wrote a dict per series; schema 3 writes flat
    arrays against `t0`. Both are accepted, so the run that first
    writes schema 3 still inherits the schema-2 file already in the
    repo instead of refilling 13 months from empty.
    """
    ser = (doc or {}).get("series") or {}
    if not ser:
        return {}
    sample = next(iter(ser.values()))
    if isinstance(sample, dict):
        return {k: dict(v or {}) for k, v in ser.items()}
    t0 = doc.get("t0")
    if not t0:
        log("hourly: array-form store without t0 - previous state "
            "discarded, refilling")
        return {}
    h0 = _naive_hour(t0)
    out = {}
    for name, arr in ser.items():
        d = {}
        for i, val in enumerate(arr or []):
            if val is None:
                continue
            d[(h0 + dt.timedelta(hours=i)).strftime("%Y-%m-%dT%H")] = val
        out[name] = d
    return out


def daily_ci_from_hourly(store, min_hours=20):
    """
    Daily mean grid carbon intensity from the hourly store.

    The daily EirGrid feed retains 50 days. The backfilled weeks are
    ten months old, so their carbon can only come from the hourly
    store's co2_ai series - and that store keeps 13 months, so its
    floor advances a day every day. Once a week is built and frozen
    its ef_electricity is stored and reused on restatement, so this is
    a one-time capture: after these weeks are in the record, the store
    rolling past them costs nothing.

    Returns {date: g/kWh}. Days with fewer than min_hours observations
    are dropped rather than averaged thin.
    """
    ser = (expand_hourly(store) or {}).get("co2_ai") or {}
    acc = {}
    for k, v in ser.items():
        acc.setdefault(k[:10], []).append(v)
    out = {d: round(sum(vs) / len(vs), 1)
           for d, vs in acc.items() if len(vs) >= min_hours}
    return out


def build_hourly_store(previous, feeds_now=None):
    """Walk back in ~28-day chunks until the window is covered, then
    keep only the most recent chunk (plus a 2-day revision re-fetch)
    on subsequent runs. Returns the store document or None."""
    prev = (previous or {})
    # NB: `prior`, not `prev_series` - there is a module-level
    # prev_series() helper for the weekly feeds and shadowing it here
    # would be a trap for the next edit.
    prior = expand_hourly(prev)
    series = {k: dict(prior.get(k) or {}) for k in HOURLY_SERIES}
    end = today_utc()
    floor = (end - dt.timedelta(days=30 * HOURLY_MONTHS)).strftime(
        "%Y-%m-%dT00")
    added = 0
    for name, (chart, region, areas) in HOURLY_SERIES.items():
        have = series[name]
        # chunk ends: newest first, walking back to the floor
        ends, cursor = [], end
        while len(ends) < HOURLY_CHUNKS_PER_RUN:
            ends.append(cursor)
            cursor = cursor - dt.timedelta(days=28)
            if cursor.strftime("%Y-%m-%dT00") < floor:
                break
        for e in ends:
            span_start = (e - dt.timedelta(days=27)).strftime("%Y-%m-%dT00")
            span_end = e.strftime("%Y-%m-%dT23")
            covered = sum(1 for k in have if span_start <= k <= span_end)
            # Newest chunk always re-fetched (2-day revision window).
            # Older chunks are skipped only once NEARLY whole: at the
            # original 600/672 threshold a chunk sitting at, say, 620
            # was never retried and its gap became permanent - which
            # is why carbon stalled at 93.6% on 7 Aug 2026. At 98% the
            # store converges over successive runs instead.
            if e != ends[0] and covered >= 660:
                continue
            got = {}
            for attempt in (1, 2):
                try:
                    got = fetch_hourly_chunk(chart, region, areas, e)
                    if got:
                        break
                except Exception as exc:
                    log(f"hourly: {name} chunk to {e.isoformat()} "
                        f"attempt {attempt} {exc.__class__.__name__}")
                if attempt == 1:
                    time.sleep(3)      # back off, then retry in-run
            if not got:
                log(f"hourly: {name} chunk to {e.isoformat()} "
                    f"EMPTY after 2 attempts - gap left for next run")
            # count what is NEW to the store, not what came back -
            # the newest chunk is deliberately re-fetched every run,
            # so len(got) double-counts an overlap the store already
            # holds and makes the run line meaningless as a measure
            # of progress.
            added += len(set(got) - set(have))
            have.update(got)
            time.sleep(0.4)            # throttle: ~56 chunk requests
                                       # per cold build trips limits
        before = len(prior.get(name) or {})
        series[name] = {k: v for k, v in sorted(have.items())
                        if k >= floor}
        after = len(series[name])
        # A series must never shrink except by the rolling floor.
        # 7 Aug 2026: demand_ai lost its oldest 600 hours in a single
        # run while the other series kept theirs - loud, not silent.
        if before and after < before - 24:
            log(f"hourly: WARNING {name} shrank {before}h -> {after}h "
                f"in one run - investigate before trusting the panel")
    # The grid series gate the store: with EirGrid down there is
    # nothing for a temperature series to be joined TO, and fetching
    # one would spend a minute of runner time on a document that is
    # not going to be written.
    if not any(len(series[k]) for k in HOURLY_SERIES):
        log("hourly: no data - store not written")
        return None

    # --- fifth series: island hourly temperature (Open-Meteo).
    prev_t = dict(prior.get("temp_ai") or {})
    try:
        t_all = fetch_hourly_temp(prev_t, floor, end)
    except Exception as exc:
        log(f"hourly: temp_ai unavailable ({exc.__class__.__name__}) - "
            "retaining previous")
        t_all = prev_t
    series["temp_ai"] = {k: v for k, v in sorted(t_all.items())
                         if k >= floor}
    # temp is fetched outside the EirGrid walk, so it was missing from
    # `added` and the run line read short (5,751 on 7 Aug 2026 while
    # temp alone had brought in 9,384 hours).
    added += len(set(series["temp_ai"]) - set(prev_t))
    if prev_t and len(series["temp_ai"]) < len(prev_t) - 24:
        log(f"hourly: WARNING temp_ai shrank {len(prev_t)}h -> "
            f"{len(series['temp_ai'])}h in one run - investigate "
            "before trusting the panel")

    # --- sixth series: SEMOpx day-ahead price, EUR/MWh.
    #
    # FILLS FORWARD ONLY. The SEMOpx report listing serves the recent
    # window; resolving an arbitrary historic trade day needs a filter
    # this pipeline has no evidence for, and guessing a parameter
    # against a live API is how the round-1 probe wasted a day. So the
    # series accumulates from the daily document already fetched, and
    # semopx_history_probe() asks the backfill question with logging
    # instead of assumptions.
    #
    # CONSEQUENCE, stated because it changes the plan: B.2.3 wants the
    # price in the tightest hour, and the tightest hour so far is
    # 5 Jan 2026. Until either the probe finds a way back or thirteen
    # months pass, B.2.3 can be computed on hours the store has priced,
    # not on the binding hour already found.
    price = dict(prev_series("semopx", "dam_hourly_eur_mwh"))
    price.update(((feeds_now or {}).get("semopx") or {})
                 .get("dam_hourly_eur_mwh") or {})
    series["price_ai"] = {k: v for k, v in sorted(price.items())
                          if k >= floor}

    counts = {k: len(v) for k, v in series.items()}
    spans = {k: (min(v), max(v)) for k, v in series.items() if v}
    ref = spans.get("demand_ai")
    # NEW hour-values, not "fetched". The EirGrid walk counted every
    # hour it received including re-fetched overlaps, while temp
    # counted only hours it did not already hold - two meanings in one
    # number, and the line read short on the day temp brought in 9,384
    # hours. Both sides now report new-to-the-store.
    log(f"hourly: {added} new hour-values this run; "
        + ", ".join(f"{k} {n}h" for k, n in counts.items())
        + (f"; span {ref[0]} .. {ref[1]}" if ref else ""))
    # completeness against the covered span (>=95% expected)
    # Completeness is judged PER SERIES against the reference span,
    # not on demand alone: the first store (7 Aug 2026) had demand at
    # 100% while carbon intensity reached only 86%, and a carbon
    # overlay drawn on that would have passed a demand-only gate.
    complete, heat_ready, price_ready, per_series = False, False, False, {}
    if ref:
        # Denominator is the INTENDED window (floor -> latest hour in
        # the store), not one series' own span. Using demand's span
        # let other series score 106.8% on 7 Aug 2026 when demand
        # shrank - a completeness figure above 100% is a bug signal,
        # so the denominator must not depend on the numerator.
        # ...and the window is the GRID window. temp_ai carries a
        # forecast tail that can run a day past the last EirGrid hour;
        # letting it set the denominator would depress every other
        # series' completeness for a reason that has nothing to do
        # with them.
        latest = max(mx for k, (_, mx) in spans.items()
                     if k in HOURLY_SERIES)
        h0 = dt.datetime.strptime(floor, "%Y-%m-%dT%H")
        h1 = dt.datetime.strptime(latest, "%Y-%m-%dT%H")
        expect = int((h1 - h0).total_seconds() // 3600) + 1
        for k, n in counts.items():
            sp = spans.get(k)
            log(f"hourly: {k} span "
                + (f"{sp[0]} .. {sp[1]}" if sp else "empty"))
            # Clamped at 100. temp_ai legitimately holds hours past
            # the grid window (whole local days, and a forecast tail),
            # so it scored 100.1% for a week - which reads as a bug
            # signal in a field whose whole job is to flag bugs. The
            # surplus is reported separately rather than inflating a
            # percentage that cannot exceed whole.
            pct = min(100.0, 100.0 * n / max(expect, 1))
            per_series[k] = round(pct, 1)
            extra = n - expect
            log(f"hourly: {k} completeness {pct:.1f}% of {expect}h "
                + ("- OK" if pct >= 95 else "- BELOW GATE")
                + (f" (+{extra}h beyond the grid window)"
                   if extra > 0 else ""))
        # the store is usable when the load/generation trio is whole;
        # carbon is an overlay and gates itself in the panel
        core = ("demand_ai", "wind_ai", "solar_ai")
        complete = all(per_series.get(k, 0) >= 95 for k in core)
        log(f"hourly: core trio {'complete' if complete else 'INCOMPLETE'}"
            f"; carbon overlay "
            f"{'available' if per_series.get('co2_ai', 0) >= 95 else 'withheld'}")
        # heat_ready is a SEPARATE gate, deliberately. The temperature
        # series is what the electrification computations need and the
        # existing panels do not; keeping it out of `complete` means
        # temp_ai filling over its first few runs can never withdraw
        # anything already shipping.
        heat_ready = complete and per_series.get("temp_ai", 0) >= 95
        # A third gate. price_ai starts empty and fills a day at a
        # time, so it must never be able to withdraw the heat layer or
        # the grid trio while it climbs.
        price_ready = complete and per_series.get("price_ai", 0) >= 95
        log(f"hourly: heat layer "
            f"{'ready' if heat_ready else 'not ready'} "
            f"(temp_ai {per_series.get('temp_ai', 0)}%)")
        log(f"hourly: price layer "
            f"{'ready' if price_ready else 'not ready'} "
            f"(price_ai {per_series.get('price_ai', 0)}% - fills forward, "
            f"see semopx_history_probe)")
    t0, n_hours, packed = compact_hourly(series)
    log(f"hourly: encoded {sum(len(v) for v in series.values())} values "
        f"as {len(packed)} arrays of {n_hours} from {t0}")
    return {"schema": HOURLY_SCHEMA,
            "t0": t0, "hours": n_hours,
            "heat_ready": heat_ready, "price_ready": price_ready,
            "generated": dt.datetime.now(dt.timezone.utc)
            .strftime("%Y-%m-%dT%H:%M:%SZ"),
            "months": HOURLY_MONTHS, "complete": complete,
            "completeness_pct": per_series,
            "encoding": ("series are flat arrays; position i is the "
                         "hour t0 + i, null where absent"),
            "basis": ("All-island 15-minute EirGrid series aggregated "
                      "to hourly means (>=3 of 4 quarters required). "
                      "Demand, wind and solar in MW; carbon intensity "
                      "in g CO2 per kWh. Source: EirGrid Smart Grid "
                      "Dashboard. temp_ai is population-weighted "
                      "island air temperature in degrees C, ERA5 via "
                      "Open-Meteo, weights as the daily HDD feed. "
                      "EVERY series is keyed on Irish local clock, "
                      "not UTC, so temperature and demand describe "
                      "the same hour year-round."),
            "series": packed}


def main():
    global PREVIOUS_FEEDS, PREVIOUS_DERIVED
    if DATA_PATH.exists():
        try:
            prev_doc = json.loads(DATA_PATH.read_text())
            PREVIOUS_FEEDS = prev_doc.get("feeds", {})
            PREVIOUS_DERIVED = prev_doc.get("derived", {})
        except Exception:
            log("warning - previous data.json unreadable, starting clean")

    feeds, failures = {}, []
    for name, fn in FEEDS.items():
        log(f"--- {name}")
        try:
            payload, status = fn()
            payload["status"] = status
            if name in FEED_FLAGS:
                payload["flags"] = FEED_FLAGS[name]
            payload["fetched_utc"] = dt.datetime.now(
                dt.timezone.utc).isoformat(timespec="seconds")
            feeds[name] = payload
            log(f"{name}: {status}, latest_day={payload.get('latest_day')}")
        except Exception as e:
            expected = (isinstance(e, NotImplementedError)
                        or name in EXPECTED_DOWN or name in SOFT_FEEDS)
            log(f"{name}: {'EXPECTED DOWN' if expected else 'FAILED'} - "
                f"{e.__class__.__name__}: {e}")
            if not expected:
                traceback.print_exc()
            prev = PREVIOUS_FEEDS.get(name, {})
            has_prev = bool(prev)
            prev["status"] = "stale"
            if name in EXPECTED_DOWN:
                prev["pending_note"] = EXPECTED_DOWN[name]
            if name in FEED_FLAGS:
                prev["flags"] = FEED_FLAGS[name]
            prev.setdefault("source", "previous run retained")
            feeds[name] = prev
            if not expected:
                # A failed feed with previous data in hand degrades to
                # stale-and-continue (transient 5xx like the CCNI 520 of
                # 27 Jul 2026 must not block the deploy); only a failure
                # with nothing to carry forward stays build-fatal.
                if has_prev:
                    log(f"{name}: previous data carried - degraded to "
                        f"stale, not build-fatal")
                else:
                    failures.append(name)

    gas = feeds.get("gni_ckan", {}).get("ndm_gwh") or {}
    hdd = feeds.get("hdd", {}).get("hdd_roi") or {}
    reg = space_heat_split(gas, hdd)
    derived = {"roi_space_heat_regression": reg} if reg else {}
    if reg:
        log("regression:", reg)
    apply_ie_fx(feeds)
    hero = derive_hero(feeds)
    if hero:
        derived["hero"] = hero
        log("hero:", {k: hero[k] for k in
                      ("week_ending", "heat_purchased_gwh",
                       "indigenous_share_pct", "bill_eur_m", "bill_gbp_m",
                       "emissions_kt_co2")})
    hg = derive_heat_gap(feeds)
    if hg:
        derived["heat_gap"] = hg
    # Carbon for the backfilled weeks. The daily feed keeps 50 days;
    # anything older comes from the hourly store, which is read here
    # BEFORE history so week_context can see it. Daily values win
    # where both exist - the store is the fallback, not the source.
    try:
        prev_store = (json.loads(HOURLY_PATH.read_text())
                      if HOURLY_PATH.exists() else {})
        hci = daily_ci_from_hourly(prev_store)
        if hci:
            eg = feeds.setdefault("eirgrid", {})
            ci = eg.setdefault("co2_intensity_g_per_kwh", {})
            added = sum(1 for d, v in hci.items() if ci.setdefault(d, v) is v)
            log(f"carbon: {len(hci)} daily means available from the "
                f"hourly store, {added} filled gaps in the daily feed "
                f"(span {min(hci)}..{max(hci)})")
    except Exception as exc:
        log(f"carbon: hourly-store fallback unavailable "
            f"({exc.__class__.__name__}) - backfilled weeks will use "
            "the anchor EF")
    derived["history"] = build_history(feeds)
    derived["history_schema"] = HISTORY_SCHEMA
    derived["anchor_epoch"] = ANCHOR_EPOCH
    # Two counters, deliberately. Weeks on record includes the weeks
    # reconstructed behind LIVE_FROM; weeks live counts only those
    # observed as they happened. The 52-live-weeks milestone is the
    # second number and is not reached by backfilling.
    _h = derived["history"] or []
    derived["weeks_on_record"] = len(_h)
    derived["weeks_live"] = sum(1 for e in _h if e.get("live"))
    derived["live_from"] = LIVE_FROM
    # Priced before compaction, while the history is still a list of
    # week objects - the columnar form is a wire format, not a
    # working one.
    try:
        hcs = derive_heat_cost_series(feeds)
        if hcs:
            derived["heat_cost_series"] = hcs
    except Exception as exc:
        log(f"heat cost: failed ({exc.__class__.__name__}) - "
            "the weekly tracker is unaffected")
    _flat = len(json.dumps(_h, separators=(",", ":")))
    derived["history"] = compact_history(_h)
    derived["history_encoding"] = HISTORY_ENCODING
    _cols = len(json.dumps(derived["history"], separators=(",", ":")))
    log(f"history: encoded {len(_h)} weeks columnar, "
        f"{_flat // 1024} kB -> {_cols // 1024} kB "
        f"({100 * _cols // max(_flat, 1)}%)")
    log(f"history: {derived['weeks_on_record']} weeks on record, "
        f"{derived['weeks_live']} live (from {LIVE_FROM}); "
        f"schema {HISTORY_SCHEMA}, anchor epoch {ANCHOR_EPOCH}")
    gw = sorted(((feeds.get("gni_live") or {}).get("ndm_gwh")
                 or {}))
    derived["gas_window"] = {"from": gw[0], "to": gw[-1],
                             "days": len(gw)} if gw else None
    # len() of the columnar block is its key count, not its week
    # count - it printed "3 complete weeks" for a 52-week record on
    # the first 4.24.0 run. Count the weeks, not the container.
    log(f"history: {derived['weeks_on_record']} complete weeks"
        + (f", gas_window {derived['gas_window']['days']}d"
           if derived["gas_window"] else ""))
    reg = derived.get("roi_space_heat_regression")
    cal = derive_gas_calibration(
        reg, (feeds.get("hdd") or {}).get("hdd_roi") or {})
    if cal:
        derived["gas_calibration"] = cal
        log("gas calibration: implied", cal["implied_annual_space_heat_gwh"],
            "vs anchor", cal["anchor_annual_space_heat_gwh"],
            "ratio", cal["ratio"],
            "(gate 0.90-1.10)" if cal["within_gate"] else
            "OUTSIDE gate 0.90-1.10 - disclosed")
    cool = derive_cool(feeds)
    if cool:
        derived["cool"] = cool
        log("cool: stranded summer share",
            cool["stranded_summer_pct"], "%")
        log("heat_gap: breakeven SPF vs oil - NI",
            hg["ni"]["breakeven_spf_vs_oil"], "ROI",
            hg["roi"]["breakeven_spf_vs_oil"])

    doc = {
        "pipeline_version": PIPELINE_VERSION,
        "built_utc": dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds"),
        "feeds": feeds,
        "derived": derived,
        "events": EVENTS,
        "geo": {**GEO, "percap": derive_geo_percap()},
        "why_heat": WHY_HEAT,
        "notes": ("Feed statuses - ok: fetched and current; lagging: fetched, "
                  "source publishes on a lag; stale: fetch failed, previous "
                  "values retained. Judgement figures are current Causeway "
                  "Energies estimates - challenge and input welcome at "
                  "contact@causewaygt.com"),
    }
    DATA_PATH.parent.mkdir(parents=True, exist_ok=True)
    DATA_PATH.write_text(json.dumps(doc, indent=1, sort_keys=True))

    # hourly store - separate file, separate schema, cannot corrupt
    # the weekly tracker if it fails
    try:
        prev_hourly = {}
        if HOURLY_PATH.exists():
            try:
                prev_hourly = json.loads(HOURLY_PATH.read_text())
            except Exception:
                log("hourly: previous store unreadable, rebuilding")
        try:
            semopx_history_probe()
        except Exception as exc:
            log(f"semopx_probe: failed ({exc.__class__.__name__})")
        try:
            semo_dispatch_probe()
        except Exception as exc:
            log(f"semo_probe: failed ({exc.__class__.__name__})")
        store = build_hourly_store(prev_hourly, feeds)
        # B.2.1 runs on the store we just wrote, log-only. Soft: a
        # failure here must never touch the weekly tracker.
        try:
            if store and store.get("heat_ready"):
                derived["tightest_hour"] = derive_tightest_hour(store)
            elif store:
                log("B.2.1 skipped - heat layer not ready "
                    f"(temp_ai {store.get('completeness_pct', {}).get('temp_ai')}%)")
        except Exception as exc:
            log(f"B.2.1 failed ({exc.__class__.__name__}) - log-only, "
                "weekly tracker unaffected")
        if store:
            HOURLY_PATH.write_text(json.dumps(store, separators=(",", ":")))
            log(f"wrote {HOURLY_PATH} "
                f"({HOURLY_PATH.stat().st_size // 1024} kB)")
    except Exception as exc:
        log(f"hourly: store step failed ({exc.__class__.__name__}: "
            f"{exc}) - weekly output unaffected")
    log(f"wrote {DATA_PATH} ({DATA_PATH.stat().st_size/1024:.0f} kB)")

    if failures:
        log("hard failures:", failures)
        if NTFY_TOPIC:
            try:
                requests.post(f"https://ntfy.sh/{NTFY_TOPIC}",
                              data=f"ioi-heatsplit build: failed {failures}",
                              timeout=15)
            except Exception:
                pass
        sys.exit(1)


if __name__ == "__main__":
    main()
